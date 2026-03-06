from datetime import date, datetime, time, timedelta, timezone, timedelta
import json
from typing import Any
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import redirect, render,get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse, reverse_lazy
from django.views import View
from django.views.generic import ListView, DetailView,CreateView, UpdateView, TemplateView
from django.contrib import messages
import pandas as pd
from userauths.models import *
from .models import *
from .forms import *
from django.contrib.auth import logout
from django.db.models import Count
from django.db.models import Sum, F, Value
import calendar
from django.db.models.functions import ExtractMonth
from django.db.models.functions import Coalesce
from django.db import transaction
# Create your views here.
from .forms import DateForm
from django.utils.decorators import method_decorator
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from userauths.mixins import CustomPermissionRequiredMixin
from django.views.decorators.http import require_POST
from calendar import monthrange
from calendar import SUNDAY
from django.utils.timezone import now

from django.core.mail import send_mail
from userauths.utils import search_vehicules

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from xhtml2pdf import pisa
from django.conf import settings
import os


def permission_denied_view(request, exception):
    return render(request, 'no_acces.html',status=403)

def custom_404_view(request, exception):
    return render(request, 'no_acces.html',status=404)

def temp_arr(request):
    # return render(request, 'perfect/dashboard.html')
    return render(request, 'perfect/tmp_arr.html')

def base(request):
    # return render(request, 'perfect/dashboard.html')
    return render(request, 'perfect/base.html')

class ResumeView(TemplateView):
    template_name = 'pbent/resume_to_day.html'

class Bilanday(TemplateView):
    form_class = DateForm
    template_name = 'perfect/bilan_day.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        dates = today
        label = [calendar.month_name[month][:1] for month in range(1, 13)]
        mois_fr = ('', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre')
        month_choices = [(i, mois_fr[i]) for i in range(1, 13)]
        form = self.form_class(self.request.GET)
        # Catégories dynamiques (comme AllVehiculeView) : nb véhicules avec car_statut=True
        catego_vehi = CategoVehi.objects.all().annotate(
            vehicule_count=Count('catego_vehicule', filter=Q(catego_vehicule__car_statut=True))
        ).order_by('id')
        date_debut = date_fin = today
        if form.is_valid() and form.cleaned_data.get('date_debut') and form.cleaned_data.get('date_fin'):
            date_debut = form.cleaned_data['date_debut']
            date_fin = form.cleaned_data['date_fin']
            date_filter = {'date_saisie__range': [date_debut, date_fin]}
        else:
            date_filter = {'date_saisie': today}
        recettes_par_categorie = []
        total_recettes = 0
        for cat in catego_vehi:
            qs = Recette.objects.filter(vehicule__category=cat, **date_filter)
            total_cat = qs.aggregate(somme=Sum('montant'))['somme'] or 0
            total_recettes += total_cat
            recettes_par_categorie.append({
                'categorie': cat,
                'total': total_cat,
                'list': qs,
            })
        list_recettes = Recette.objects.filter(**date_filter)

        # Compatibilité template : 1ère et 2ème catégorie (si existantes)
        total_recettes_vtc = recettes_par_categorie[0]['total'] if len(recettes_par_categorie) > 0 else 0
        list_recettes_vtc = recettes_par_categorie[0]['list'] if len(recettes_par_categorie) > 0 else Recette.objects.none()
        total_recettes_taxi = recettes_par_categorie[1]['total'] if len(recettes_par_categorie) > 1 else 0
        list_recettes_taxi = recettes_par_categorie[1]['list'] if len(recettes_par_categorie) > 1 else Recette.objects.none()

        # ---- Pièces ----
        total_piece = Piece.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_piece = Piece.objects.filter(**date_filter)
        total_piec_echange = PiecEchange.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_piec_echange = PiecEchange.objects.filter(**date_filter)

        # ---- Charges ----
        total_charg_fix = ChargeFixe.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_charg_fix = ChargeFixe.objects.filter(**date_filter)
        total_charg_var = ChargeVariable.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_charg_var = ChargeVariable.objects.filter(**date_filter)
        total_charg_admin = ChargeVariable.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_charg_admin = ChargeVariable.objects.filter(**date_filter)
        total_charg = total_charg_fix + total_charg_var

        # ---- Réparations, assurances, visites, etc. ----
        total_reparation = Reparation.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_reparation = Reparation.objects.filter(**date_filter)
        total_assurances = Assurance.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_assurance = Assurance.objects.filter(**date_filter)
        total_visite = VisiteTechnique.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_visite = VisiteTechnique.objects.filter(**date_filter)
        total_entretien = Entretien.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_entretien = Entretien.objects.filter(**date_filter)
        total_patente = Patente.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_patente = Patente.objects.filter(**date_filter)
        total_vignette = Vignette.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_vignette = Vignette.objects.filter(**date_filter)
        total_encaissement = Encaissement.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_encaissement = Encaissement.objects.filter(**date_filter)
        total_decaissement = Decaissement.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
        list_decaissement = Decaissement.objects.filter(**date_filter)

        # ---- Marges et taux ----
        marge_contribution = total_recettes - total_charg_var
        taux_marge = (marge_contribution * 100 / total_recettes) if total_recettes else 0
        taux_marge_format = '{:.2f}'.format(taux_marge)
        marge_brute = total_recettes - total_charg
        marge_brute_format = '{:,}'.format(marge_brute).replace(',', ' ')

        # ---- Comptabilité : Débit / Crédit / Solde ----
        # Produits (recettes) → Crédit ; Charges et mouvements sortants → Débit ; Encaissements → Débit ; Décaisseements → Crédit
        total_debit = (
            total_charg_fix + total_charg_var + total_reparation + total_assurances
            + total_visite + total_entretien + total_patente + total_vignette
            + total_piece + total_piec_echange + total_encaissement
        )
        total_credit = total_recettes + total_decaissement
        solde_periode = total_credit - total_debit
        period_label = (
            f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
            if date_debut != date_fin
            else f"Jour du {date_debut.strftime('%d/%m/%Y')}"
        )

        def _fmt(n):
            return '{:,}'.format(n).replace(',', ' ') if n else '0'

        # Lignes du tableau bilan (Débit / Crédit / Solde par ligne)
        lignes_bilan = [
            {'section': '70', 'intitule': "Recettes d'exploitation", 'debit': 0, 'credit': total_recettes, 'solde': total_recettes},
            {'section': '61', 'intitule': 'Charges fixes', 'debit': total_charg_fix, 'credit': 0, 'solde': -total_charg_fix},
            {'section': '62', 'intitule': 'Charges variables', 'debit': total_charg_var, 'credit': 0, 'solde': -total_charg_var},
            {'section': '63', 'intitule': 'Réparations', 'debit': total_reparation, 'credit': 0, 'solde': -total_reparation},
            {'section': '64', 'intitule': 'Assurances', 'debit': total_assurances, 'credit': 0, 'solde': -total_assurances},
            {'section': '65', 'intitule': 'Visites techniques', 'debit': total_visite, 'credit': 0, 'solde': -total_visite},
            {'section': '66', 'intitule': 'Entretiens', 'debit': total_entretien, 'credit': 0, 'solde': -total_entretien},
            {'section': '67', 'intitule': 'Patentes', 'debit': total_patente, 'credit': 0, 'solde': -total_patente},
            {'section': '68', 'intitule': 'Vignettes', 'debit': total_vignette, 'credit': 0, 'solde': -total_vignette},
            {'section': '69', 'intitule': 'Pièces (réparations)', 'debit': total_piece, 'credit': 0, 'solde': -total_piece},
            {'section': '69', 'intitule': 'Pièces échangées', 'debit': total_piec_echange, 'credit': 0, 'solde': -total_piec_echange},
            {'section': '53', 'intitule': 'Encaissements (caisse)', 'debit': total_encaissement, 'credit': 0, 'solde': total_encaissement},
            {'section': '53', 'intitule': 'Décaisseements (caisse)', 'debit': 0, 'credit': total_decaissement, 'solde': -total_decaissement},
        ]
        for L in lignes_bilan:
            L['debit_fmt'] = _fmt(L['debit']) if L['debit'] else ''
            L['credit_fmt'] = _fmt(L['credit']) if L['credit'] else ''
            L['solde_fmt'] = _fmt(abs(L['solde']))
            L['solde_créditeur'] = L['solde'] > 0
            L['solde_débiteur'] = L['solde'] < 0

        context.update({
            'date_debut': date_debut,
            'date_fin': date_fin,
            'period_label': period_label,
            'lignes_bilan': lignes_bilan,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'solde_periode': solde_periode,
            'total_debit_fmt': _fmt(total_debit),
            'total_credit_fmt': _fmt(total_credit),
            'solde_periode_fmt': _fmt(abs(solde_periode)),
            'solde_periode_crediteur': solde_periode >= 0,
            'catego_vehi': catego_vehi,
            'recettes_par_categorie': recettes_par_categorie,
            'total_recettes_taxi': total_recettes_taxi,
            'list_recettes_taxi': list_recettes_taxi,
            'total_recettes_vtc': total_recettes_vtc,
            'list_recettes_vtc': list_recettes_vtc,
            'total_recettes': total_recettes,
            'list_recettes': list_recettes,
            'total_piece': total_piece,
            'list_piece': list_piece,
            'total_piec_echange': total_piec_echange,
            'list_piec_echange': list_piec_echange,
            'total_charg_fix': total_charg_fix,
            'list_charg_fix': list_charg_fix,
            'total_charg_var': total_charg_var,
            'list_charg_var': list_charg_var,
            'total_charg_admin': total_charg_admin,
            'list_charg_admin': list_charg_admin,
            'total_reparation': total_reparation,
            'list_reparation': list_reparation,
            'total_assurances': total_assurances,
            'list_assurance': list_assurance,
            'total_visite': total_visite,
            'list_visite': list_visite,
            'total_entretien': total_entretien,
            'list_entretien': list_entretien,
            'total_patente': total_patente,
            'list_patente': list_patente,
            'total_vignette': total_vignette,
            'list_vignette': list_vignette,
            'total_encaissement': total_encaissement,
            'list_encaissement': list_encaissement,
            'total_decaissement': total_decaissement,
            'list_decaissement': list_decaissement,
            'marge_contribution': marge_contribution,
            'taux_marge_format': taux_marge_format,
            'marge_brute': marge_brute,
            'marge_brute_format': marge_brute_format,
            'labels': label,
            'month_choices': month_choices,
            'years': range(today.year - 2, today.year + 1),
            'form': form,
            'dates': dates,
        })
        return context

def _get_bilanday_data_for_period(date_debut, date_fin):
    """Calcule les données du bilan pour une période (ou un jour). Utilisé par Bilanday et ExportBilandayExcelView."""
    today = date.today()
    if date_debut != date_fin:
        date_filter = {'date_saisie__range': [date_debut, date_fin]}
    else:
        date_filter = {'date_saisie': date_debut}

    total_recettes = Recette.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_charg_fix = ChargeFixe.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_charg_var = ChargeVariable.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_reparation = Reparation.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_assurances = Assurance.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_visite = VisiteTechnique.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_entretien = Entretien.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_patente = Patente.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_vignette = Vignette.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_piece = Piece.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_piec_echange = PiecEchange.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_encaissement = Encaissement.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0
    total_decaissement = Decaissement.objects.filter(**date_filter).aggregate(somme=Sum('montant'))['somme'] or 0

    total_debit = (
        total_charg_fix + total_charg_var + total_reparation + total_assurances
        + total_visite + total_entretien + total_patente + total_vignette
        + total_piece + total_piec_echange + total_encaissement
    )
    total_credit = total_recettes + total_decaissement
    solde_periode = total_credit - total_debit
    period_label = (
        f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        if date_debut != date_fin
        else f"Jour du {date_debut.strftime('%d/%m/%Y')}"
    )

    def _fmt(n):
        return '{:,}'.format(n).replace(',', ' ') if n else '0'

    lignes_bilan = [
        {'section': '70', 'intitule': "Recettes d'exploitation", 'debit': 0, 'credit': total_recettes, 'solde': total_recettes},
        {'section': '61', 'intitule': 'Charges fixes', 'debit': total_charg_fix, 'credit': 0, 'solde': -total_charg_fix},
        {'section': '62', 'intitule': 'Charges variables', 'debit': total_charg_var, 'credit': 0, 'solde': -total_charg_var},
        {'section': '63', 'intitule': 'Réparations', 'debit': total_reparation, 'credit': 0, 'solde': -total_reparation},
        {'section': '64', 'intitule': 'Assurances', 'debit': total_assurances, 'credit': 0, 'solde': -total_assurances},
        {'section': '65', 'intitule': 'Visites techniques', 'debit': total_visite, 'credit': 0, 'solde': -total_visite},
        {'section': '66', 'intitule': 'Entretiens', 'debit': total_entretien, 'credit': 0, 'solde': -total_entretien},
        {'section': '67', 'intitule': 'Patentes', 'debit': total_patente, 'credit': 0, 'solde': -total_patente},
        {'section': '68', 'intitule': 'Vignettes', 'debit': total_vignette, 'credit': 0, 'solde': -total_vignette},
        {'section': '69', 'intitule': 'Pièces (réparations)', 'debit': total_piece, 'credit': 0, 'solde': -total_piece},
        {'section': '69', 'intitule': 'Pièces échangées', 'debit': total_piec_echange, 'credit': 0, 'solde': -total_piec_echange},
        {'section': '53', 'intitule': 'Encaissements (caisse)', 'debit': total_encaissement, 'credit': 0, 'solde': total_encaissement},
        {'section': '53', 'intitule': 'Décaisseements (caisse)', 'debit': 0, 'credit': total_decaissement, 'solde': -total_decaissement},
    ]
    for L in lignes_bilan:
        L['debit_fmt'] = _fmt(L['debit']) if L['debit'] else ''
        L['credit_fmt'] = _fmt(L['credit']) if L['credit'] else ''
        L['solde_fmt'] = _fmt(abs(L['solde']))
        L['solde_créditeur'] = L['solde'] > 0
        L['solde_débiteur'] = L['solde'] < 0

    return {
        'period_label': period_label,
        'lignes_bilan': lignes_bilan,
        'total_debit_fmt': _fmt(total_debit),
        'total_credit_fmt': _fmt(total_credit),
        'solde_periode_fmt': _fmt(abs(solde_periode)),
        'solde_periode_crediteur': solde_periode >= 0,
    }

class ExportBilandayExcelView(LoginRequiredMixin, View):
    """Export du bilan journalier / période au format Excel. Par défaut : jour en cours ; si date_debut et date_fin en GET : période du filtre."""
    login_url = 'login'

    def get(self, request, *args, **kwargs):
        today = date.today()
        date_debut = date_fin = today
        form = DateForm(request.GET)
        if form.is_valid() and form.cleaned_data.get('date_debut') and form.cleaned_data.get('date_fin'):
            date_debut = form.cleaned_data['date_debut']
            date_fin = form.cleaned_data['date_fin']

        data = _get_bilanday_data_for_period(date_debut, date_fin)
        period_label = data['period_label']
        lignes_bilan = data['lignes_bilan']
        total_debit_fmt = data['total_debit_fmt']
        total_credit_fmt = data['total_credit_fmt']
        solde_periode_fmt = data['solde_periode_fmt']
        solde_periode_crediteur = data['solde_periode_crediteur']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bilan"

        header_font = Font(bold=True, color="035C9F")
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        ws.append(["Bilan — " + period_label])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.cell(row=1, column=1).font = header_font
        ws.append([])
        ws.append(["Section", "Intitulé de section", "Débit", "Crédit", "Soldes", "Exercice précédent"])
        for col in range(1, 7):
            ws.cell(row=3, column=col).font = Font(bold=True)
            ws.cell(row=3, column=col).fill = header_fill

        row = 4
        for L in lignes_bilan:
            solde_str = ''
            if L['solde_créditeur']:
                solde_str = 'C-' + L['solde_fmt']
            elif L['solde_débiteur']:
                solde_str = 'D-' + L['solde_fmt']
            ws.append([L['section'], L['intitule'], L['debit_fmt'], L['credit_fmt'], solde_str, ''])
            row += 1

        ws.append([])
        row += 1
        ws.append(["Total", "", total_debit_fmt, total_credit_fmt,
                   ('C-' if solde_periode_crediteur else 'D-') + solde_periode_fmt, ''])
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1
        ws.append(["A reporter (résultat période)", "", total_debit_fmt, total_credit_fmt,
                   ('C-' if solde_periode_crediteur else 'D-') + solde_periode_fmt, ''])
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = Font(bold=True)

        for col, width in enumerate([12, 35, 18, 18, 18, 18], 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Bilan-{date_debut.strftime('%d-%m-%Y')}"
        if date_debut != date_fin:
            filename += f"_au_{date_fin.strftime('%d-%m-%Y')}"
        filename += ".xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class TableaustopView(CustomPermissionRequiredMixin,TemplateView):
    model = Vehicule
    permission_url = 'temps'
    template_name = "perfect/temp_arret.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        month = self.request.GET.get('month', timezone.now().month)
        year = self.request.GET.get('year', timezone.now().year)
        month = int(month)
        year = int(year)
        days_in_month = monthrange(year, month)[1]
        mois_fr = ('', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre')
        month_name = mois_fr[month] if 1 <= month <= 12 else datetime(year, month, 1).strftime("%B")
        month_choices = [(i, mois_fr[i]) for i in range(1, 13)]
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant, car_statut=True)
                else:
                    vehicules = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.filter(car_statut=True)

        # Filtre par catégorie si sélectionnée
        selected_categorie_id = self.request.GET.get('categorie', '').strip()
        if selected_categorie_id:
            try:
                selected_categorie_id = int(selected_categorie_id)
                vehicules = vehicules.filter(category_id=selected_categorie_id)
            except (ValueError, TypeError):
                selected_categorie_id = None
        else:
            selected_categorie_id = None
        categories_list = CategoVehi.objects.all().order_by('category')

        context['current_date'] = date.today()
        total_actions_sum = total_cost_parts_sum = total_income_sum = total_piece_sum = total_visit_sum = total_panne_sum = total_accident_sum = total_autrarret_sum = total_visitechique_sum = total_entretien_sum = total_repairs_by_motifs = total_motif_arrets = 0
        # Calculer les totaux par jour
        daily_totals = [0] * days_in_month
        for day in range(1, days_in_month + 1):
            for model in [Reparation, VisiteTechnique, Entretien, Autrarret]:
                count = model.objects.filter(
                    date_saisie__day=day,
                    date_saisie__month=month,
                    date_saisie__year=year
                )
                if user.user_type == "4":
                    try:
                        gerant = user.gerants.get()
                        categories_gerant = gerant.gerant_voiture.all()
                        if categories_gerant.exists():
                            count = count.filter(vehicule__category__in=categories_gerant)
                        else:
                            count = count.none()
                    except Gerant.DoesNotExist:
                        count = count.none()
                if selected_categorie_id:
                    count = count.filter(vehicule__category_id=selected_categorie_id)
                daily_totals[day - 1] += count.count()
        
        vehicule_data = []
        for vehicule in vehicules:
            total_actions = (
                Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count() +
                VisiteTechnique.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count() +
                Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count() +
                Reparation.objects.filter(vehicule=vehicule,motif="Visite", date_saisie__month=month, date_saisie__year=year).count() +
                Reparation.objects.filter(vehicule=vehicule,motif="Panne", date_saisie__month=month, date_saisie__year=year).count() +
                Reparation.objects.filter(vehicule=vehicule,motif="Accident", date_saisie__month=month, date_saisie__year=year).count()
            )
            total_cost_parts = Piece.objects.filter(
                reparation__vehicule=vehicule, date_saisie__month=month, date_saisie__year=year
            ).aggregate(total_cost=models.Sum('montant'))['total_cost'] or 0

            total_income = Recette.objects.filter(
                vehicule=vehicule, date_saisie__month=month, date_saisie__year=year
            ).aggregate(total_income=models.Sum('montant'))['total_income'] or 0

            part_details_queryset = Piece.objects.filter(
                reparation__vehicule=vehicule, date_saisie__month=month, date_saisie__year=year
            ).values('libelle').annotate(count=models.Count('libelle'), total_price=models.Sum('montant'))
            
            all_piece = Piece.objects.filter(reparation__vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            all_visitechnique = VisiteTechnique.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            all_entretien = Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()

            part_details = "; ".join(
                f"{part['libelle']} ({part['count']}) {part['total_price']}" for part in part_details_queryset
            )

            daily_actions = [0] * days_in_month
            for day in range(1, days_in_month + 1):
                for model in [Reparation, VisiteTechnique, Entretien, Autrarret]:
                    count = model.objects.filter(
                        vehicule=vehicule, date_saisie__day=day, date_saisie__month=month, date_saisie__year=year
                    ).count()
                    daily_actions[day - 1] += count
            
            repairs_by_motif = {
                'P-vis': Reparation.objects.filter(vehicule=vehicule, motif="Visite", date_saisie__month=month, date_saisie__year=year).count(),
                'pan': Reparation.objects.filter(vehicule=vehicule, motif="Panne", date_saisie__month=month, date_saisie__year=year).count(),
                'acc': Reparation.objects.filter(vehicule=vehicule, motif="Accident", date_saisie__month=month, date_saisie__year=year).count(),
            }
            total_repairs_by_motif = (
                Reparation.objects.filter(vehicule=vehicule, motif="Visite", date_saisie__month=month, date_saisie__year=year).count()+
                Reparation.objects.filter(vehicule=vehicule, motif="Panne", date_saisie__month=month, date_saisie__year=year).count()+
                Reparation.objects.filter(vehicule=vehicule, motif="Accident", date_saisie__month=month, date_saisie__year=year).count()
            )
            motif_arret = {
                'vis': VisiteTechnique.objects.filter(vehicule=vehicule,date_saisie__month=month, date_saisie__year=year).count(),
                'ent': Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count(),
                'aut': Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count(),
            }
            total_motif_arret = (
                VisiteTechnique.objects.filter(vehicule=vehicule,date_saisie__month=month, date_saisie__year=year).count()+
                Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()+
                Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            )
            total_repairs_by_motifs += total_repairs_by_motif
            total_motif_arrets += total_motif_arret 
            
            all_rep_visit = Reparation.objects.filter(vehicule=vehicule, motif="Visite", date_saisie__month=month, date_saisie__year=year).count()
            all_rep_panne = Reparation.objects.filter(vehicule=vehicule, motif="Panne", date_saisie__month=month, date_saisie__year=year).count()
            all_rep_accident = Reparation.objects.filter(vehicule=vehicule, motif="Accident", date_saisie__month=month, date_saisie__year=year).count()
            all_autre_arret = Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            
            vehicule_data.append({
                'immatriculation': vehicule.immatriculation,
                'marque': vehicule.marque,
                'category': vehicule.category.category,
                'total_actions': total_actions,
                'total_cost_parts': total_cost_parts,
                'total_income': total_income,
                'part_details': part_details,
                'daily_actions': daily_actions,
                'repairs_by_motif': repairs_by_motif,
                'motif_arret': motif_arret,
            })
            total_actions_sum += total_actions
            total_cost_parts_sum += total_cost_parts
            total_income_sum += total_income
            total_piece_sum += all_piece
            total_visitechique_sum += all_visitechnique
            total_entretien_sum += all_entretien
            total_visit_sum += all_rep_visit
            total_panne_sum += all_rep_panne
            total_accident_sum += all_rep_accident
            total_autrarret_sum += all_autre_arret

        context['vehicule_data'] = vehicule_data
        context['categories_list'] = categories_list
        context['selected_categorie_id'] = selected_categorie_id
        
        context['total_repairs_by_motifs'] = total_repairs_by_motifs
        context['total_motif_arrets'] = total_motif_arrets
        
        context['total_visit_sum'] = total_visit_sum
        context['total_panne_sum'] = total_panne_sum
        context['total_accident_sum'] = total_accident_sum
        context['total_autrarret_sum'] = total_autrarret_sum
        
        context['total_visitechique_sum'] = total_visitechique_sum
        context['total_entretien_sum'] = total_entretien_sum
        
        context['total_actions_sum'] = total_actions_sum
        context['total_cost_parts_sum'] = total_cost_parts_sum
        context['total_income_sum'] = total_income_sum
        context['total_piece_sum'] = total_piece_sum
        
        context['days_in_month'] = range(1, days_in_month + 1)
        context['month_name'] = month_name
        context['month'] = month
        context['year'] = year
        context['years'] = range(timezone.now().year - 4, timezone.now().year + 1)
        context['month_range'] = range(1, 13)
        context['month_choices'] = month_choices
        context['days_in_month_plus_two'] = days_in_month + 2
        context['daily_totals'] = daily_totals
        return context

class ExportTempsArretExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        month = request.GET.get('month', timezone.now().month)
        year = request.GET.get('year', timezone.now().year)
        month = int(month)
        year = int(year)
        days_in_month = monthrange(year, month)[1]
        month_name = datetime(year, month, 1).strftime("%B")
        
        user = request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        selected_categorie_id = request.GET.get('categorie', '').strip()
        if selected_categorie_id:
            try:
                selected_categorie_id = int(selected_categorie_id)
                vehicules = vehicules.filter(category_id=selected_categorie_id)
            except (ValueError, TypeError):
                selected_categorie_id = None
        else:
            selected_categorie_id = None

        total_actions_sum = total_cost_parts_sum = total_income_sum = total_piece_sum = 0
        total_visit_sum = total_panne_sum = total_accident_sum = total_autrarret_sum = 0
        total_visitechique_sum = total_entretien_sum = total_repairs_by_motifs = total_motif_arrets = 0

        vehicule_data = []
        daily_totals = [0] * days_in_month

        for vehicule in vehicules:
            total_actions = (
                Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count() +
                VisiteTechnique.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count() +
                Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count() +
                Reparation.objects.filter(vehicule=vehicule, motif="Visite", date_saisie__month=month, date_saisie__year=year).count() +
                Reparation.objects.filter(vehicule=vehicule, motif="Panne", date_saisie__month=month, date_saisie__year=year).count() +
                Reparation.objects.filter(vehicule=vehicule, motif="Accident", date_saisie__month=month, date_saisie__year=year).count()
            )
            total_cost_parts = Piece.objects.filter(
                reparation__vehicule=vehicule, date_saisie__month=month, date_saisie__year=year
            ).aggregate(total_cost=Sum('montant'))['total_cost'] or 0

            total_income = Recette.objects.filter(
                vehicule=vehicule, date_saisie__month=month, date_saisie__year=year
            ).aggregate(total_income=Sum('montant'))['total_income'] or 0

            part_details_queryset = Piece.objects.filter(
                reparation__vehicule=vehicule, date_saisie__month=month, date_saisie__year=year
            ).values('libelle').annotate(count=models.Count('libelle'), total_price=models.Sum('montant'))
            
            all_piece = Piece.objects.filter(reparation__vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            all_visitechnique = VisiteTechnique.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            all_entretien = Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()

            part_details = "; ".join(
                f"{part['libelle']} ({part['count']}) {part['total_price']}" for part in part_details_queryset
            )

            daily_actions = [0] * days_in_month
            for day in range(1, days_in_month + 1):
                for model in [Reparation, VisiteTechnique, Entretien, Autrarret]:
                    count = model.objects.filter(
                        vehicule=vehicule, date_saisie__day=day, date_saisie__month=month, date_saisie__year=year
                    ).count()
                    daily_actions[day - 1] += count
                    daily_totals[day - 1] += count
            
            repairs_by_motif = {
                'P-vis': Reparation.objects.filter(vehicule=vehicule, motif="Visite", date_saisie__month=month, date_saisie__year=year).count(),
                'pan': Reparation.objects.filter(vehicule=vehicule, motif="Panne", date_saisie__month=month, date_saisie__year=year).count(),
                'acc': Reparation.objects.filter(vehicule=vehicule, motif="Accident", date_saisie__month=month, date_saisie__year=year).count(),
            }
            total_repairs_by_motif = (
                Reparation.objects.filter(vehicule=vehicule, motif="Visite", date_saisie__month=month, date_saisie__year=year).count()+
                Reparation.objects.filter(vehicule=vehicule, motif="Panne", date_saisie__month=month, date_saisie__year=year).count()+
                Reparation.objects.filter(vehicule=vehicule, motif="Accident", date_saisie__month=month, date_saisie__year=year).count()
            )
            motif_arret = {
                'vis': VisiteTechnique.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count(),
                'ent': Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count(),
                'aut': Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count(),
            }
            total_motif_arret = (
                VisiteTechnique.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()+
                Entretien.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()+
                Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            )
            
            all_rep_visit = Reparation.objects.filter(vehicule=vehicule, motif="Visite", date_saisie__month=month, date_saisie__year=year).count()
            all_rep_panne = Reparation.objects.filter(vehicule=vehicule, motif="Panne", date_saisie__month=month, date_saisie__year=year).count()
            all_rep_accident = Reparation.objects.filter(vehicule=vehicule, motif="Accident", date_saisie__month=month, date_saisie__year=year).count()
            all_autre_arret = Autrarret.objects.filter(vehicule=vehicule, date_saisie__month=month, date_saisie__year=year).count()
            
            vehicule_data.append({
                'immatriculation': vehicule.immatriculation,
                'marque': vehicule.marque,
                'total_actions': total_actions,
                'total_cost_parts': total_cost_parts,
                'total_income': total_income,
                'part_details': part_details,
                'daily_actions': daily_actions,
                'repairs_by_motif': repairs_by_motif,
                'motif_arret': motif_arret,
            })
            total_actions_sum += total_actions
            total_cost_parts_sum += total_cost_parts
            total_income_sum += total_income
            total_piece_sum += all_piece
            total_visitechique_sum += all_visitechnique
            total_entretien_sum += all_entretien
            total_visit_sum += all_rep_visit
            total_panne_sum += all_rep_panne
            total_accident_sum += all_rep_accident
            total_autrarret_sum += all_autre_arret
            total_repairs_by_motifs += total_repairs_by_motif
            total_motif_arrets += total_motif_arret
        
        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Temps Arret {month_name} {year}"
        
        # En-tête principal
        ws.merge_cells(f'A1:{get_column_letter(days_in_month + 8)}1')
        header_cell = ws.cell(row=1, column=1)
        header_cell.value = f"TEMPS D'ARRET - {month_name.upper()} {year}"
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center')
        
        # Ligne vide
        ws.append([])
        
        # Première ligne d'en-têtes
        headers_row1 = ['Immat', 'Marque']
        headers_row1.extend([''] * days_in_month)  # Espace pour les jours
        headers_row1.extend(['T arrêt', 'T Pièce', 'Recette', 'Pièce', 'Motif rép', 'Motif arrêt'])
        ws.append(headers_row1)
        
        # Fusionner les cellules pour le mois dans la première ligne
        if days_in_month > 0:
            start_col = get_column_letter(3)
            end_col = get_column_letter(2 + days_in_month)
            ws.merge_cells(f'{start_col}3:{end_col}3')
            month_cell = ws.cell(row=3, column=3)
            month_cell.value = f'{month_name} - {year}'
            month_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Deuxième ligne d'en-têtes (jours)
        headers_row2 = ['', '']
        headers_row2.extend([f'Jour {day}' for day in range(1, days_in_month + 1)])
        headers_row2.extend(['', '', '', '', '', ''])
        ws.append(headers_row2)
        
        # Style des en-têtes
        for row_num in [3, 4]:
            for col_num in range(1, len(headers_row1) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données des véhicules
        for detail in vehicule_data:
            row = [
                detail['immatriculation'],
                detail['marque'],
            ]
            row.extend(detail['daily_actions'])
            row.extend([
                detail['total_actions'],
                detail['total_cost_parts'],
                detail['total_income'],
                detail['part_details'],
                f"P-vis:{detail['repairs_by_motif']['P-vis']}, pan:{detail['repairs_by_motif']['pan']}, acc:{detail['repairs_by_motif']['acc']}",
                f"vis:{detail['motif_arret']['vis']}, ent:{detail['motif_arret']['ent']}, aut:{detail['motif_arret']['aut']}"
            ])
            ws.append(row)
        
        # Ligne de totaux par jour dans le footer
        footer_row = ['TOTAL PAR JOUR', '']
        footer_row.extend(daily_totals)
        footer_row.extend([
            total_actions_sum,
            total_cost_parts_sum,
            total_income_sum,
            total_piece_sum,
            total_repairs_by_motifs,
            total_motif_arrets
        ])
        ws.append(footer_row)
        
        # Style de la ligne de totaux
        for col_num in range(1, len(footer_row) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Ligne de résumé
        summary_row = ['', '']
        summary_row.extend([''] * days_in_month)
        summary_row.extend([
            '',
            '',
            '',
            '',
            f"P-vis:{total_visit_sum}, pan:{total_panne_sum}, acc:{total_accident_sum}",
            f"vis:{total_visitechique_sum}, ent:{total_entretien_sum}, aut:{total_autrarret_sum}"
        ])
        ws.append(summary_row)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        for col_num in range(3, days_in_month + 3):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 10
        
        # Colonnes finales
        for col_num in range(days_in_month + 3, len(headers_row1) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20
        
        # Préparer la réponse HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Temps-Arret-{month_name}-{year}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

class AnalytiqueFicheView(LoginRequiredMixin, CustomPermissionRequiredMixin,TemplateView):
    login_url = 'login'
    permission_url = 'analytique_fiche'
    model = Vehicule
    template_name = 'perfect/aanalytique_fiche.html'
    form_class = DateFormAnalytique

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = now().date()
        # Mois et année par défaut (mois en cours)
        month = int(self.request.GET.get('month', today.month))
        year = int(self.request.GET.get('year', today.year))
        # Véhicules de base
        vehicules = Vehicule.objects.select_related('category').filter(car_statut=True).order_by('immatriculation')
        form = self.form_class(self.request.GET)
        date_debut = None
        date_fin = None
        selected_categorie_id = None
        immatriculation_filter = None

        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            categorie_obj = form.cleaned_data.get('categorie')
            if categorie_obj:
                selected_categorie_id = categorie_obj.id
            immatriculation_filter = form.cleaned_data.get('immatriculation')

        # Override catégorie/immat depuis GET si présents (compat MyRecetteView)
        cat_get = self.request.GET.get('categorie', '').strip()
        if cat_get:
            try:
                selected_categorie_id = int(cat_get)
            except (ValueError, TypeError):
                selected_categorie_id = None
        immat_get = self.request.GET.get('immatriculation', '').strip()
        if immat_get:
            immatriculation_filter = immat_get

        # Appliquer les filtres
        if selected_categorie_id:
            vehicules = vehicules.filter(category_id=selected_categorie_id)
        if immatriculation_filter:
            vehicules = vehicules.filter(immatriculation__icontains=immatriculation_filter)

        # Plage de dates : dates fournies OU mois/année
        if date_debut and date_fin:
            start_date = date_debut
            end_date = date_fin
        else:
            from calendar import monthrange
            _, last_day = monthrange(year, month)
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)

        mois_fr = ('', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre')
        month_name = mois_fr[month] if 1 <= month <= 12 else datetime(year, month, 1).strftime("%B")
        month_choices = [(i, mois_fr[i]) for i in range(1, 13)]
        categories_list = CategoVehi.objects.all().order_by('category')

        # Calcul par véhicule
        resultat_vehicule = []
        total_recettes = 0
        total_charg_fix = 0
        total_charg_var = 0
        total_charges = 0
        total_marge_contrib = 0

        for idx, vehicule in enumerate(vehicules, start=1):
            recettes = Recette.objects.filter(
                vehicule=vehicule,
                date_saisie__range=[start_date, end_date]
            ).aggregate(s=Sum('montant'))['s'] or 0

            charge_fix = ChargeFixe.objects.filter(
                vehicule=vehicule,
                date_saisie__range=[start_date, end_date]
            ).aggregate(s=Sum('montant'))['s'] or 0

            charge_var = ChargeVariable.objects.filter(
                vehicule=vehicule,
                date_saisie__range=[start_date, end_date]
            ).aggregate(s=Sum('montant'))['s'] or 0

            total_charge = charge_fix + charge_var
            marge_contribution = recettes - charge_var
            marge_cout_direct = recettes - total_charge

            if recettes > 0:
                taux_marge = round((marge_contribution * 100) / recettes, 2)
                taux_marge_str = f"{taux_marge:.2f}%"
            else:
                taux_marge = 0
                taux_marge_str = "0,00%"

            resultat_vehicule.append({
                'numero': idx,
                'vehicule': vehicule,
                'recettes': recettes,
                'charge_fix': charge_fix,
                'charge_var': charge_var,
                'total_charge': total_charge,
                'marge_contribution': marge_contribution,
                'taux_marge': taux_marge,
                'taux_marge_str': taux_marge_str,
                'marge_cout_direct': marge_cout_direct,
            })

            total_recettes += recettes
            total_charg_fix += charge_fix
            total_charg_var += charge_var
            total_charges += total_charge
            total_marge_contrib += marge_contribution

        total_marge_cout_direct = total_recettes - total_charges
        # Charges administratives sur la période (mois en cours par défaut)
        total_charge_admin = ChargeAdminis.objects.filter(
            date_saisie__range=[start_date, end_date]
        ).aggregate(s=Sum('montant'))['s'] or 0
        # TOTAL GENERALE = TOTAL MARGE SUR COÛT DIRECT - CHARGES ADMINISTRATIVES
        total_generale = total_marge_cout_direct - total_charge_admin
        if total_recettes > 0:
            taux_marge_global = round((total_marge_contrib * 100) / total_recettes, 2)
        else:
            taux_marge_global = 0

        period_label = f"{month_name} {year}" if not (date_debut and date_fin) else f"Du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"

        context.update({
            'form': form,
            'resultat_vehicule': resultat_vehicule,
            'month': month,
            'year': year,
            'month_name': month_name,
            'period_label': period_label,
            'month_range': range(1, 13),
            'month_choices': month_choices,
            'years': range(today.year - 4, today.year + 1),
            'categories_list': categories_list,
            'selected_categorie_id': selected_categorie_id,
            'total_recettes': total_recettes,
            'total_charg_fix': total_charg_fix,
            'total_charg_var': total_charg_var,
            'total_charges': total_charges,
            'total_charge_admin': total_charge_admin,
            'total_marge_contrib': total_marge_contrib,
            'total_marge_cout_direct': total_marge_cout_direct,
            'total_generale': total_generale,
            'taux_marge_global': taux_marge_global,
        })
        return context

def _get_analytique_fiche_data(request):
    """Récupère les données de la fiche analytique (partagé entre vue et export Excel)."""
    today = now().date()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    vehicules = Vehicule.objects.select_related('category').filter(car_statut=True).order_by('immatriculation')
    form = DateFormAnalytique(request.GET)
    date_debut = None
    date_fin = None
    selected_categorie_id = None
    immatriculation_filter = None

    if form.is_valid():
        date_debut = form.cleaned_data.get('date_debut')
        date_fin = form.cleaned_data.get('date_fin')
        categorie_obj = form.cleaned_data.get('categorie')
        if categorie_obj:
            selected_categorie_id = categorie_obj.id
        immatriculation_filter = form.cleaned_data.get('immatriculation')

    cat_get = request.GET.get('categorie', '').strip()
    if cat_get:
        try:
            selected_categorie_id = int(cat_get)
        except (ValueError, TypeError):
            selected_categorie_id = None
    immat_get = request.GET.get('immatriculation', '').strip()
    if immat_get:
        immatriculation_filter = immat_get

    if selected_categorie_id:
        vehicules = vehicules.filter(category_id=selected_categorie_id)
    if immatriculation_filter:
        vehicules = vehicules.filter(immatriculation__icontains=immatriculation_filter)

    if date_debut and date_fin:
        start_date = date_debut
        end_date = date_fin
    else:
        from calendar import monthrange
        _, last_day = monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, last_day)

    mois_fr = ('', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre')
    month_name = mois_fr[month] if 1 <= month <= 12 else datetime(year, month, 1).strftime("%B")
    period_label = f"{month_name} {year}" if not (date_debut and date_fin) else f"Du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"

    resultat_vehicule = []
    total_recettes = 0
    total_charg_fix = 0
    total_charg_var = 0
    total_charges = 0
    total_marge_contrib = 0

    for idx, vehicule in enumerate(vehicules, start=1):
        recettes = Recette.objects.filter(vehicule=vehicule, date_saisie__range=[start_date, end_date]).aggregate(s=Sum('montant'))['s'] or 0
        charge_fix = ChargeFixe.objects.filter(vehicule=vehicule, date_saisie__range=[start_date, end_date]).aggregate(s=Sum('montant'))['s'] or 0
        charge_var = ChargeVariable.objects.filter(vehicule=vehicule, date_saisie__range=[start_date, end_date]).aggregate(s=Sum('montant'))['s'] or 0
        total_charge = charge_fix + charge_var
        marge_contribution = recettes - charge_var
        marge_cout_direct = recettes - total_charge
        taux_marge = round((marge_contribution * 100) / recettes, 2) if recettes > 0 else 0
        taux_marge_str = f"{taux_marge:.2f}%" if recettes > 0 else "0,00%"

        resultat_vehicule.append({
            'numero': idx, 'vehicule': vehicule, 'recettes': recettes, 'charge_fix': charge_fix,
            'charge_var': charge_var, 'total_charge': total_charge, 'marge_contribution': marge_contribution,
            'taux_marge': taux_marge, 'taux_marge_str': taux_marge_str, 'marge_cout_direct': marge_cout_direct,
        })
        total_recettes += recettes
        total_charg_fix += charge_fix
        total_charg_var += charge_var
        total_charges += total_charge
        total_marge_contrib += marge_contribution

    total_marge_cout_direct = total_recettes - total_charges
    # Charges administratives sur la période (mois en cours par défaut)
    total_charge_admin = ChargeAdminis.objects.filter(
        date_saisie__range=[start_date, end_date]
    ).aggregate(s=Sum('montant'))['s'] or 0
    # TOTAL GENERALE = TOTAL MARGE SUR COÛT DIRECT - CHARGES ADMINISTRATIVES
    total_generale = total_marge_cout_direct - total_charge_admin
    taux_marge_global = round((total_marge_contrib * 100) / total_recettes, 2) if total_recettes > 0 else 0

    return {
        'resultat_vehicule': resultat_vehicule,
        'period_label': period_label,
        'total_recettes': total_recettes,
        'total_charg_fix': total_charg_fix,
        'total_charg_var': total_charg_var,
        'total_charges': total_charges,
        'total_charge_admin': total_charge_admin,
        'total_marge_contrib': total_marge_contrib,
        'total_marge_cout_direct': total_marge_cout_direct,
        'total_generale': total_generale,
        'taux_marge_global': taux_marge_global,
    }

class ExportAnalytiqueFicheExcelView(LoginRequiredMixin, View):
    """Export Excel de la fiche analytique exploitation."""
    def get(self, request, *args, **kwargs):
        data = _get_analytique_fiche_data(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fiche Analytique"

        headers = [
            "N°", "MATRICULE", "MARQUE", "CATEGORIE",
            "RECETTES", "CHARGES FIXES", "CHARGE VARIABLE", "TOTAL CHARGES",
            "MARGES DE CONTRIBUTIONS", "TAUX MARGE (%)", "MARGE SUR COÛT DIRECT"
        ]

        ws.merge_cells('A1:K1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"FICHE ANALYTIQUE EXPLOITATION P&B ENTREPRISE - {data['period_label'].upper()}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        ws.append([])
        ws.append(headers)

        header_fill = PatternFill(start_color="1170EC", end_color="1170EC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        taux_ok_fill = PatternFill(start_color="D5F779", end_color="D5F779", fill_type="solid")

        for row in data['resultat_vehicule']:
            rec = row['recettes'] if row['recettes'] else "-"
            cf = row['charge_fix'] if row['charge_fix'] else "-"
            cv = row['charge_var'] if row['charge_var'] else "-"
            tc = row['total_charge'] if row['total_charge'] else "-"
            mc = row['marge_contribution'] if row['marge_contribution'] else "-"
            mcd = row['marge_cout_direct'] if row['marge_cout_direct'] else "-"
            row_data = [
                row['numero'],
                row['vehicule'].immatriculation,
                row['vehicule'].marque,
                row['vehicule'].category.category,
                rec, cf, cv, tc, mc, row['taux_marge_str'], mcd
            ]
            ws.append(row_data)
            taux_cell = ws.cell(row=ws.max_row, column=10)
            if row['taux_marge'] >= 90:
                taux_cell.fill = taux_ok_fill

        total_row = [
            "TOTAL", "", "", "",
            data['total_recettes'], data['total_charg_fix'], data['total_charg_var'],
            data['total_charges'], data['total_marge_contrib'],
            f"{data['taux_marge_global']}%", data['total_marge_cout_direct']
        ]
        ws.append(total_row)
        for col_num in range(1, len(total_row) + 1):
            ws.cell(row=ws.max_row, column=col_num).font = Font(bold=True)

        # Ligne TOTAL GENERALE (marge sur coût direct - charge variable globale)
        generale_row = [
            "TOTAL GENERALE", "", "", "",
            "", "", "", "", "", "",
            data['total_generale']
        ]
        ws.append(generale_row)
        for col_num in range(1, len(generale_row) + 1):
            ws.cell(row=ws.max_row, column=col_num).font = Font(bold=True)

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Fiche-Analytique-{data['period_label'].replace(' ', '-').replace('/', '-')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class MyRecetteView(CustomPermissionRequiredMixin, LoginRequiredMixin, TemplateView):
    login_url = 'login'
    template_name = "perfect/myrecette.html"
    permission_url = 'rec_day'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = now().date()
        month = self.request.GET.get('month', today.month)
        year = self.request.GET.get('year', today.year)
        month = int(month)
        year = int(year)
        days_in_month = monthrange(year, month)[1]
        mois_fr = ('', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre')
        month_name = mois_fr[month] if 1 <= month <= 12 else datetime(year, month, 1).strftime("%B")
        month_choices = [(i, mois_fr[i]) for i in range(1, 13)]
        # Calculer les dimanches dans le mois
        dimanches = [
            day for day in range(1, days_in_month + 1)
            if datetime(year, month, day).weekday() == SUNDAY
        ]
        jours_ouvrables = days_in_month - len(dimanches)
        vehicules = Vehicule.objects.select_related('category').filter(car_statut=True)
        # Filtre par catégorie si sélectionnée
        selected_categorie_id = self.request.GET.get('categorie', '').strip()
        if selected_categorie_id:
            try:
                selected_categorie_id = int(selected_categorie_id)
                vehicules = vehicules.filter(category_id=selected_categorie_id)
            except (ValueError, TypeError):
                selected_categorie_id = None
        else:
            selected_categorie_id = None
        categories_list = CategoVehi.objects.all().order_by('category')
        recette_details = []
        total_recette_mensuelle = 0
        total_recette_annuelle = 0
        sum_recets_jours = 0
        sum_difference_mensuelle = 0
        sum_recettes_vehicule_mois = 0
        sum_difference = 0
        sum_motif_arrets = 0
        for vehicule in vehicules:
            # Recette journalière
            recettes_vehicule_jour = Recette.objects.filter(
                vehicule=vehicule,
                date_saisie=date.today()
            ).aggregate(total_recette=Sum('montant'))['total_recette'] or 0
            # Recette mensuelle
            recettes_vehicule_mois = Recette.objects.filter(
                vehicule=vehicule,
                date_saisie__month=month,
                date_saisie__year=year
            ).aggregate(total_recette=Sum('montant'))['total_recette'] or 0
            # Recette annuelle
            recettes_vehicule_an = Recette.objects.filter(
                vehicule=vehicule,
                date_saisie__year=year
            ).aggregate(total_recette=Sum('montant'))['total_recette'] or 0
            # Recette par défaut de la catégorie
            recette_defaut = vehicule.category.recette_defaut
            # Calcul de la recette mensuelle attendue sans dimanches
            recette_attendue_mensuelle = recette_defaut * jours_ouvrables
            # Calcul de la différence pour aujourd'hui
            difference = recettes_vehicule_jour - recette_defaut
            difference_mensuelle = recettes_vehicule_mois - recette_attendue_mensuelle
            sum_difference_mensuelle += difference_mensuelle
            sum_recettes_vehicule_mois += recettes_vehicule_mois
            sum_difference += difference
            
            motif_arrets = {
                'vis': VisiteTechnique.objects.filter(vehicule=vehicule,date_saisie=date.today()).count(),
                'ent': Entretien.objects.filter(vehicule=vehicule, date_saisie=date.today()).count(),
                'rep': Reparation.objects.filter(vehicule=vehicule, date_saisie=date.today()).count(),
            }
            visite = VisiteTechnique.objects.filter(vehicule=vehicule,date_saisie=date.today()).count()
            entretien = Entretien.objects.filter(vehicule=vehicule, date_saisie=date.today()).count()
            reparation = Reparation.objects.filter(vehicule=vehicule, date_saisie=date.today()).count()
            som_des_motifs = visite+entretien+reparation
            
            sum_motif_arrets += som_des_motifs
            sum_recets_jours += recettes_vehicule_jour
            daily_actions = [0] * days_in_month
            for day in range(1, days_in_month + 1):
                for model in [Recette]:
                    motant = model.objects.filter(
                        vehicule=vehicule, date_saisie__day=day, date_saisie__month=month, date_saisie__year=year
                    ).aggregate(somme=Sum('montant'))['somme'] or 0
                    daily_actions[day - 1] += motant
            
            recette_details.append({
                'vehicule': vehicule.immatriculation,
                'marque': vehicule.marque,
                'category': vehicule.category.category,
                'recette_versee': recettes_vehicule_jour,
                'recette_attendue': recette_defaut,
                'daily_actions': daily_actions,
                'difference': difference,
                'difference_mensuelle': difference_mensuelle,
                'recette_mensuelle': recettes_vehicule_mois,
                'recette_annuelle': recettes_vehicule_an,
                'motif_arrets': motif_arrets,
            })
        
        recette_par_categorie = (
            Recette.objects.filter(date_saisie__month=month, date_saisie__year=year)
            .values('vehicule__category__id', 'vehicule__category__category')
            .annotate(total_verse=Sum('montant'))
            .order_by('vehicule__category__category')
        )

        verse_dict = {item['vehicule__category__id']: item['total_verse'] for item in recette_par_categorie}
        categories = CategoVehi.objects.annotate(nb_vehicules=Count('catego_vehicule'))
        if selected_categorie_id:
            categories = categories.filter(id=selected_categorie_id)
        recap_categorie = []

        for cat in categories:
            nb_vehicules = cat.nb_vehicules
            recette_defaut = cat.recette_defaut
            recette_attendue = recette_defaut * jours_ouvrables * nb_vehicules
            recette_verse = verse_dict.get(cat.id, 0) or 0
            ecart = recette_verse - recette_attendue

            recap_categorie.append({
                'categorie': cat.category,
                'nb_vehicules': nb_vehicules,
                'recette_defaut': recette_defaut,
                'recette_attendue': recette_attendue,
                'recette_verse': recette_verse,
                'ecart': ecart,
            })
        # Calculer les totaux par date (colonne)
        total_cost_parts_sum = sum(vehicule.category.recette_defaut for vehicule in vehicules)
        totals_by_day = []
        for day in range(1, days_in_month + 1):
            qs = Recette.objects.filter(
                date_saisie__day=day,
                date_saisie__month=month,
                date_saisie__year=year
            )
            if selected_categorie_id:
                qs = qs.filter(vehicule__category_id=selected_categorie_id)
            total_day = qs.aggregate(somme=Sum('montant'))['somme'] or 0
            totals_by_day.append(total_day)
        
        context={
            'recap_categorie': recap_categorie,
            'sum_motif_arrets': sum_motif_arrets,
            'sum_difference': sum_difference,
            'sum_recettes_vehicule_mois': sum_recettes_vehicule_mois,
            'sum_difference_mensuelle': sum_difference_mensuelle,
            'sum_recets_jours': sum_recets_jours,
            'recette_details': recette_details,
            'total_recette_mensuelle': total_recette_mensuelle,
            'total_recette_annuelle': total_recette_annuelle,
            'current_date': today,
            'days_in_month': range(1, days_in_month + 1),
            'month_name': month_name,
            'month': month,
            'year': year,
            'years': range(today.year - 4, today.year + 1),
            'month_range': range(1, 13),
            'month_choices': month_choices,
            'days_in_month_plus_two': days_in_month + 2,
            'totals_by_day': totals_by_day,
            'total_cost_parts_sum': total_cost_parts_sum,
            'categories_list': categories_list,
            'selected_categorie_id': selected_categorie_id,
        }
        return context

class ExportRecetteMensuelleExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = now().date()
        month = request.GET.get('month', today.month)
        year = request.GET.get('year', today.year)
        month = int(month)
        year = int(year)
        days_in_month = monthrange(year, month)[1]
        month_name = datetime(year, month, 1).strftime("%B")
        
        # Calculer les dimanches dans le mois
        dimanches = [
            day for day in range(1, days_in_month + 1)
            if datetime(year, month, day).weekday() == SUNDAY
        ]
        jours_ouvrables = days_in_month - len(dimanches)

        vehicules = Vehicule.objects.select_related('category').all()
        selected_categorie_id = request.GET.get('categorie', '').strip()
        if selected_categorie_id:
            try:
                selected_categorie_id = int(selected_categorie_id)
                vehicules = vehicules.filter(category_id=selected_categorie_id)
            except (ValueError, TypeError):
                selected_categorie_id = None
        else:
            selected_categorie_id = None
        recette_details = []

        # Calculer les totaux par date
        totals_by_day = []
        for day in range(1, days_in_month + 1):
            qs = Recette.objects.filter(
                date_saisie__day=day,
                date_saisie__month=month,
                date_saisie__year=year
            )
            if selected_categorie_id:
                qs = qs.filter(vehicule__category_id=selected_categorie_id)
            total_day = qs.aggregate(somme=Sum('montant'))['somme'] or 0
            totals_by_day.append(total_day)

        for vehicule in vehicules:
            recettes_vehicule_jour = Recette.objects.filter(
                vehicule=vehicule,
                date_saisie=date.today()
            ).aggregate(total_recette=Sum('montant'))['total_recette'] or 0
            
            recettes_vehicule_mois = Recette.objects.filter(
                vehicule=vehicule,
                date_saisie__month=month,
                date_saisie__year=year
            ).aggregate(total_recette=Sum('montant'))['total_recette'] or 0
            
            recette_defaut = vehicule.category.recette_defaut
            recette_attendue_mensuelle = recette_defaut * jours_ouvrables
            difference = recettes_vehicule_jour - recette_defaut
            difference_mensuelle = recettes_vehicule_mois - recette_attendue_mensuelle
            
            motif_arrets = {
                'vis': VisiteTechnique.objects.filter(vehicule=vehicule, date_saisie=date.today()).count(),
                'ent': Entretien.objects.filter(vehicule=vehicule, date_saisie=date.today()).count(),
                'rep': Reparation.objects.filter(vehicule=vehicule, date_saisie=date.today()).count(),
            }
            
            daily_actions = [0] * days_in_month
            for day in range(1, days_in_month + 1):
                motant = Recette.objects.filter(
                    vehicule=vehicule, 
                    date_saisie__day=day, 
                    date_saisie__month=month, 
                    date_saisie__year=year
                ).aggregate(somme=Sum('montant'))['somme'] or 0
                daily_actions[day - 1] = motant
            
            recette_details.append({
                'vehicule': vehicule.immatriculation,
                'marque': vehicule.marque,
                'recette_versee': recettes_vehicule_jour,
                'recette_attendue': recette_defaut,
                'daily_actions': daily_actions,
                'difference': difference,
                'difference_mensuelle': difference_mensuelle,
                'recette_mensuelle': recettes_vehicule_mois,
                'motif_arrets': motif_arrets,
            })
        
        # Calculer les totaux
        sum_recets_jours = sum(r['recette_versee'] for r in recette_details)
        sum_difference = sum(r['difference'] for r in recette_details)
        sum_recettes_vehicule_mois = sum(r['recette_mensuelle'] for r in recette_details)
        sum_difference_mensuelle = sum(r['difference_mensuelle'] for r in recette_details)
        total_cost_parts_sum = sum(vehicule.category.recette_defaut for vehicule in vehicules)
        
        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Recettes {month_name} {year}"
        
        # En-tête principal
        ws.merge_cells(f'A1:{get_column_letter(days_in_month + 8)}1')
        header_cell = ws.cell(row=1, column=1)
        header_cell.value = f"RECETTES MENSUELLES - {month_name.upper()} {year}"
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center')
        # Première ligne d'en-têtes (avec fusion pour le mois)
        ws.append([])  # Ligne vide
        
        # En-têtes des colonnes
        headers = ['Immatriculation', 'Marque']
        headers.extend([f'Jour {day}' for day in range(1, days_in_month + 1)])
        headers.extend(['Aujourd\'hui', 'A verser', 'Ecart', 'T Recette', 'A payer', 'Motif du jour'])
        ws.append(headers)
        
        # Style des en-têtes
        header_row = 3
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données des véhicules
        for detail in recette_details:
            row = [
                detail['vehicule'],
                detail['marque'],
            ]
            row.extend(detail['daily_actions'])
            row.extend([
                detail['recette_versee'],
                detail['recette_attendue'],
                detail['difference'],
                detail['recette_mensuelle'],
                detail['difference_mensuelle'],
                f"Vis:{detail['motif_arrets']['vis']}, Ent:{detail['motif_arrets']['ent']}, Rep:{detail['motif_arrets']['rep']}"
            ])
            ws.append(row)
        
        # Ligne de totaux par date dans le footer
        footer_row = ['TOTAL PAR JOUR', '']
        footer_row.extend(totals_by_day)
        footer_row.extend([
            sum_recets_jours,
            total_cost_parts_sum,
            sum_difference,
            sum_recettes_vehicule_mois,
            sum_difference_mensuelle,
            ''
        ])
        ws.append(footer_row)
        
        # Style de la ligne de totaux
        for col_num in range(1, len(footer_row) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        for col_num in range(3, days_in_month + 3):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 10
        
        # Colonnes finales
        for col_num in range(days_in_month + 3, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
        
        # Préparer la réponse HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Recettes-{month_name}-{year}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

class DashboardView(CustomPermissionRequiredMixin,LoginRequiredMixin,TemplateView):
    login_url = 'login'
    permission_url = 'dash'
    template_name = 'perfect/dashboard.html'
    form_class = DateFormMJR
    timeout_minutes = 600
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get(self, request, *args, **kwargs):
        request.session['last_activity'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return super().get(request, *args, **kwargs)
    def post(self, request, *args, **kwargs):
        request.session['last_activity'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return super().post(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois = date.today().month
        mois_en_cours =date.today().month
        libelle_mois_en_cours = calendar.month_name[mois_en_cours]
        label = [calendar.month_name[month][:1] for month in range(1, 13)]
        marge_cont, best_recets, best_marge, best_taux, labelscat, datacat, datasets, best_recets, best_marge, best_taux = [],[],[],[],[],[],[],[],[],[]
        
        jours_semaine = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]
        categories = CategoVehi.objects.all()
        vehicules = Vehicule.objects.filter(car_statut=True)

        recette_queryset = Recette.objects.all()
        chargfix_queryset = ChargeFixe.objects.all()
        chargvar_queryset = ChargeVariable.objects.all()
        reparation_queryset = Reparation.objects.all()
        piechan_queryset = PiecEchange.objects.all()
        piece_queryset = Piece.objects.all()

        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=5)

        all_vehicule = vehicules
        # Initialiser les variables de filtre par défaut (mois en cours)
        date_debut = None
        date_fin = None
        categorie_filter = None
        immatriculation = None
        
        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            
            if categorie_filter:
                recette_queryset = recette_queryset.filter(vehicule__category__category=categorie_filter)
                chargfix_queryset = chargfix_queryset.filter(vehicule__category__category=categorie_filter)
                chargvar_queryset = chargvar_queryset.filter(vehicule__category__category=categorie_filter)
                reparation_queryset = reparation_queryset.filter(vehicule__category__category=categorie_filter)
                piechan_queryset = piechan_queryset.filter(vehicule__category__category=categorie_filter)
                piece_queryset = piece_queryset.filter(reparation__vehicule__category__category=categorie_filter)

            if date_debut and date_fin:
                recette_queryset = recette_queryset.filter(date_saisie__range=[date_debut, date_fin])
                chargfix_queryset = chargfix_queryset.filter(date_saisie__range=[date_debut, date_fin])
                chargvar_queryset = chargvar_queryset.filter(date_saisie__range=[date_debut, date_fin])
                reparation_queryset = reparation_queryset.filter(date_saisie__range=[date_debut, date_fin])
                piechan_queryset = piechan_queryset.filter(date_saisie__range=[date_debut, date_fin])
                piece_queryset = piece_queryset.filter(date_saisie__range=[date_debut, date_fin])

            if immatriculation:
                recette_queryset = recette_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                chargfix_queryset = chargfix_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                chargvar_queryset = chargvar_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                reparation_queryset = reparation_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                piechan_queryset = piechan_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                piece_queryset = piece_queryset.filter(reparation__vehicule__immatriculation__icontains=immatriculation)

        # Appliquer le filtre par date : si dates fournies, les utiliser, sinon filtrer par mois en cours
        if date_debut and date_fin:
            filtre_recette = recette_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_piece = piece_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_piec_echange = piechan_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_chargfix = chargfix_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_chargvar = chargvar_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_reparation = reparation_queryset.filter(date_saisie__range=[date_debut, date_fin])
            start_of_week, end_of_week = date_debut, date_fin
        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_recette = recette_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_piece = piece_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_piec_echange = piechan_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_chargfix = chargfix_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_chargvar = chargvar_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_reparation = reparation_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)

        # ################################----Recettes----#############################
        total_recettes = filtre_recette.aggregate(somme=Sum('montant'))['somme'] or 1
        total_recette_format ='{:,}'.format(total_recettes).replace(',', ' ')
        # ################################----Pieces----#############################
        total_piece=filtre_piece.aggregate(somme=Sum('montant'))['somme'] or 0
        total_piece_format ='{:,}'.format(total_piece).replace(',', ' ')
        # ################################-----#----Pieces echanges----#-----#############################
        total_piec_echange= filtre_piec_echange.aggregate(somme=Sum('montant'))['somme'] or 0
        total_piece_echang_format ='{:,}'.format(total_piec_echange).replace(',', ' ')
        # ################################----Charges----#############################
        total_charg_fix = filtre_chargfix.aggregate(somme=Sum('montant'))['somme'] or 0
        total_chargfix_format ='{:,}'.format(total_charg_fix).replace(',', ' ')
        total_charg_var = filtre_chargvar.aggregate(somme=Sum('montant'))['somme'] or 0
        total_chargvar_format ='{:,}'.format(total_charg_var).replace(',', ' ')
        # ################################----Charge Totale----#############################
        total_charg = total_charg_fix + total_charg_var
        total_charge_format ='{:,}'.format(total_charg).replace(',', ' ')
        ################################----Marge de contribution----#############################
        marge_contribution = total_recettes - total_charg
        marge_contribution = total_recettes - total_charg_var
        # ################################----Taux----#############################
        if total_recettes == 1:
            taux_marge = 0
        else:
            taux_marge = (marge_contribution*100/(total_recettes))
        taux_marge_format ='{:.2f}'.format(taux_marge)
        
        # ################################----Marge brute----#############################
        marge_brute = total_recettes - total_charg
        marge_brute_format ='{:,}'.format(marge_brute).replace(',', ' ')
        
        # ################################----Graphiques----#############################
        recet_data = recette_queryset.filter(date_saisie__year=datetime.now().year)
        recet_mois_data = {month: 0 for month in range(1, 13)}
        for commande in recet_data:
            recet_mois_data[commande.date_saisie.month] += commande.montant
        recet_mois_data = [recet_mois_data[month] for month in range(1, 13)]
        
        rep_data = reparation_queryset.filter(date_saisie__year=datetime.now().year)
        rep_mois_data = {month: 0 for month in range(1, 13)}
        for commande in rep_data:
            rep_mois_data[commande.date_saisie.month] += commande.montant
        rep_mois_data = [rep_mois_data[month] for month in range(1, 13)]
        
        piec_data = piece_queryset.filter(reparation__in=rep_data, date_saisie__year=datetime.now().year)
        piec_mois_data = {month: 0 for month in range(1, 13)}
        for commande in piec_data:
            piec_mois_data[commande.date_saisie.month] += commande.montant
        piec_mois_data = [piec_mois_data[month] for month in range(1, 13)]
        
        piecha_data = piechan_queryset.filter(date_saisie__year=datetime.now().year)
        piecha_mois_data = {month: 0 for month in range(1, 13)}
        for commande in piecha_data:
            piecha_mois_data[commande.date_saisie.month] += commande.montant
        piecha_mois_data = [piecha_mois_data[month] for month in range(1, 13)]
        
        chargvar_data = chargvar_queryset.filter(date_saisie__year=datetime.now().year)
        chargvar_mois_data = {month: 0 for month in range(1, 13)}
        for commande in chargvar_data:
            chargvar_mois_data[commande.date_saisie.month] += commande.montant
        chargvar_mois_data = [chargvar_mois_data[month] for month in range(1, 13)]
        
        # ####################################################################################################
        chargfix_data = chargfix_queryset.filter(date_saisie__year=datetime.now().year)
        chargfix_mois_data = {month: 0 for month in range(1, 13)}
        for commande in chargfix_data:
            chargfix_mois_data[commande.date_saisie.month] += commande.montant
        chargfix_mois_data = [chargfix_mois_data[month] for month in range(1, 13)]
        
        # #####################################################################################################
        chargvar_data = chargvar_queryset.filter(date_saisie__year=datetime.now().year)
        chargvar_mois_data = {month: 0 for month in range(1, 13)}
        for commande in chargvar_data:
            chargvar_mois_data[commande.date_saisie.month] += commande.montant
        chargvar_mois_data = [chargvar_mois_data[month] for month in range(1, 13)]
        
        # ################################VTC################################
        marge_data = [recet_mois_data[i] - chargvar_mois_data[i] for i in range(12)]
        taux_data = [(marge_data[i] * 100) / recet_mois_data[i] if recet_mois_data[i] > 0 else 0 for i in range(12)]
        
        for vehicule in all_vehicule:
            recs = filtre_recette.filter(vehicule = vehicule).aggregate(somme=Sum('montant'))['somme'] or 0
            rece_all = recs if recs is not None else 0
            charges_variables = chargvar_queryset.filter(vehicule = vehicule, date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
            chargvari_all = charges_variables if charges_variables is not None else 0
            marge_cont = rece_all - chargvari_all
            marge = marge_cont if marge_cont is not None else 0
            if rece_all == 0:
                taux=0
            else:
                taux = round((marge*100)/rece_all,2)
                taux = taux if taux is not None else 0
            best_taux.append({'vehicule': vehicule, 'taux':taux})
            best_marge.append({'vehicule': vehicule, 'marge_cont':marge_cont})
            best_recets.append({'vehicule': vehicule, 'recs':recs})
        best_marge = sorted([x for x in best_marge if x['marge_cont'] is not None], key=lambda x: x['marge_cont'], reverse=True)[:5] 
        best_taux = sorted(best_taux, key=lambda x: x['taux'], reverse=True)[:5]  
        best_recets = sorted([x for x in best_recets if x['recs'] is not None], key=lambda x: x['recs'], reverse=True)[:5]

        rectte_par_categorie = filtre_recette.values(
            'vehicule__category__category'
        ).annotate(
            total=Sum('montant')
        ).order_by('-total')
        # Extraire les données pour le graphique
        labelscat = [item['vehicule__category__category'] for item in rectte_par_categorie]
        datacat = [item['total'] for item in rectte_par_categorie]

        for categorie in categories:
            ventes = [0] * 6
            lignes = (
                filtre_recette.filter(
                    vehicule__category__category=categorie,
                    date_saisie__range=[start_of_week, end_of_week]
                ).annotate(
                    jour=F("date_saisie"),
                    total=Sum("montant")
                ).values("jour").annotate(total=Sum("montant")).order_by('-total')
            )
            for ligne in lignes:
                day_index = ligne["jour"].weekday()
                if day_index < 6:
                    ventes[day_index] = float(ligne["total"]) if ligne["total"] else 0

            datasets.append({
                "label": categorie.category.upper(),
                "data": ventes,
                "fill": True,
            })

        top_reparations = (
            filtre_reparation.values(
                'vehicule__immatriculation'
            )
            .annotate(total_reparations=Count('id'))
            .order_by('-total_reparations')[:5]
        )

        context={
            'total_recette_format':total_recette_format,
            "datasets_json": json.dumps(datasets),
            "jours_semaine" : jours_semaine,
            
            'vehicules':vehicules,
            'tot_piece':total_piece_format,
            'tot_piec_echang':total_piece_echang_format,
            'tot_chargfix':total_chargfix_format,
            'tot_chargvar':total_chargvar_format,
            'charge_totale':total_charge_format,
            'marge_brute_format':marge_brute_format,
            'taux_marge':taux_marge_format,
            'best_recettes':best_recets,
            'top_reparations':top_reparations,
            'bests_taux':best_taux,
            'mois':libelle_mois_en_cours,
            
            #################---###############-----Graphiq-----#################---###############
            'recet_mois_data':recet_mois_data,
            'taux_data':taux_data,
            'chargvar_mois_data':chargvar_mois_data,
            'chargfix_mois_data':chargfix_mois_data,
            
            'piecha_mois_data':piecha_mois_data,
            'piec_mois_data':piec_mois_data,
            'labels':label,
            'form':form,
            'dates':dates,

            'labelscat':labelscat,
            'datacat':datacat,
            # 'permissions':permissions
        }
        return context

class DashboardGaragView(LoginRequiredMixin, CustomPermissionRequiredMixin, TemplateView):
    login_url = 'login'
    model = Vehicule
    template_name = 'perfect/dash_garag.html'
    permission_url = 'dashgarage'
    timeout_minutes = 600
    form_class = DateFormMJR

    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois = date.today().month
        mois_en_cours =date.today().month
        libelle_mois_en_cours = calendar.month_name[mois_en_cours]
        label = [calendar.month_name[month][:1] for month in range(1, 13)]
        user = self.request.user
        labelscat, datacat = [], []
        # Initialiser les variables de filtre par défaut (pour le filtrage du mois en cours)
        date_debut = None
        date_fin = None
        categorie_filter = None
        immatriculation = None
        # Gestion des véhicules selon le type d'utilisateur
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        all_vehicule = vehicules
        # Initialiser les querysets
        visitech_queryset = VisiteTechnique.objects.all()
        reparation_queryset = Reparation.objects.all()
        piechan_queryset = PiecEchange.objects.all()
        ligne_piechan_queryset = LignePiecEchange.objects.all()
        piece_queryset = Piece.objects.all()
        entretien_queryset = Entretien.objects.all()
        stationnement_queryset = Stationnement.objects.all()
        assurance_queryset = Assurance.objects.all()
        patente_queryset = Patente.objects.all()
        vignette_queryset = Vignette.objects.all()
        # Calculer la semaine en cours par défaut
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=5)
        
        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            
            if categorie_filter:
                reparation_queryset = reparation_queryset.filter(vehicule__category__category=categorie_filter)
                piechan_queryset = piechan_queryset.filter(vehicule__category__category=categorie_filter)
                piece_queryset = piece_queryset.filter(reparation__vehicule__category__category=categorie_filter)
                entretien_queryset = entretien_queryset.filter(vehicule__category__category=categorie_filter)
                visitech_queryset = visitech_queryset.filter(vehicule__category__category=categorie_filter)

                stationnement_queryset = stationnement_queryset.filter(vehicule__category__category=categorie_filter)
                assurance_queryset = assurance_queryset.filter(vehicule__category__category=categorie_filter)
                patente_queryset = patente_queryset.filter(vehicule__category__category=categorie_filter)
                vignette_queryset = vignette_queryset.filter(vehicule__category__category=categorie_filter)
                
            if date_debut and date_fin:
                visitech_queryset = visitech_queryset.filter(date_saisie__range=[date_debut, date_fin])
                reparation_queryset = reparation_queryset.filter(date_saisie__range=[date_debut, date_fin])
                piechan_queryset = piechan_queryset.filter(date_saisie__range=[date_debut, date_fin])
                piece_queryset = piece_queryset.filter(date_saisie__range=[date_debut, date_fin])
                entretien_queryset = entretien_queryset.filter(date_saisie__range=[date_debut, date_fin])
                
                stationnement_queryset = stationnement_queryset.filter(date_saisie__range=[date_debut, date_fin])
                assurance_queryset = assurance_queryset.filter(date_saisie__range=[date_debut, date_fin])
                patente_queryset = patente_queryset.filter(date_saisie__range=[date_debut, date_fin])
                vignette_queryset = vignette_queryset.filter(date_saisie__range=[date_debut, date_fin])

            if immatriculation:
                visitech_queryset = visitech_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                reparation_queryset = reparation_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                piechan_queryset = piechan_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                piece_queryset = piece_queryset.filter(reparation__vehicule__immatriculation__icontains=immatriculation)
                entretien_queryset = entretien_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                assurance_queryset = assurance_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
               
                stationnement_queryset = stationnement_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                patente_queryset = patente_queryset.filter(vehicule__immatriculation__icontains=immatriculation)
                vignette_queryset = vignette_queryset.filter(vehicule__immatriculation__icontains=immatriculation)

        # Appliquer le filtre par date : si dates fournies, les utiliser, sinon filtrer par mois en cours
        if date_debut and date_fin:
            filtre_visitech = visitech_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_reparation = reparation_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_piechan = piechan_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_piece = piece_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_entretien = entretien_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_stationnement = stationnement_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_patente = patente_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_vignette = vignette_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_assurance = assurance_queryset.filter(date_saisie__range=[date_debut, date_fin])
            start_of_week, end_of_week = date_debut, date_fin
        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_visitech = visitech_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_reparation = reparation_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_piechan = piechan_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_piece = piece_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_entretien = entretien_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_stationnement = stationnement_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_patente = patente_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_vignette = vignette_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_assurance = assurance_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)

        nb_reparat = filtre_reparation.count()
        total_reparat = filtre_reparation.aggregate(somme=Sum('montant'))['somme'] or 0
        total_reparat_format ='{:,}'.format(total_reparat).replace(',', ' ')
        
        nb_reparatvtc = filtre_reparation.count()
        total_reparatvtc = filtre_reparation.aggregate(somme=Sum('montant'))['somme'] or 0
        total_reparatvtc_format ='{:,}'.format(total_reparatvtc).replace(',', ' ')
        
        nb_reparataxi = filtre_reparation.count()
        total_reparataxi = filtre_reparation.aggregate(somme=Sum('montant'))['somme'] or 0
        total_reparataxi_format ='{:,}'.format(total_reparataxi).replace(',', ' ')
        
        total_visit = filtre_visitech.aggregate(somme=Sum('montant'))['somme'] or 0
        total_visit_format ='{:,}'.format(total_visit).replace(',', ' ')
        
        total_entret = filtre_entretien.aggregate(somme=Sum('montant'))['somme'] or 0
        total_ent_format ='{:,}'.format(total_entret).replace(',', ' ')
        
        total_piece = filtre_piece.aggregate(somme=Sum('montant'))['somme'] or 0
        total_piece_format ='{:,}'.format(total_piece).replace(',', ' ')
        
        total_piecechang = filtre_piechan.aggregate(somme=Sum('montant'))['somme'] or 0
        total_piececha_format ='{:,}'.format(total_piecechang).replace(',', ' ')
        
        total_piece_int = filtre_piece.filter(lieu="INTERNE").aggregate(somme=Sum('montant'))['somme'] or 0

        total_piecechang_int = filtre_piechan.aggregate(somme=Sum('montant'))['somme'] or 0
        # total_piecechang_int = filtre_piechan.filter(lieu="INTERNE").aggregate(somme=Sum('montant'))['somme'] or 0

        total_pieces_int = total_piece_int + total_piecechang_int
        total_pieces_format_int = '{:,}'.format(total_pieces_int).replace(',', ' ')

        total_stationnement = filtre_stationnement.aggregate(somme=Sum('montant'))['somme'] or 0
        total_stationnement_format ='{:,}'.format(total_stationnement).replace(',', ' ')
        
        total_patente = filtre_patente.aggregate(somme=Sum('montant'))['somme'] or 0
        total_patente_format ='{:,}'.format(total_patente).replace(',', ' ')
        
        total_vignette = filtre_vignette.aggregate(somme=Sum('montant'))['somme'] or 0
        total_vignette_format ='{:,}'.format(total_vignette).replace(',', ' ')
        
        total_assurance = filtre_assurance.aggregate(somme=Sum('montant'))['somme'] or 0
        total_assurance_format ='{:,}'.format(total_assurance).replace(',', ' ')
       
        ################################----Graphiques----#############################
        rep_data = reparation_queryset.filter(date_saisie__year=datetime.now().year)
        rep_mois_data = {month: 0 for month in range(1, 13)}
        for commande in rep_data:
            rep_mois_data[commande.date_saisie.month] += commande.montant
        rep_mois_data = [rep_mois_data[month] for month in range(1, 13)]

        vis_data = visitech_queryset.filter(date_saisie__year=datetime.now().year)
        vis_mois_data = {month: 0 for month in range(1, 13)}
        for commande in vis_data:
            vis_mois_data[commande.date_saisie.month] += commande.montant
        vis_mois_data = [vis_mois_data[month] for month in range(1, 13)]
        
        ent_data = entretien_queryset.filter(date_saisie__year=datetime.now().year)
        ent_mois_data = {month: 0 for month in range(1, 13)}
        for commande in ent_data:
            ent_mois_data[commande.date_saisie.month] += commande.montant
        ent_mois_data = [ent_mois_data[month] for month in range(1, 13)]
        
        piec_data = piece_queryset.filter(date_saisie__year=datetime.now().year)
        piec_mois_data = {month: 0 for month in range(1, 13)}
        for commande in piec_data:
            piec_mois_data[commande.date_saisie.month] += commande.montant

        piec_mois_data = [piec_mois_data[month] for month in range(1, 13)]
        
        piecha_data = piechan_queryset.filter(date_saisie__year=datetime.now().year)
        piecha_mois_data = {month: 0 for month in range(1, 13)}
        for commande in piecha_data:
            piecha_mois_data[commande.date_saisie.month] += commande.montant
        piecha_mois_data = [piecha_mois_data[month] for month in range(1, 13)]


        reparat_par_categorie = filtre_reparation.values(
            'vehicule__category__category'
        ).annotate(
            total=Sum('montant')
        ).order_by('-total')
        # Extraire les données pour le graphique
        labelscat = [item['vehicule__category__category'] for item in reparat_par_categorie]
        datacat = [item['total'] for item in reparat_par_categorie]
        
        top_reparations = (
            filtre_reparation.values(
                'vehicule__immatriculation'
            )
            .annotate(total_reparations=Count('id'))
            .order_by('-total_reparations')[:5]
        )
        
        context.update({
                'rep_mois_data':rep_mois_data,
                
                'vis_mois_data':vis_mois_data,
                
                'ent_mois_data':ent_mois_data,
                
                'piec_mois_data':piec_mois_data,
                
                'piecha_mois_data':piecha_mois_data,

                'total_assurance_format':total_assurance_format,
                
                'nb_reparat':nb_reparat,
                'total_reparat_format':total_reparat_format,
                'total_visit_format':total_visit_format,
                'total_ent_format':total_ent_format,
                'total_piece_format':total_piece_format,
                'total_piececha_format':total_piececha_format,

                'total_stationnement_format':total_stationnement_format,
                'total_vignette_format':total_vignette_format,
                'total_patente_format':total_patente_format,

                'nb_reparatvtc':nb_reparatvtc,
                'total_reparatvtc_format':total_reparatvtc_format,

                'nb_reparataxi':nb_reparataxi,
                'total_reparataxi_format':total_reparataxi_format,
                'total_pieces_format_int':total_pieces_format_int,
                
                'labels':label,
                'form':form,
                'dates':dates,
                'vehicules':vehicules, 
                'top_reparations':top_reparations,

                'labelscat':labelscat,  
                'datacat':datacat  
            })
        return context 
 
class BilletageView(CustomPermissionRequiredMixin, CreateView):
    model = Billetage
    permission_url = 'billetage'
    form_class = BilletageForm
    template_name = 'perfect/caisse.html'
    success_message = 'Saisie enrégistrée avec succès✓✓'
    error_message = "Erreur de saisie ✘✘"
    success_url = reverse_lazy ('billetage')
    timeout_minutes = 230
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        reponse =  super().form_valid(form)
        messages.success(self.request, self.success_message)
        return reponse
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        form = self.get_form()

        if self.request.POST:
            formset = BilletageFormSet(self.request.POST, queryset=Billetage.objects.none())
        else:
            formset = BilletageFormSet(queryset=Billetage.objects.none())

        forms = DatebilanForm(self.request.GET)
        if forms.is_valid():
            date_bilan = forms.cleaned_data['date_bilan']
            dates = date_bilan
            encaisser = Encaissement.objects.filter(date_saisie=date_bilan)
            decaisser = Decaissement.objects.filter(date_saisie=date_bilan)
            solde_jour = SoldeJour.objects.filter(date_saisie=date_bilan).aggregate(Sum('montant'))['montant__sum'] or 0
            # Calculer le total des encaissements pour la journée actuelle
            tot_entree = Encaissement.objects.filter(date_saisie=date_bilan).aggregate(somme=Sum('montant'))['somme'] or 0
            tot_sortie = Decaissement.objects.filter(date_saisie=date_bilan).aggregate(somme=Sum('montant'))['somme'] or 0
            
            solde_initial = None
            # Vérifier le solde des 3 derniers jours, en commençant par hier
            for i in range(1, 4):
                solde_temp = SoldeJour.objects.filter(date_saisie=date_bilan - timedelta(days=i)).first()
                if solde_temp:
                    solde_initial = solde_temp.montant
                    break
                
            # Si aucun solde trouvé dans les 3 derniers jours, utiliser le premier solde initial enregistré
            if not solde_initial:
                solde_initial_record = SoldeJour.objects.filter(date_saisie=date_bilan - timedelta(days=1)).first()
                solde_initial = solde_initial_record.montant if solde_initial_record else 0
                
            som_entree = tot_entree + solde_initial
            solde_fin_journee = som_entree - tot_sortie
            
            som_billets = Billetage.objects.filter(type='Billet', date_saisie=date_bilan).aggregate(somme=Sum('valeur'))['somme'] or 0
            
            som_pieces = Billetage.objects.filter(type='Pièce', date_saisie=date_bilan).aggregate(somme=Sum('valeur'))['somme'] or 0
            som_billets = som_billets
            som_pieces = som_pieces

            bill = Billetage.objects.filter(type='Billet', date_saisie=date_bilan)
            bil = []
            som_tot_bil = 0
            for b in bill:
                val = b.valeur
                nb = b.nombre
                res = b.valeur * b.nombre
                som_tot_bil += res or 0
                bil.append({'val': val, 'nb':nb,'res':res})
            som_tot_bil = som_tot_bil
            bil = bil

            piece = Billetage.objects.filter(type='Pièce', date_saisie=date_bilan)
            pie = []
            som_tot_piec = 0
            for p in piece:
                val = p.valeur or 0
                nb = p.nombre or 0
                res = p.valeur * p.nombre
                som_tot_piec += res or 0
                pie.append({'val': val, 'nb':nb,'res':res})
            som_tot_piec = som_tot_piec
            pie = pie
            
            Total_piec_bill = som_tot_piec + som_tot_bil
            ecart = solde_fin_journee - Total_piec_bill
            self.request.user.save()
            
        else:
            encaisser = Encaissement.objects.filter(date_saisie=date.today())
            decaisser = Decaissement.objects.filter(date_saisie=date.today())
            solde_jour = SoldeJour.objects.filter(date_saisie=date.today()).aggregate(Sum('montant'))['montant__sum'] or 0
            # Calculer le total des encaissements pour la journée actuelle
            tot_entree = Encaissement.objects.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
            tot_sortie = Decaissement.objects.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
            
            solde_initial = None

            # Vérifier le solde des 3 derniers jours, en commençant par hier
            for i in range(1, 4):
                solde_temp = SoldeJour.objects.filter(date_saisie=date.today() - timedelta(days=i)).first()
                if solde_temp:
                    solde_initial = solde_temp.montant
                    break
            # Si aucun solde trouvé dans les 3 derniers jours, utiliser le premier solde initial enregistré
            if not solde_initial:
                solde_initial_record = SoldeJour.objects.filter(date_saisie=date.today() - timedelta(days=1)).first()
                solde_initial = solde_initial_record.montant if solde_initial_record else 0
            som_entree = tot_entree + solde_initial
            solde_fin_journee = som_entree - tot_sortie
            
            som_billets = Billetage.objects.filter(type='Billet', date_saisie=date.today()).aggregate(somme=Sum('valeur'))['somme'] or 0
            nb_bill = Billetage.objects.filter(type='Billet', date_saisie=date.today()).count() 

            som_pieces = Billetage.objects.filter(type='Pièce', date_saisie=date.today()).aggregate(somme=Sum('valeur'))['somme'] or 0
            som_billets = som_billets
            som_pieces = som_pieces

            bill = Billetage.objects.filter(type='Billet', date_saisie=date.today())
            bil = []
            som_tot_bil = 0
            for b in bill:
                val = b.valeur
                nb = b.nombre
                res = b.valeur * b.nombre
                som_tot_bil += res or 0
                bil.append({'val': val, 'nb':nb,'res':res})
            som_tot_bil = som_tot_bil
            bil = bil

            piece = Billetage.objects.filter(type='Pièce', date_saisie=date.today())
            pie = []
            som_tot_piec = 0
            for p in piece:
                val = p.valeur or 0
                nb = p.nombre or 0
                res = p.valeur * p.nombre
                som_tot_piec += res or 0
                pie.append({'val': val, 'nb':nb,'res':res})
            som_tot_piec = som_tot_piec
            pie = pie
            
            Total_piec_bill = som_tot_piec + som_tot_bil
            ecart = solde_fin_journee - Total_piec_bill

            self.request.user.save()
            
        context={
            'forms':forms,
            'formset':formset,
            'solde_initial':solde_initial,
            
            'ent':encaisser,
            'sort':decaisser,
            'ecart':ecart,
            
            'piecs':pie,
            'biels':bil,
            
            'sompiecs':som_tot_piec,
            'sombiels':som_tot_bil ,
            
            'tot_ent':tot_entree,
            'tot_sor':tot_sortie,
            'form':form,
            
            'solde_day':solde_jour,
            'tot_bi_pi':Total_piec_bill,
            'dates':dates,
        }
        return context
    def post(self, request, *args, **kwargs):
        formset = BilletageFormSet(self.request.POST, queryset=Billetage.objects.none())
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.auteur = request.user
                instance.save()
            messages.success(request, self.success_message)
            return redirect(self.success_url)
        else:
            messages.error(request, self.error_message)
        return self.render_to_response(self.get_context_data(formset=formset))

class ExportCaisseExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Récupérer la date depuis les paramètres GET
        date_bilan = request.GET.get('date_bilan')
        
        # Si aucune date n'est fournie, utiliser la date du jour
        if date_bilan:
            try:
                date_bilan = datetime.strptime(date_bilan, "%Y-%m-%d").date()
            except ValueError:
                date_bilan = date.today()
        else:
            date_bilan = date.today()
        
        # Récupérer les encaissements
        encaisser = Encaissement.objects.filter(date_saisie=date_bilan).select_related('auteur')
        tot_entree = encaisser.aggregate(somme=Sum('montant'))['somme'] or 0
        
        # Récupérer les décaissements
        decaisser = Decaissement.objects.filter(date_saisie=date_bilan).select_related('auteur')
        tot_sortie = decaisser.aggregate(somme=Sum('montant'))['somme'] or 0
        
        # Récupérer le solde du jour
        solde_jour = SoldeJour.objects.filter(date_saisie=date_bilan).aggregate(Sum('montant'))['montant__sum'] or 0
        
        # Calculer le solde initial (même logique que dans BilletageView)
        solde_initial = None
        for i in range(1, 4):
            solde_temp = SoldeJour.objects.filter(date_saisie=date_bilan - timedelta(days=i)).first()
            if solde_temp:
                solde_initial = solde_temp.montant
                break
        
        if not solde_initial:
            solde_initial_record = SoldeJour.objects.filter(date_saisie=date_bilan - timedelta(days=1)).first()
            solde_initial = solde_initial_record.montant if solde_initial_record else 0
        
        som_entree = tot_entree + solde_initial
        solde_fin_journee = som_entree - tot_sortie
        
        # Récupérer le billetage
        bill = Billetage.objects.filter(type='Billet', date_saisie=date_bilan)
        piece = Billetage.objects.filter(type='Pièce', date_saisie=date_bilan)
        
        som_tot_bil = sum(b.valeur * b.nombre for b in bill)
        som_tot_piec = sum(p.valeur * p.nombre for p in piece)
        Total_piec_bill = som_tot_piec + som_tot_bil
        ecart = solde_fin_journee - Total_piec_bill
        
        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # ========== ONGLET 1: RÉSUMÉ ==========
        ws_resume = wb.create_sheet("Résumé", 0)
        ws_resume.append(["SITUATION DE LA CAISSE DU " + date_bilan.strftime("%d-%m-%Y")])
        ws_resume.append([])
        
        # Informations générales
        ws_resume.append(["Solde initial de début de journée", solde_initial or 0])
        ws_resume.append(["Total des entrées", tot_entree])
        ws_resume.append(["Total des sorties", tot_sortie])
        ws_resume.append(["Solde de fin de journée", solde_fin_journee])
        ws_resume.append(["Solde du jour (enregistré)", solde_jour])
        ws_resume.append([])
        
        # Billetage
        ws_resume.append(["BILLETAGE"])
        ws_resume.append(["Total Billets", som_tot_bil])
        ws_resume.append(["Total Pièces", som_tot_piec])
        ws_resume.append(["Total Billets + Pièces", Total_piec_bill])
        ws_resume.append(["Écart", ecart])
        
        # Ajuster la largeur des colonnes
        ws_resume.column_dimensions['A'].width = 35
        ws_resume.column_dimensions['B'].width = 20
        
        # ========== ONGLET 2: ENCAISSEMENTS ==========
        ws_encaissements = wb.create_sheet("Encaissements")
        headers_enc = ["Date", "Numéro Pièce", "Libellé", "Montant", "Auteur"]
        ws_encaissements.append(headers_enc)
        
        for e in encaisser:
            ws_encaissements.append([
                e.date_saisie.strftime("%d-%m-%Y") if e.date_saisie else "",
                e.Num_piece or "",
                e.libelle or "",
                e.montant,
                e.auteur.username if e.auteur else "N/A"
            ])
        
        # Ligne de total
        ws_encaissements.append([])
        ws_encaissements.append(["TOTAL", "", "", tot_entree, ""])
        
        # Ajuster la largeur des colonnes
        for col_num, column_title in enumerate(headers_enc, 1):
            col_letter = get_column_letter(col_num)
            ws_encaissements.column_dimensions[col_letter].width = 22
        
        # ========== ONGLET 3: DÉCAISSEMENTS ==========
        ws_decaissements = wb.create_sheet("Décaissements")
        headers_dec = ["Date", "Numéro Pièce", "Libellé", "Montant", "Auteur"]
        ws_decaissements.append(headers_dec)
        
        for d in decaisser:
            ws_decaissements.append([
                d.date_saisie.strftime("%d-%m-%Y") if d.date_saisie else "",
                d.Num_piece or "",
                d.libelle or "",
                d.montant,
                d.auteur.username if d.auteur else "N/A"
            ])
        
        # Ligne de total
        ws_decaissements.append([])
        ws_decaissements.append(["TOTAL", "", "", tot_sortie, ""])
        
        # Ajuster la largeur des colonnes
        for col_num, column_title in enumerate(headers_dec, 1):
            col_letter = get_column_letter(col_num)
            ws_decaissements.column_dimensions[col_letter].width = 22
        
        # ========== ONGLET 4: BILLETAGE ==========
        ws_billetage = wb.create_sheet("Billetage")
        
        # Section Billets
        ws_billetage.append(["BILLETS"])
        headers_bil = ["Valeur", "Nombre", "Total"]
        ws_billetage.append(headers_bil)
        
        for b in bill:
            ws_billetage.append([
                b.valeur,
                b.nombre,
                b.valeur * b.nombre
            ])
        
        ws_billetage.append(["TOTAL BILLETS", "", som_tot_bil])
        ws_billetage.append([])
        
        # Section Pièces
        ws_billetage.append(["PIÈCES"])
        headers_piec = ["Valeur", "Nombre", "Total"]
        ws_billetage.append(headers_piec)
        
        for p in piece:
            ws_billetage.append([
                p.valeur,
                p.nombre,
                p.valeur * p.nombre
            ])
        
        ws_billetage.append(["TOTAL PIÈCES", "", som_tot_piec])
        ws_billetage.append([])
        ws_billetage.append(["TOTAL BILLETS + PIÈCES", "", Total_piec_bill])
        ws_billetage.append(["ÉCART", "", ecart])
        
        # Ajuster la largeur des colonnes
        ws_billetage.column_dimensions['A'].width = 25
        ws_billetage.column_dimensions['B'].width = 15
        ws_billetage.column_dimensions['C'].width = 20
        
        # Préparer la réponse HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Point-de-caisse-{date_bilan.strftime('%d-%m-%Y')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

class AddDecaissementView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_decaisse'
    model = Decaissement
    form_class = DecaissementForm
    template_name = 'perfect/sortie_caiss.html'
    success_message = 'Sortie de caisse enregistrée avec succès✓✓'
    error_message = "Erreur de saisie ✘✘"
    timeout_minutes = 500
    success_url = reverse_lazy ('add_decaisse')
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        reponse =  super().form_valid(form)
        messages.success(self.request, self.success_message)
        return reponse
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        forms = self.get_form()
        annee_en_cours =date.today().year
        today = date.today()
        mois_en_cours =date.today().month
        date_debut = date_fin = None
        if self.request.POST:
            formset = DecaissementFormSet(self.request.POST, queryset=Decaissement.objects.none())
        else:
            formset = DecaissementFormSet(queryset=Decaissement.objects.none())
        sorties_queryset = Decaissement.objects.all()
        form = DateForm(self.request.GET)
        if form.is_valid():
            date_debut = form.cleaned_data['date_debut'] 
            date_fin = form.cleaned_data['date_fin'] 
            if date_debut and date_fin:
                sorties_queryset = sorties_queryset.filter(date_saisie__range=[date_debut, date_fin])

        if date_debut and date_fin:
            filtre_sorties = sorties_queryset
        else:
            filtre_sorties = sorties_queryset.filter(date_saisie__month=date.today().month)
        
        sorties_liste = filtre_sorties.order_by('-id')
        tot_sort_jour = sorties_queryset.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
        tot_sort_mois =  sorties_queryset.filter(date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
        tot_sort_annuel = sorties_queryset.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 0

        sorti_jours_format = '{:,}'.format(tot_sort_jour).replace(',', ' ')
        sorti_mois_format = '{:,}'.format(tot_sort_mois).replace(',', ' ')
        sorti_an_format = '{:,}'.format(tot_sort_annuel).replace(',', ' ')
        context = {
            'form': form,
            'forms': forms,
            'formset': formset,
            'tot_sort_jours': sorti_jours_format,
            'tot_sort_mois': sorti_mois_format,
            'tot_sort_annuel': sorti_an_format,
            'sorties_liste': sorties_liste,
            'dates': today,
        }
        return context
    # def get_success_url(self):
    #     return reverse('add_decaisse')
    def post(self, request, *args, **kwargs):
        formset = DecaissementFormSet(self.request.POST, queryset=Decaissement.objects.none())
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.auteur = request.user
                instance.save()
            messages.success(request, self.success_message)
            return redirect(self.success_url)
        else:
            messages.error(request, self.error_message)
        return self.render_to_response(self.get_context_data(formset=formset))

class ExportDecaissementExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        decaissements = Decaissement.objects.all().select_related('auteur')

        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        num_piece = request.GET.get('num_piece')
        libelle = request.GET.get('libelle')
        auteur = request.GET.get('auteur')
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                decaissements = decaissements.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pass
        if num_piece:
            decaissements = decaissements.filter(Num_piece__icontains=num_piece)
        if libelle:
            decaissements = decaissements.filter(libelle__icontains=libelle)
        if auteur:
            decaissements = decaissements.filter(auteur__username__icontains=auteur)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sortie de caisse"
        headers = [
            "Numéro Pièce",
            "Libellé",
            "Montant",
            "Date Saisie",
            "Auteur"
        ]
        ws.append(headers)
        for e in decaissements:
            ws.append([
                e.Num_piece or "",
                e.libelle or "",
                e.montant,
                e.date_saisie.strftime("%d-%m-%Y") if e.date_saisie else "",
                e.auteur.username if e.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Sortie-caisse.xlsx'
        wb.save(response)
        return response

class UpdatDecaissementView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_decaisse'
    model = Decaissement
    form_class = UpdatDecaissementForm
    template_name = "perfect/partials/sorti_form.html"
    success_url = reverse_lazy('add_decaisse')
    success_message = 'Sortir de caisse modifiée avec succès✓✓'
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class UpdatEncaissementView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_encaisse'
    model = Encaissement
    form_class = UpdatEncaissementForm
    template_name = "perfect/partials/encais_form.html"
    success_url = reverse_lazy('addencaisse')
    success_message = 'Entrée de caisse modifiée avec succès✓✓'
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

def delete_sortie_caisse(request, pk):
    try:
        decaissements = get_object_or_404(Decaissement, id=pk)
        decaissements.delete()
        messages.success(request, f"la sortie de caisse de {decaissements.Num_piece} - {decaissements.libelle} a été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('add_decaisse')

class AddEncaissementView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'addencaisse'
    model = Encaissement
    form_class = EncaissementForm
    template_name = 'perfect/entre_caisse.html'
    success_message = 'Entrée de caisse enregistrée avec succès✓✓'
    error_message = "Erreur de saisie ✘✘"
    success_url = reverse_lazy ('addencaisse')
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        reponse =  super().form_valid(form)
        messages.success(self.request, self.success_message)
        return reponse
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        forms = self.get_form()
        annee_en_cours =date.today().year
        today =date.today()
        mois_en_cours =date.today().month
        libelle_mois_en_cours = calendar.month_name[mois_en_cours]
        date_debut = date_fin = None
        if self.request.POST:
            formset = EncaissementFormSet(self.request.POST, queryset=Encaissement.objects.none())
        else:
            formset = EncaissementFormSet(queryset=Encaissement.objects.none())
        entres_queryset = Encaissement.objects.all()
        form = DateForm(self.request.GET)
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
        # Appliquer le filtre par date : si dates fournies, les utiliser, sinon filtrer par mois en cours
        if date_debut and date_fin:
            filtre_entres = entres_queryset.filter(date_saisie__range=[date_debut, date_fin])
        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_entres = entres_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours).order_by('-date_saisie')
        
        liste_entret = filtre_entres.order_by('-id')
        # Calculer les totaux du jour, mois et année en cours (toujours sur toutes les données, pas filtrées)
        tot_entre_jour = Encaissement.objects.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
        tot_entre_mois = Encaissement.objects.filter(date_saisie__month=date.today().month, date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 0
        tot_entre_annuel = Encaissement.objects.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 0
        
        # Formater les valeurs pour l'affichage
        entret_jours_format = '{:,}'.format(tot_entre_jour).replace(',', ' ')
        entret_mois_format = '{:,}'.format(tot_entre_mois).replace(',', ' ')
        entret_an_format = '{:,}'.format(tot_entre_annuel).replace(',', ' ')
        
        context = {
            'form': form,
            'formset': formset,
            'entret_jours_format': entret_jours_format,
            'entret_mois_format': entret_mois_format,
            'entret_an_format': entret_an_format,
            'dates': today,
            'forms': forms,
            'entres_liste': liste_entret,
            # 'liste_entret': liste_entret,
        }
        return context
    def post(self, request, *args, **kwargs):
        formset = EncaissementFormSet(self.request.POST, queryset=Encaissement.objects.none())
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.auteur = request.user
                instance.save()
            messages.success(request, self.success_message)
            return redirect(self.success_url)
        else:
            messages.error(request, self.error_message)
        return self.render_to_response(self.get_context_data(formset=formset))

class ExportEncaissementExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        encaissements = Encaissement.objects.all().select_related('auteur')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        num_piece = request.GET.get('num_piece')
        libelle = request.GET.get('libelle')
        auteur = request.GET.get('auteur')
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                encaissements = encaissements.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pass
        if num_piece:
            encaissements = encaissements.filter(Num_piece__icontains=num_piece)
        if libelle:
            encaissements = encaissements.filter(libelle__icontains=libelle)
        if auteur:
            encaissements = encaissements.filter(auteur__username__icontains=auteur)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entrée caisse"
        headers = [
            "Numéro Pièce",
            "Libellé",
            "Montant",
            "Date Saisie",
            "Auteur"
        ]
        ws.append(headers)
        for e in encaissements:
            ws.append([
                e.Num_piece or "",
                e.libelle or "",
                e.montant,
                e.date_saisie.strftime("%d-%m-%Y") if e.date_saisie else "",
                e.auteur.username if e.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Entrée-caisse.xlsx'
        wb.save(response)
        return response

def delete_entre_caisse(request, pk):
    try:
        encaissements = get_object_or_404(Encaissement, id=pk)
        encaissements.delete()
        messages.success(request, f"l'entrée de caisse de {encaissements.Num_piece} - {encaissements.libelle} a été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('addencaisse')

class AddSoldeJourView(CustomPermissionRequiredMixin, CreateView):
    permission_url = 'add_solde'
    model = SoldeJour
    form_class = Solde_JourForm
    template_name = 'perfect/solde.html'
    success_message = 'le solde de la journée a été enregistré avec succès✓✓'
    error_message = "Un solde existe deja pour cette journée ✘✘ "
    success_url = reverse_lazy ('add_solde')
    timeout_minutes = 200
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        # Vérifie si un solde existe déjà pour la date spécifiée
        date = form.cleaned_data['date_saisie']
        form.instance.auteur = self.request.user
        messages.success(self.request, self.success_message)
        solde_exist = SoldeJour.objects.filter(date_saisie=date).exists()
        if solde_exist:
            # Si un solde existe déjà pour cette date, renvoie une erreur
            form.add_error('date', 'Un solde existe déjà pour cette date.')
            return self.form_invalid(form)
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        forms = self.get_form()
        form = DateForm(self.request.GET)
        if form.is_valid():
            date_debut = form.cleaned_data['date_debut'] 
            date_fin = form.cleaned_data['date_fin'] 
            solde_liste = SoldeJour.objects.filter(date_saisie__range=[date_debut, date_fin])
        else:
            solde_liste = SoldeJour.objects.filter(date_saisie__month=date.today().month)
        solde_jour = solde_liste.order_by('-id')
        context = {
            'form': form,
            'soldes_liste': solde_jour,
            'forms': forms,
        }
        return context

class ExportSoldeJourExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        soldes = SoldeJour.objects.all()

        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        auteur = request.GET.get('auteur')
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                solde = solde.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pass
        if auteur:
            soldes = soldes.filter(auteur__username__icontains=auteur)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Solde journalier"
        headers = [
            "Montant",
            "Date Saisie",
            "Auteur"
        ]
        ws.append(headers)
        for e in soldes:
            ws.append([
                e.montant,
                e.date_saisie.strftime("%d-%m-%Y") if e.date_saisie else "",
                e.auteur.username if e.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Solde-journalier.xlsx'
        wb.save(response)
        return response

def delete_solde(request, pk):
    try:
        solde = get_object_or_404(SoldeJour, id=pk)
        solde.delete()
        messages.success(request, f"le solde de la journee du {solde.date_saisie} - {solde.montant} frscfa a été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('add_solde')
     
class GestionalerteView(LoginRequiredMixin, CustomPermissionRequiredMixin, TemplateView):  
    permission_url = 'alerte'  
    login_url = 'login'                                                                      
    template_name = 'perfect/alerte.html'
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self,*args, **kwargs):  
        context = super().get_context_data(*args,**kwargs)  
        dates = date.today()
        annee = date.today().year
        now = datetime.now()
        mois_en_cours =date.today().month
        libelle_mois_en_cours = calendar.month_name[mois_en_cours]
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant, car_statut=True)
                else:
                    vehicules = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.filter(car_statut=True)
        resultat_vehicule = []
        alert_color = " "
        forms = DateForm(self.request.GET)
        if forms.is_valid():
            date_debut = forms.cleaned_data['date_debut'] 
            date_fin = forms.cleaned_data['date_fin']

            assu_all = Assurance.objects.filter(date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 0
            vign_all = Vignette.objects.filter(date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 0
            patente_all = Patente.objects.filter(date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 0
            stat_all = Stationnement.objects.filter(date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 0
            entretiens = Entretien.objects.filter(Q(date_saisie__lte=now) & Q(date_proch__gte=now)).count()
            visites = VisiteTechnique.objects.filter(Q(date_saisie__lte=now) & Q(date_proch__gte=now)).count()
            assurances = Assurance.objects.filter(Q(date_saisie__lte=now) & Q(date_proch__gte=now)).count()
            vignette = Vignette.objects.filter(Q(date__lte=now) & Q(date_proch__gte=now)).count()                
            patente = Patente.objects.filter(Q(date__lte=now) & Q(date_proch__gte=now)).count()                 
            for vehicule in vehicules:
                visite = VisiteTechnique.objects.filter(vehicule = vehicule).order_by('date_saisie')
                entretien = Entretien.objects.filter(vehicule = vehicule).order_by('date_saisie')
                assurances = Assurance.objects.filter(vehicule = vehicule).order_by('date_saisie')
                vignettes = Vignette.objects.filter(vehicule = vehicule).order_by('date_saisie')
                patentes = Patente.objects.filter(vehicule = vehicule).order_by('date_saisie')
                cartstaions = Stationnement.objects.filter(vehicule = vehicule).order_by('date_saisie')

                jours_cartsta_restant = cartstaions
                if jours_cartsta_restant:
                    for cartstaion in jours_cartsta_restant:
                        jours_cartsta_restant = cartstaion.jours_cartsta_restant
                        alert_cartsta_color = 'danger' if jours_cartsta_restant <=10 else 'warning' if jours_cartsta_restant < 30 else 'success'
                else: 
                    jours_cartsta_restant = "0"    
                    alert_cartsta_color = "info"
                # 
                jours_pate_restant = patentes
                if jours_pate_restant:
                    for patente in jours_pate_restant:
                        jours_pate_restant = patente.jours_pate_restant
                        alert_pate_color = 'danger' if jours_pate_restant <=10 else 'warning' if jours_pate_restant < 30 else 'success'
                else: 
                    jours_pate_restant = "0"    
                    alert_pate_color = "info"
                    # 
                jours_vign_restant = vignettes
                if jours_vign_restant:
                    for vignette in jours_vign_restant:
                        jours_vign_restant = vignette.jours_vign_restant
                        alert_vign_color = 'danger' if jours_vign_restant <=10 else 'warning' if jours_vign_restant < 334 else 'success'
                else: 
                    jours_vign_restant = "0"    
                    alert_vign_color = "info"
                    # 
                jours_assu_restant = assurances
                if jours_assu_restant:
                    for assurance in jours_assu_restant:
                        jours_assu_restant = assurance.jours_assu_restant
                        alert_assu_color = 'danger' if jours_assu_restant <=7 else 'warning' if jours_assu_restant < 15 else 'success'
                else: 
                    jours_assu_restant = "0"    
                    alert_assu_color = "info"
                    # 
                jours_ent_restant = entretien
                if jours_ent_restant:
                    for entretien in jours_ent_restant:
                        jours_ent_restant = entretien.jours_ent_restant
                        alert_ent_color = 'danger' if jours_ent_restant <=3 else 'warning' if jours_ent_restant < 8 else 'success'
                else: 
                    jours_ent_restant = "0"     
                    alert_ent_color = "info"    
                jours_restant = visite  
                # 
                if jours_restant:       
                    for visit in jours_restant:         
                        jours_restant = visit.jour_restant      
                        alert_color = 'danger' if jours_restant <=32 else 'warning' if jours_restant <92 else 'success'
                else: 
                    jours_restant = "0"    
                    alert_color = "info"
                # Calculer les alertes critiques
                alert_types = []
                if isinstance(jours_restant, int) and 1 <= jours_restant <= 32:
                    alert_types.append("visite technique")
                if isinstance(jours_ent_restant, int) and 1 <= jours_ent_restant <= 3:
                    alert_types.append("entretien")
                if isinstance(jours_assu_restant, int) and 1 <= jours_assu_restant <= 7:
                    alert_types.append("assurance")
                if isinstance(jours_vign_restant, int) and 1 <= jours_vign_restant <= 10:
                    alert_types.append("vignette")
                if isinstance(jours_pate_restant, int) and 1 <= jours_pate_restant <= 10:
                    alert_types.append("patente")
                if isinstance(jours_cartsta_restant, int) and 1 <= jours_cartsta_restant <= 10:
                    alert_types.append("stationnement")
                
                resultat_vehicule.append({
                    'vehicule': vehicule, 
                    'jours_restant': jours_restant, 
                    'alert_color': alert_color, 
                    'alert_ent_color': alert_ent_color, 
                    'jours_ent_restant': jours_ent_restant,
                    'alert_assu_color': alert_assu_color,
                    'jours_assu_restant': jours_assu_restant,
                    'jours_vign_restant': jours_vign_restant,
                    'alert_vign_color': alert_vign_color, 
                    'jours_pate_restant': jours_pate_restant,
                    'alert_pate_color': alert_pate_color, 
                    'jours_cartsta_restant': jours_cartsta_restant,
                    'alert_cartsta_color': alert_cartsta_color,
                    'alert_types': alert_types
                })
        else:
            assu_all = Assurance.objects.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
            vign_all = Vignette.objects.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
            patente_all = Patente.objects.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
            stat_all = Stationnement.objects.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
            entretiens = Entretien.objects.filter(Q(date_saisie__lte=now) & Q(date_proch__gte=now)).count()
            visites = VisiteTechnique.objects.filter(Q(date_saisie__lte=now) & Q(date_proch__gte=now)).count()
            assurances = Assurance.objects.filter(Q(date_saisie__lte=now) & Q(date_proch__gte=now)).count()
            vignette = Vignette.objects.filter(Q(date__lte=now) & Q(date_proch__gte=now)).count()                
            patente = Patente.objects.filter(Q(date__lte=now) & Q(date_proch__gte=now)).count()                 
            # -------------------------#-*-#------------------------- #
            for vehicule in vehicules:
                visite = VisiteTechnique.objects.filter(vehicule = vehicule).order_by('date_saisie')
                entretien = Entretien.objects.filter(vehicule = vehicule).order_by('date_saisie')
                assurances = Assurance.objects.filter(vehicule = vehicule).order_by('date_saisie')
                vignettes = Vignette.objects.filter(vehicule = vehicule).order_by('date_saisie')
                patentes = Patente.objects.filter(vehicule = vehicule).order_by('date_saisie')
                cartstaions = Stationnement.objects.filter(vehicule = vehicule).order_by('date_saisie')
                jours_cartsta_restant = cartstaions
                if jours_cartsta_restant:
                    for cartstaion in jours_cartsta_restant:
                        jours_cartsta_restant = cartstaion.jours_cartsta_restant
                        alert_cartsta_color = 'danger' if jours_cartsta_restant <=10 else 'warning' if jours_cartsta_restant < 30 else 'success'
                else: 
                    jours_cartsta_restant = "0"    
                    alert_cartsta_color = "info"
                # 
                jours_pate_restant = patentes
                if jours_pate_restant:
                    for patente in jours_pate_restant:
                        jours_pate_restant = patente.jours_pate_restant
                        alert_pate_color = 'danger' if jours_pate_restant <=10 else 'warning' if jours_pate_restant < 30 else 'success'
                else: 
                    jours_pate_restant = "0"    
                    alert_pate_color = "info"
                    # 
                jours_vign_restant = vignettes
                if jours_vign_restant:
                    for vignette in jours_vign_restant:
                        jours_vign_restant = vignette.jours_vign_restant
                        alert_vign_color = 'danger' if jours_vign_restant <=10 else 'warning' if jours_vign_restant < 334 else 'success'
                else: 
                    jours_vign_restant = "0"    
                    alert_vign_color = "info"
                    # 
                jours_assu_restant = assurances
                if jours_assu_restant:
                    for assurance in jours_assu_restant:
                        jours_assu_restant = assurance.jours_assu_restant
                        alert_assu_color = 'danger' if jours_assu_restant <=7 else 'warning' if jours_assu_restant < 15 else 'success'
                else: 
                    jours_assu_restant = "0"    
                    alert_assu_color = "info"
                    # 
                jours_ent_restant = entretien
                if jours_ent_restant:
                    for entretien in jours_ent_restant:
                        jours_ent_restant = entretien.jours_ent_restant
                        alert_ent_color = 'danger' if jours_ent_restant <=3 else 'warning' if jours_ent_restant < 8 else 'success'
                else: 
                    jours_ent_restant = "0"     
                    alert_ent_color = "info"    
                jours_restant = visite  
                # 
                if jours_restant:       
                    for visit in jours_restant:         
                        jours_restant = visit.jour_restant      
                        alert_color = 'danger' if jours_restant <=32 else 'warning' if jours_restant <92 else 'success'
                else: 
                    jours_restant = "0"    
                    alert_color = "info"  
                # Calculer les alertes critiques
                alert_types = []
                if isinstance(jours_restant, int) and 1 <= jours_restant <= 32:
                    alert_types.append("visite technique")
                if isinstance(jours_ent_restant, int) and 1 <= jours_ent_restant <= 3:
                    alert_types.append("entretien")
                if isinstance(jours_assu_restant, int) and 1 <= jours_assu_restant <= 7:
                    alert_types.append("assurance")
                if isinstance(jours_vign_restant, int) and 1 <= jours_vign_restant <= 10:
                    alert_types.append("vignette")
                if isinstance(jours_pate_restant, int) and 1 <= jours_pate_restant <= 10:
                    alert_types.append("patente")
                if isinstance(jours_cartsta_restant, int) and 1 <= jours_cartsta_restant <= 10:
                    alert_types.append("stationnement")
                resultat_vehicule.append({
                    'vehicule': vehicule, 
                    'jours_restant': jours_restant, 
                    'alert_color': alert_color, 
                    'alert_ent_color': alert_ent_color, 
                    'jours_ent_restant': jours_ent_restant,
                    'alert_assu_color': alert_assu_color,
                    'jours_assu_restant': jours_assu_restant,
                    'jours_vign_restant': jours_vign_restant,
                    'alert_vign_color': alert_vign_color, 
                    'jours_pate_restant': jours_pate_restant,
                    'alert_pate_color': alert_pate_color, 
                    'jours_cartsta_restant': jours_cartsta_restant,
                    'alert_cartsta_color': alert_cartsta_color,
                    'alert_types': alert_types
                })
        context={
            'dates':dates,
            'vehicules':vehicules,
            'resultat_vehicule':resultat_vehicule,
            'annees':annee,
            'visites':visites,
            'assu_all':assu_all,
            'vign_all':vign_all,
            'pat_all':patente_all,
            'stat_all':stat_all,
            'entr_all':entretiens,
            'form':forms,
            'mois_en_cours':libelle_mois_en_cours,
        }
        return context 

class AddCategoriVehi(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):  
    login_url = 'login'
    permission_url = 'add_catego_vehi'
    model = CategoVehi      
    form_class = CategorieForm      
    template_name = 'perfect/add_categorie.html'
    success_message = 'Categorie enregistré avec succès✓✓'
    error_message = "Erreur de saisie, cette categorie ou cet identifiant existe✘✘"
    success_url= reverse_lazy('add_catego_vehi')
    timeout_minutes = 120
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        messages.success(self.request,self.success_message)
        return super().form_valid(form)
    def form_invalid(self, form):
        messages.success(self.request,self.error_message)
        return super().form_invalid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        catego_vehi = CategoVehi.objects.all()
        categories = catego_vehi.annotate(
            nb_vehicules=Count('catego_vehicule', filter=Q(catego_vehicule__car_statut=True))
        ).order_by('id')
        total_veh = Vehicule.objects.filter(car_statut=True).order_by('id').count()
        forms = self.get_form()
        context = {
            'form':forms,
            'catego_vehi':catego_vehi,
            'categories':categories,
            'total_veh':total_veh,
        }
        return context

class UpdateCategoView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_catego_vehi'
    model = CategoVehi
    form_class = CategorieForm
    template_name = "perfect/partials/categ_form.html"
    success_url = reverse_lazy('add_catego_vehi')
    success_message = 'Categorie modifiée avec succès✓✓'
    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

def delete_catego(request, pk):
    try:
        catego = get_object_or_404(CategoVehi, id=pk)
        catego.delete()
        messages.success(request, f"la categorie de véhicule {catego.category} a été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('add_catego_vehi')
    
def CategoVehiculeListView(request, cid):
    categorys= CategoVehi.objects.get(cid=cid)
    cars = Vehicule.objects.filter(category=categorys, car_statut=True)
    user_group = request.user.groups.first()
    user_group = user_group.name if user_group else None
    context = {
        'user_group':user_group,
        'categorys':categorys,
        'cars':cars,
    }
    return render(request, 'news/applist/list_vehi_categor.html',context)

class AllVehiculeView(LoginRequiredMixin, CustomPermissionRequiredMixin, View):
    login_url = 'login'
    permission_url = 'all_vehi'
    template_name = 'perfect/add_vehicule.html'
    def get(self, request, pk=None):
        user = request.user
        if user.user_type == "4":
            try:
                gerant = Gerant.objects.get(user=user)
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():
                    cars = Vehicule.objects.filter(category__in=categories_gerant, car_statut=True)
                else:
                    cars = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                cars = Vehicule.objects.none()
        else:
            cars = Vehicule.objects.filter(car_statut=True)
        categories = CategoVehi.objects.annotate(
            nb_vehicules=Count('catego_vehicule', filter=Q(catego_vehicule__car_statut=True))
        ).order_by('id')
        cout_totals = 0
        total_veh = 0
        categorie = None
        if pk:
            categorie = get_object_or_404(CategoVehi, pk=pk)
            cars = cars.filter(category=categorie)
        cout_totals = cars.aggregate(total=Sum('cout_acquisition'))['total'] or 0
        cout_total = '{:,}'.format(cout_totals).replace(',', ' ')
        total_veh = cars.count()
        # Véhicules hors parc (même filtres catégorie / gérant) pour l’affichage du compteur
        cars_hors_parc = Vehicule.objects.filter(car_statut=False)
        if user.user_type == "4":
            try:
                gerant = Gerant.objects.get(user=user)
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():
                    cars_hors_parc = cars_hors_parc.filter(category__in=categories_gerant)
                else:
                    cars_hors_parc = cars_hors_parc.none()
            except Gerant.DoesNotExist:
                cars_hors_parc = cars_hors_parc.none()
        if pk and categorie:
            cars_hors_parc = cars_hors_parc.filter(category=categorie)
        total_veh_hors_parc = cars_hors_parc.count()
        return render(request, self.template_name, {
            'vehicules': cars,
            'categorie': categorie,
            'categories': categories,
            'cout_total': cout_total,
            'total_veh': total_veh,
            'total_veh_hors_parc': total_veh_hors_parc
        })

class AllVehiculeHorsParrcView(LoginRequiredMixin, CustomPermissionRequiredMixin, View):
    login_url = 'login'
    permission_url = 'all_vehi'
    template_name = 'perfect/vehi_hors_parc.html'
    def get(self, request, pk=None):
        user = request.user
        # Vue « hors parc » : n’afficher que les véhicules avec car_statut=False
        if user.user_type == "4":
            try:
                gerant = Gerant.objects.get(user=user)
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():
                    cars = Vehicule.objects.filter(category__in=categories_gerant, car_statut=False)
                else:
                    cars = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                cars = Vehicule.objects.none()
        else:
            cars = Vehicule.objects.filter(car_statut=False)
        categories = CategoVehi.objects.annotate(
            nb_vehicules=Count('catego_vehicule', filter=Q(catego_vehicule__car_statut=True))
        ).order_by('id')
        cout_totals = 0
        total_veh = 0
        categorie = None
        if pk:
            categorie = get_object_or_404(CategoVehi, pk=pk)
            cars = cars.filter(category=categorie)
        cout_totals = cars.aggregate(total=Sum('cout_acquisition'))['total'] or 0
        cout_total = '{:,}'.format(cout_totals).replace(',', ' ')
        # Nombre de véhicules au parc (car_statut=True) avec le même périmètre que la liste hors parc
        cars_in_parc = Vehicule.objects.filter(car_statut=True)
        if user.user_type == "4":
            try:
                gerant = Gerant.objects.get(user=user)
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():
                    cars_in_parc = cars_in_parc.filter(category__in=categories_gerant)
                else:
                    cars_in_parc = cars_in_parc.none()
            except Gerant.DoesNotExist:
                cars_in_parc = cars_in_parc.none()
        if pk and categorie:
            cars_in_parc = cars_in_parc.filter(category=categorie)
        total_veh = cars_in_parc.count()
        total_veh_hors_parc = cars.count()
        return render(request, self.template_name, {
            'vehicules': cars,
            'categorie': categorie,
            'categories': categories,
            'cout_total': cout_total,
            'total_veh': total_veh,
            'total_veh_hors_parc': total_veh_hors_parc
        })

class AddVehiculeExcelView(LoginRequiredMixin, View):
    login_url = 'login'
    template_name = 'perfect/add_vehicule.html'
    success_message = 'Véhicule enregistré avec succès ✓✓'
    error_message = "Erreur de saisie ✘✘"
    def get(self, request, pk=None):
        cout_totals = 0
        total_veh = 0
        categories = CategoVehi.objects.annotate(
            nb_vehicules=Count('catego_vehicule', filter=Q(catego_vehicule__car_statut=True))
        ).order_by('id')
        if pk:
            categorie = get_object_or_404(CategoVehi, pk=pk)
            cars = Vehicule.objects.filter(category=categorie, car_statut=True).order_by('id')
            cout_totals = cars.aggregate(total=Sum('cout_acquisition'))['total'] or 0
            cout_total = '{:,}'.format(cout_totals).replace(',', ' ')
            total_veh = cars.filter(car_statut=True).count()
        else:
            categorie = None
            cars = Vehicule.objects.none()
            cout_total = 0
            total_veh = 0
        form = VehiculeForm()
        return render(request, self.template_name, {
            'forms': form,
            'vehicules': cars.filter(car_statut=True),
            'categorie': categorie,
            'categories': categories,
            'cout_total': cout_total,
            'total_veh': total_veh,
        })
    def post(self, request, pk=None):
        categorie = get_object_or_404(CategoVehi, pk=pk)
        categories = CategoVehi.objects.all().order_by('id')
        # --- Importation depuis Excel ---
        if request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            try:
                df = pd.read_excel(excel_file)
                df = df.fillna('')
                required_columns = [
                    'immatriculation', 'marque', 'duree', 'num_cart_grise',
                    'num_Chassis', 'date_acquisition', 'cout_acquisition',
                    'dat_edit_carte_grise', 'date_mis_service',
                ]
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f"Colonnes manquantes : {', '.join(missing_columns)}")
                    return redirect('add_vehi', pk=pk)
                created_count = 0
                updated_count = 0
                for _, row in df.iterrows():
                    immatriculation = str(row.get('immatriculation', '')).strip()
                    if not immatriculation:
                        continue
                    date_acquisition = pd.to_datetime(row.get('date_acquisition'), errors='coerce')
                    dat_edit_carte_grise = pd.to_datetime(row.get('dat_edit_carte_grise'), errors='coerce')
                    date_mis_service = pd.to_datetime(row.get('date_mis_service'), errors='coerce')
                    vehicule_data = {
                        'marque': str(row.get('marque', '')).strip(),
                        'duree': int(float(row.get('duree', 0) or 0)),
                        'num_cart_grise': str(row.get('num_cart_grise', '')).strip(),
                        'num_Chassis': str(row.get('num_Chassis', '')).strip(),
                        'date_acquisition': date_acquisition.date() if pd.notna(date_acquisition) else None,
                        'cout_acquisition': int(float(row.get('cout_acquisition', 0) or 0)),
                        'dat_edit_carte_grise': dat_edit_carte_grise.date() if pd.notna(dat_edit_carte_grise) else None,
                        'date_mis_service': date_mis_service.date() if pd.notna(date_mis_service) else None,
                        'category': categorie,
                        'auteur': request.user,
                    }
                    cars, created = Vehicule.objects.update_or_create(
                        immatriculation=immatriculation,
                        defaults=vehicule_data
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                messages.success(request, f"✅ {created_count} véhicule(s) créé(s), {updated_count} mis à jour.")
                return redirect('add_vehi', pk=pk)

            except Exception as e:
                messages.error(request, f"Erreur d'importation : {str(e)}")
                return redirect('add_vehi', pk=pk)
        form = VehiculeForm(request.POST, request.FILES)
        if form.is_valid():
            cars = form.save(commit=False)
            cars.auteur = request.user
            cars.category = categorie
            cars.save()
            messages.success(request, self.success_message)
            return redirect('all_vehi')
        else:
            messages.error(request, self.error_message)
            cars = Vehicule.objects.filter(category=categorie, car_statut=True).order_by('-id')
            cout_total = cars.aggregate(total=Sum('cout_acquisition'))['total'] or 0
            total_veh = cars.count()
            return render(request, self.template_name, {
                'forms': form,
                'vehicules': cars,
                'categorie': categorie,
                'categories': categories,
                'cout_total': cout_total,
                'total_veh': total_veh,
            })

class ExportVehiculeExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        # Gestion des permissions pour les gérants
        if user.user_type == "4":
            try:
                gerant = Gerant.objects.get(user=user)
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():
                    cars = Vehicule.objects.filter(category__in=categories_gerant, car_statut=True)
                else:
                    cars = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                cars = Vehicule.objects.none()
        else:
            cars = Vehicule.objects.filter(car_statut=True)
        
        # Application des filtres
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        if date_debut and date_fin:
            cars = cars.filter(date_saisie__range=[date_debut, date_fin])
        if immatriculation:
            cars = cars.filter(immatriculation__icontains=immatriculation)
        if categorie:
            try:
                # Si c'est un ID (entier)
                categorie_id = int(categorie)
                cars = cars.filter(category__id=categorie_id)
            except (ValueError, TypeError):
                # Si c'est un nom de catégorie (chaîne)
                cars = cars.filter(category__category=categorie)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Véhicules"
        headers = [
            "Immatriculation",
            "Marque",
            "Durée",
            "Numéro Carte Grise",
            "Numéro Châssis",
            "Date Acquisition",
            "Coût Acquisition",
            "Date Édition Carte Grise",
            "Date Mise en Service",
            "Catégorie",
            "Âge",
            "Auteur",
            "Date de saisie",
        ]
        ws.append(headers)
        for v in cars:
            ws.append([
            v.immatriculation,
            v.marque,
            v.duree,
            v.num_cart_grise,
            v.num_Chassis,
            v.date_acquisition.strftime("%d/%m/%Y") if v.date_acquisition else "",
            v.cout_acquisition,
            v.dat_edit_carte_grise.strftime("%d/%m/%Y") if v.dat_edit_carte_grise else "",
            v.date_mis_service.strftime("%d/%m/%Y") if v.date_mis_service else "",
            v.category.category if v.category else "",
            v.age,
            v.auteur.username if v.auteur else "",
            v.date_saisie.strftime("%d/%m/%Y") if v.date_saisie else "",
        ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Véhicules.xlsx'
        wb.save(response)
        return response

class ExportVehiculeHorsParcExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        # Gestion des permissions pour les gérants
        if user.user_type == "4":
            try:
                gerant = Gerant.objects.get(user=user)
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():
                    cars = Vehicule.objects.filter(category__in=categories_gerant, car_statut=False)
                else:
                    cars = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                cars = Vehicule.objects.none()
        else:
            cars = Vehicule.objects.filter(car_statut=False)
        # Application des filtres
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        if date_debut and date_fin:
            cars = cars.filter(date_saisie__range=[date_debut, date_fin])
        if immatriculation:
            cars = cars.filter(immatriculation__icontains=immatriculation)
        if categorie:
            try:
                # Si c'est un ID (entier)
                categorie_id = int(categorie)
                cars = cars.filter(category__id=categorie_id)
            except (ValueError, TypeError):
                # Si c'est un nom de catégorie (chaîne)
                cars = cars.filter(category__category=categorie)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Véhicules hors parc"
        headers = [
            "Immatriculation",
            "Marque",
            "Durée",
            "Numéro Carte Grise",
            "Numéro Châssis",
            "Date Acquisition",
            "Coût Acquisition",
            "Date Édition Carte Grise",
            "Date Mise en Service",
            "Catégorie",
            "Âge",
            "Motif de sortie",
            "Auteur",
            "Date de saisie",
        ]
        ws.append(headers)
        for v in cars:
            ws.append([
                v.immatriculation,
                v.marque,
                v.duree,
                v.num_cart_grise,
                v.num_Chassis,
                v.date_acquisition.strftime("%d/%m/%Y") if v.date_acquisition else "",
                v.cout_acquisition,
                v.dat_edit_carte_grise.strftime("%d/%m/%Y") if v.dat_edit_carte_grise else "",
                v.date_mis_service.strftime("%d/%m/%Y") if v.date_mis_service else "",
                v.category.category if v.category else "",
                v.age,
                v.motif_sorti or "",
                v.auteur.username if v.auteur else "",
                v.date_saisie.strftime("%d/%m/%Y") if v.date_saisie else "",
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Véhicules_hors_parc.xlsx'
        wb.save(response)
        return response

class HistoriqueVehiculeView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    """Vue pour afficher l'historique des véhicules avec filtres dynamiques"""
    login_url = 'login'
    permission_url = 'historique_vehicule'
    template_name = 'perfect/historiq_vehicule.html'
    context_object_name = 'liste_vehicules'
    form_class = HistoriqueVehiculeFilterForm
    timeout_minutes = 500
    
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        # Accéder au modèle historique via Vehicule.history.model
        HistoricalVehicule = Vehicule.history.model
        queryset = HistoricalVehicule.objects.all().order_by('-history_date').select_related('category', 'auteur', 'history_user')
        
        form = self.form_class(self.request.GET)
        if form.is_valid():
            # Filtres sur les dates d'historique
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            if date_debut and date_fin:
                queryset = queryset.filter(history_date__range=[date_debut, date_fin])
            elif date_debut:
                queryset = queryset.filter(history_date__gte=date_debut)
            elif date_fin:
                queryset = queryset.filter(history_date__lte=date_fin)
            
            # Filtre sur l'immatriculation
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                queryset = queryset.filter(immatriculation__icontains=immatriculation)
            
            # Filtre sur la catégorie
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                queryset = queryset.filter(category__id=categorie.id)
            
            # Filtre sur la marque
            marque = form.cleaned_data.get('marque')
            if marque:
                queryset = queryset.filter(marque__icontains=marque)
            
            # Filtre sur le motif de sortie
            motif_sorti = form.cleaned_data.get('motif_sorti')
            if motif_sorti:
                queryset = queryset.filter(motif_sorti__icontains=motif_sorti)
            
            # Filtre sur l'année de sortie (année où motif_sorti a été défini)
            annee_sortie = form.cleaned_data.get('annee_sortie')
            if annee_sortie is not None:
                # Filtrer les entrées historiques où motif_sorti n'est pas None et où l'année du history_date correspond
                queryset = queryset.filter(
                    motif_sorti__isnull=False,
                    history_date__year=annee_sortie
                )
            
            # Filtre sur l'année d'enregistrement (date_saisie)
            annee_enregistrement = form.cleaned_data.get('annee_enregistrement')
            if annee_enregistrement is not None:
                queryset = queryset.filter(date_saisie__year=annee_enregistrement)
            
            # Filtre sur le type d'historique (Créé, Modifié, Supprimé)
            history_type = form.cleaned_data.get('history_type')
            if history_type:
                queryset = queryset.filter(history_type=history_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        
        # Formulaire de filtre
        form = self.form_class(self.request.GET)
        context['form'] = form
        
        # Calculer l'âge pour chaque véhicule historique
        liste_vehicules = context.get('liste_vehicules', [])
        # S'assurer que c'est une liste pour pouvoir ajouter des attributs
        if hasattr(liste_vehicules, '__iter__') and not isinstance(liste_vehicules, (list, tuple)):
            liste_vehicules = list(liste_vehicules)
        
        today = date.today()
        for vehicule_hist in liste_vehicules:
            if vehicule_hist.date_mis_service:
                age = (today.year - vehicule_hist.date_mis_service.year) - int(
                    (vehicule_hist.date_mis_service.month, today.day) < (vehicule_hist.date_mis_service.month, today.day)
                )
                vehicule_hist.calculated_age = age
                # Calculer aussi la couleur selon l'âge
                if vehicule_hist.duree and vehicule_hist.duree > 0:
                    if age <= vehicule_hist.duree / 2:
                        vehicule_hist.calculated_color_age = "success"
                    elif age < vehicule_hist.duree:
                        vehicule_hist.calculated_color_age = "warning"
                    else:
                        vehicule_hist.calculated_color_age = "danger"
                else:
                    vehicule_hist.calculated_color_age = "info"
            else:
                vehicule_hist.calculated_age = None
                vehicule_hist.calculated_color_age = "info"
        
        # Mettre à jour le contexte avec la liste modifiée
        context['liste_vehicules'] = liste_vehicules
        
        # Calculer les statistiques sur les véhicules actuels filtrés
        vehicule_queryset = Vehicule.objects.all()
        
        # Appliquer les mêmes filtres sur les véhicules actuels pour les statistiques
        if form.is_valid():
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                vehicule_queryset = vehicule_queryset.filter(immatriculation__icontains=immatriculation)
            
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                vehicule_queryset = vehicule_queryset.filter(category__id=categorie.id)
            
            marque = form.cleaned_data.get('marque')
            if marque:
                vehicule_queryset = vehicule_queryset.filter(marque__icontains=marque)
            
            motif_sorti = form.cleaned_data.get('motif_sorti')
            if motif_sorti:
                vehicule_queryset = vehicule_queryset.filter(motif_sorti__icontains=motif_sorti)
            
            annee_enregistrement = form.cleaned_data.get('annee_enregistrement')
            if annee_enregistrement is not None:
                vehicule_queryset = vehicule_queryset.filter(date_saisie__year=annee_enregistrement)
        
        # Statistiques
        categories = CategoVehi.objects.annotate(nb_vehicules=Count('catego_vehicule', filter=Q(catego_vehicule__car_statut=True))).order_by('id')
        cout_totals = vehicule_queryset.filter(car_statut=True).aggregate(total=Sum('cout_acquisition'))['total'] or 0
        cout_total = '{:,}'.format(cout_totals).replace(',', ' ')
        total_veh = vehicule_queryset.filter(car_statut=True).count()
        total_veh_hors_parc = vehicule_queryset.filter(car_statut=False).count()
        
        context.update({
            'categories': categories,
            'cout_total': cout_total,
            'total_veh': total_veh,
            'total_veh_hors_parc': total_veh_hors_parc,
            'dates': dates,
            'annees': annee,
        })
        
        # Ajouter les permissions groupées si nécessaire
        user = self.request.user
        if hasattr(user, 'custom_permissions'):
            permissions = user.custom_permissions.all().select_related('categorie')
            grouped_permissions = {}
            for perm in permissions:
                if perm.categorie not in grouped_permissions:
                    grouped_permissions[perm.categorie] = []
                grouped_permissions[perm.categorie].append(perm)
            context['grouped_permissions'] = grouped_permissions
        
        return context

class ExportHistoriqueVehiculeExcelView(LoginRequiredMixin, View):
    """Vue pour exporter l'historique des véhicules au format Excel"""
    login_url = 'login'
    
    def get(self, request, *args, **kwargs):
        # Accéder au modèle historique via Vehicule.history.model
        HistoricalVehicule = Vehicule.history.model
        queryset = HistoricalVehicule.objects.all().order_by('-history_date').select_related('category', 'auteur', 'history_user')
        form = HistoriqueVehiculeFilterForm(request.GET)
        if form.is_valid():
            # Filtres sur les dates d'historique
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            if date_debut and date_fin:
                queryset = queryset.filter(history_date__range=[date_debut, date_fin])
            elif date_debut:
                queryset = queryset.filter(history_date__gte=date_debut)
            elif date_fin:
                queryset = queryset.filter(history_date__lte=date_fin)
            
            # Filtre sur l'immatriculation
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                queryset = queryset.filter(immatriculation__icontains=immatriculation)
            
            # Filtre sur la catégorie
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                queryset = queryset.filter(category__id=categorie.id)
            
            # Filtre sur la marque
            marque = form.cleaned_data.get('marque')
            if marque:
                queryset = queryset.filter(marque__icontains=marque)
            
            # Filtre sur le motif de sortie
            motif_sorti = form.cleaned_data.get('motif_sorti')
            if motif_sorti:
                queryset = queryset.filter(motif_sorti__icontains=motif_sorti)
            
            # Filtre sur l'année de sortie
            annee_sortie = form.cleaned_data.get('annee_sortie')
            if annee_sortie is not None:
                queryset = queryset.filter(
                    motif_sorti__isnull=False,
                    history_date__year=annee_sortie
                )
            
            # Filtre sur l'année d'enregistrement
            annee_enregistrement = form.cleaned_data.get('annee_enregistrement')
            if annee_enregistrement is not None:
                queryset = queryset.filter(date_saisie__year=annee_enregistrement)
            
            # Filtre sur le type d'historique
            history_type = form.cleaned_data.get('history_type')
            if history_type:
                queryset = queryset.filter(history_type=history_type)
        
        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historique Véhicules"
        
        # En-têtes de colonnes
        headers = [
            'Type',
            'Immatriculation',
            'Marque',
            'Catégorie',
            'Durée',
            'Numéro Carte Grise',
            'Numéro Châssis',
            'Date Acquisition',
            'Coût Acquisition',
            'Date Édition Carte Grise',
            'Date Mise en Service',
            'Statut',
            'Motif de sortie',
            'Date Saisie',
            'Auteur',
            'Date Historique',
            'Utilisateur Historique'
        ]
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[col_letter].width = 20
 
        # Données
        for h in queryset:
            history_type_label = ''
            if h.history_type == '+':
                history_type_label = 'Créé'
            elif h.history_type == '~':
                history_type_label = 'Modifié'
            elif h.history_type == '-':
                history_type_label = 'Supprimé'
            
            # Calculer l'âge du véhicule à partir de date_mis_service
            age = 0
            if h.date_mis_service:
                from datetime import date as date_module
                today = date_module.today()
                age = (today.year - h.date_mis_service.year) - int((h.date_mis_service.month, today.day) < (h.date_mis_service.month, today.day))
            
            ws.append([
                history_type_label,
                h.immatriculation or 'N/A',
                h.marque or 'N/A',
                h.category.category if h.category else 'N/A',
                h.duree,
                h.num_cart_grise or 'N/A',
                h.num_Chassis or 'N/A',
                h.date_acquisition.strftime("%d-%m-%Y") if h.date_acquisition else 'N/A',
                h.cout_acquisition,
                h.dat_edit_carte_grise.strftime("%d-%m-%Y") if h.dat_edit_carte_grise else 'N/A',
                h.date_mis_service.strftime("%d-%m-%Y") if h.date_mis_service else 'N/A',
                'En parc' if h.car_statut else 'Hors parc',
                h.motif_sorti or 'N/A',
                h.date_saisie.strftime("%d-%m-%Y") if h.date_saisie else 'N/A',
                h.auteur.username if h.auteur else 'N/A',
                h.history_date.strftime("%d-%m-%Y %H:%M") if h.history_date else 'N/A',
                h.history_user.username if h.history_user else 'N/A'
            ])
        
        # Réponse HTTP (fichier Excel téléchargeable)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Historique-Vehicules.xlsx'
        wb.save(response)
        return response

@login_required(login_url='/login/')
@require_POST
def toggle_car_statut(request, pk):
    vehicule = get_object_or_404(Vehicule, id=pk)
    if vehicule.car_statut:
        motif_sorti = request.POST.get('motif_sorti', '').strip()
        if not motif_sorti:
            return JsonResponse({
                "success": False, 
                "error": "Le motif de sortie est requis",
                "requires_motif": True
            })
        vehicule.motif_sorti = motif_sorti
        vehicule.car_statut = False
    else:
        vehicule.car_statut = True
        vehicule.motif_sorti = None
    vehicule.save()
    return JsonResponse({
        "success": True, 
        "car_statut": vehicule.car_statut,
        "message": "Statut du véhicule mis à jour avec succès"
    })

@login_required(login_url='login')
def delete_multiple_vehicules(request):
    ids = request.POST.getlist('vehicule_ids')
    if ids:
        vehicules = Vehicule.objects.filter(id__in=ids)
        deleted_count = vehicules.count()
        vehicules.delete()
        messages.success(request, f"{deleted_count} véhicule(s) supprimé(s) avec succès.")
    else:
        messages.warning(request, "Aucun véhicule sélectionné.")
    return redirect('all_vehi')

class UpdatVehiculeView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_car'
    model = Vehicule
    form_class = UpdatVehiculeForm
    template_name = "perfect/partials/car_form.html"  
    success_url = reverse_lazy('list_recet')
    success_message = 'véhicule modifiée avec succès✓✓'
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class CarFinanceView(LoginRequiredMixin, CustomPermissionRequiredMixin,TemplateView):
    login_url = 'login'
    permission_url = 'detail_car_financier'
    model = Vehicule
    template_name = 'perfect/dash_car.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()
        context={
            'vehicule':vehicules,
        }
        return context

class DetailVehiculeView(LoginRequiredMixin, DetailView):
    login_url = 'login'
    model = Vehicule
    template_name = 'perfect/dash_car.html'
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        now = datetime.now()
        mois_en_cours =date.today().month
        libelle_mois_en_cours = calendar.month_name[mois_en_cours]
        label = [calendar.month_name[month][:1] for month in range(1, 13)]

        vehicule = Vehicule.objects.all()
        vehicules = self.get_object()
        form = DateForm(self.request.GET)
        if form.is_valid():
            date_debut = form.cleaned_data['date_debut'] 
            date_fin = form.cleaned_data['date_fin']
            
            recettes = Recette.objects.filter(vehicule=vehicules, date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 1 
            charge_var = ChargeVariable.objects.filter(vehicule = vehicules, date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 0
            charge_fix = ChargeFixe.objects.filter(vehicule = vehicules, date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 0
            piecechang = PiecEchange.objects.filter(vehicule = vehicules, date_saisie__range=[date_debut, date_fin]).aggregate(somme=Sum('montant'))['somme'] or 0

            Total_charge = charge_fix + charge_var
            marg_contr = recettes - charge_var
            taux_marge = (marg_contr*100/(recettes))
            taux_marge_format='{:.2f}'.format(taux_marge)
            resultat = recettes-Total_charge

            nbreparation = Reparation.objects.filter(date_entree__range=[date_debut, date_fin], vehicule = vehicules).count() 
            reparations = Reparation.objects.filter(date_entree__range=[date_debut, date_fin],vehicule = vehicules)
            som_piece = Piece.objects.filter(reparation__in=reparations).aggregate(somme=Sum('montant'))['somme'] or 0
            
            ################################----Graphiques----#############################
            recet_data = Recette.objects.filter(vehicule=vehicules, date_saisie__range=[date_debut, date_fin])
            recet_mois_data = {month: 0 for month in range(1, 13)}
            for commande in recet_data:
                recet_mois_data[commande.date_saisie.month] += commande.montant
            recet_mois_data = [recet_mois_data[month] for month in range(1, 13)]
            
            chargfix_data = ChargeFixe.objects.filter(vehicule=vehicules, date_saisie__range=[date_debut, date_fin])
            chargfix_mois_data = {month: 0 for month in range(1, 13)}
            for commande in chargfix_data:
                chargfix_mois_data[commande.date_saisie.month] += commande.montant
            chargfix_mois_data = [chargfix_mois_data[month] for month in range(1, 13)]
            
            chargvar_data = ChargeVariable.objects.filter(vehicule=vehicules, date_saisie__range=[date_debut, date_fin])
            chargvar_mois_data = {month: 0 for month in range(1, 13)}
            for commande in chargvar_data:
                chargvar_mois_data[commande.date_saisie.month] += commande.montant
            chargvar_mois_data = [chargvar_mois_data[month] for month in range(1, 13)]
            
            reparations = Reparation.objects.filter(date_entree__range=[date_debut, date_fin],vehicule = vehicules)
            piece_data = Piece.objects.filter(reparation__in = reparations)
            piece_mois_data = {month: 0 for month in range(1, 13)}
            for commande in piece_data:
                piece_mois_data[commande.date_saisie.month] += commande.montant
            piece_mois_data = [piece_mois_data[month] for month in range(1, 13)]
            
            piechag_data = PiecEchange.objects.filter(vehicule=vehicules, date_saisie__range=[date_debut, date_fin])
            piechang_mois_data = {month: 0 for month in range(1, 13)}
            for commande in piechag_data:
                piechang_mois_data[commande.date_saisie.month] += commande.montant
            piechang_mois_data = [piechang_mois_data[month] for month in range(1, 13)]
            
            # label = [calendar.month_name[month][:1] for month in range(1, 13)]
            
            # marges_mensuelles = recet_mois_data - chargvar_mois_data
            marge_data = [recet_mois_data[i] - chargvar_mois_data[i] for i in range(12)]
            # Calcul des taux mensuels
            taux_data = [(marge_data[i] * 100) / recet_mois_data[i] if recet_mois_data[i] > 0 else 0 for i in range(12)]
        else:
            recettes = Recette.objects.filter(vehicule = vehicules, date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 1
            charge_fix = ChargeFixe.objects.filter(vehicule = vehicules, date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
            charge_var = ChargeVariable.objects.filter(vehicule = vehicules, date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
            piecechang = PiecEchange.objects.filter(vehicule = vehicules, date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0

            Total_charge = charge_fix + charge_var
            marg_contr = recettes - charge_var
            taux_marge = (marg_contr*100/(recettes))
            
            taux_marge_format='{:.2f}'.format(taux_marge)
            resultat = recettes-Total_charge
            
            
            nbreparation = Reparation.objects.filter(vehicule = vehicules, date_saisie__month=date.today().month).count()
            reparations = Reparation.objects.filter(vehicule = vehicules, date_saisie__month=date.today().month)
            som_piece = Piece.objects.filter(reparation__in = reparations).aggregate(somme=Sum('montant'))['somme'] or 0
            
            # reparation_mois = Reparation.objects.filter(vehicule = vehicule).annotate(month=ExtractMonth("date_entree")).values("month").annotate(total=Count("id")).values("month","total").order_by('month')

            recet_data = Recette.objects.filter(vehicule=vehicules, date_saisie__month=date.today().month)
            recet_mois_data = {month: 0 for month in range(1, 13)}
            for commande in recet_data:
                recet_mois_data[commande.date_saisie.month] += commande.montant
            recet_mois_data = [recet_mois_data[month] for month in range(1, 13)]
            
            chargfix_data = ChargeFixe.objects.filter(vehicule=vehicules, date_saisie__month=date.today().month)
            chargfix_mois_data = {month: 0 for month in range(1, 13)}
            for commande in chargfix_data:
                chargfix_mois_data[commande.date_saisie.month] += commande.montant
            chargfix_mois_data = [chargfix_mois_data[month] for month in range(1, 13)]
            
            chargvar_data = ChargeVariable.objects.filter(vehicule=vehicules, date_saisie__month=date.today().month)
            chargvar_mois_data = {month: 0 for month in range(1, 13)}
            for commande in chargvar_data:
                chargvar_mois_data[commande.date_saisie.month] += commande.montant
            chargvar_mois_data = [chargvar_mois_data[month] for month in range(1, 13)]
            
            piechag_data = PiecEchange.objects.filter(vehicule=vehicules, date_saisie__month=date.today().month)
            piechang_mois_data = {month: 0 for month in range(1, 13)}
            for commande in piechag_data:
                piechang_mois_data[commande.date_saisie.month] += commande.montant
            piechang_mois_data = [piechang_mois_data[month] for month in range(1, 13)]
            
            reparations = Reparation.objects.filter(vehicule = vehicules,date_saisie__month=date.today().month)
            piece_data = Piece.objects.filter(reparation__in = reparations)
            piece_mois_data = {month: 0 for month in range(1, 13)}
            for commande in piece_data:
                piece_mois_data[commande.date_saisie.month] += commande.montant
            piece_mois_data = [piece_mois_data[month] for month in range(1, 13)]

            marge_data = [recet_mois_data[i] - chargvar_mois_data[i] for i in range(12)]

            # Calcul des taux mensuels
            taux_data = [(marge_data[i] * 100) / recet_mois_data[i] if recet_mois_data[i] > 0 else 0 for i in range(12)]

        context.update({
            'recettes':recettes,
            'charge_fix':charge_fix,
            'charge_var':charge_var,
            # 'taux_marge':taux_marge,
            'taux_marge_format':taux_marge_format,
            'resultat':resultat,
            'nbreparation':nbreparation,
            'resultat':resultat,
            'som_piece':som_piece,
            'piecechang':piecechang,
            
            'recet_mois_data':recet_mois_data,
            'chargfix_mois_data':chargfix_mois_data,
            'chargvar_mois_data':chargvar_mois_data,
            'piece_mois_data':piece_mois_data,
            'piechang_mois_data':piechang_mois_data,
            
            'taux_data':taux_data,
            
            'dates':dates,
            'vehicules':vehicules,
            'vehicule':vehicule,
            'annees':annee,
            'mois_en_cours':libelle_mois_en_cours,
            
            'form':form,
            'labels':label,
        })    
        return context 
    
class SaisieGaragView(LoginRequiredMixin, CustomPermissionRequiredMixin, TemplateView):
    login_url = 'login'
    permission_url = 'saisi_garag'
    template_name = 'perfect/saisi_garag.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()
        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        
        # Initialiser les variables de filtre par défaut (pour le filtrage du mois en cours)
        date_debut = None
        date_fin = None
        categorie_filter = None
        
        assure_queryset = Assurance.objects.all()
        vigne_queryset = Vignette.objects.all()
        patent_queryset = Patente.objects.all()
        station_queryset = Stationnement.objects.all()
        piechang_queryset = PiecEchange.objects.all()

        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            
        if categorie_filter:
            assure_queryset = assure_queryset.filter(vehicule__category__category=categorie_filter)
            vigne_queryset = vigne_queryset.filter(vehicule__category__category=categorie_filter)
            patent_queryset = patent_queryset.filter(vehicule__category__category=categorie_filter)
            station_queryset = station_queryset.filter(vehicule__category__category=categorie_filter)
            piechang_queryset = piechang_queryset.filter(vehicule__category__category=categorie_filter)

        # Appliquer le filtre par date : si dates fournies, les utiliser, sinon filtrer par mois en cours
        if date_debut and date_fin:
            filtre_assure = assure_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_vigne = vigne_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_patent = patent_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_station = station_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_piechang = piechang_queryset.filter(date_saisie__range=[date_debut, date_fin])
        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_assure = assure_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_vigne = vigne_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_patent = patent_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_station = station_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_piechang = piechang_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)

        assure_mois = filtre_assure.aggregate(somme=Sum('montant'))['somme'] or 0
        assure_mois_format ='{:,}'.format(assure_mois).replace(',', ' ')

        vigne_mois = filtre_vigne.aggregate(somme=Sum('montant'))['somme'] or 0
        vigne_mois_format ='{:,}'.format(vigne_mois).replace(',', ' ')

        patent_mois = filtre_patent.aggregate(somme=Sum('montant'))['somme'] or 0
        patent_mois_format ='{:,}'.format(patent_mois).replace(',', ' ')

        station_mois = filtre_station.aggregate(somme=Sum('montant'))['somme'] or 0
        station_mois_format ='{:,}'.format(station_mois).replace(',', ' ')

        piechang_mois = filtre_piechang.aggregate(somme=Sum('montant'))['somme'] or 0
        piechang_mois_format ='{:,}'.format(piechang_mois).replace(',', ' ')

        # Permissions groupées pour le template (liens Assurance, Vignette, etc.)
        grouped_permissions = {}
        if hasattr(user, 'custom_permissions'):
            permissions = user.custom_permissions.all().select_related('categorie')
            for perm in permissions:
                if perm.categorie not in grouped_permissions:
                    grouped_permissions[perm.categorie] = []
                grouped_permissions[perm.categorie].append(perm)

        context={
            'vehicules': vehicules,
            'assure_mois_format': assure_mois_format,
            'vigne_mois_format': vigne_mois_format,
            'patent_mois_format': patent_mois_format,
            'station_mois_format': station_mois_format,
            'piechang_mois_format': piechang_mois_format,
            'form': form,
            'search_query': search_query,
            'grouped_permissions': grouped_permissions,
        }
        return context

class TempsArretsView(LoginRequiredMixin, CustomPermissionRequiredMixin, TemplateView):
    login_url = 'login'
    permission_url = 'temps_arrets'
    template_name = 'perfect/saisi_temp_arret.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois = date.today().month
        user = self.request.user

        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()
            
        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        
        visit_queryset = VisiteTechnique.objects.all()
        entretien_queryset = Entretien.objects.all()
        reparat_queryset = Reparation.objects.all()
        autarret_queryset = Autrarret.objects.all()

        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            
        if categorie_filter:
            visit_queryset = visit_queryset.filter(vehicule__category__category=categorie_filter)
            entretien_queryset = entretien_queryset.filter(vehicule__category__category=categorie_filter)
            reparat_queryset = reparat_queryset.filter(vehicule__category__category=categorie_filter)
            autarret_queryset = autarret_queryset.filter(vehicule__category__category=categorie_filter)

        if date_debut and date_fin:
            visit_queryset = visit_queryset.filter(date_saisie__range=[date_debut, date_fin])
            entretien_queryset = entretien_queryset.filter(date_saisie__range=[date_debut, date_fin])
            reparat_queryset = reparat_queryset.filter(date_saisie__range=[date_debut, date_fin])
            autarret_queryset = autarret_queryset.filter(date_saisie__range=[date_debut, date_fin])

        total_visit = visit_queryset.filter(date_saisie__month=date.today().month).count()
        total_entretien = entretien_queryset.filter(date_saisie__month=date.today().month).count()
        total_reparat = reparat_queryset.filter(date_saisie__month=date.today().month).count()
        total_autarret = autarret_queryset.filter(date_saisie__month=date.today().month).count()

        visit_mois = visit_queryset.filter(date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
        visit_mois_format ='{:,}'.format(visit_mois).replace(',', ' ')
        entretien_mois = entretien_queryset.filter(date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
        entretien_mois_format ='{:,}'.format(entretien_mois).replace(',', ' ')
        reparat_mois = reparat_queryset.filter(date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_mois_format ='{:,}'.format(reparat_mois).replace(',', ' ')
        autarret_mois = autarret_queryset.filter(date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 0
        autarret_mois_format ='{:,}'.format(autarret_mois).replace(',', ' ')

        context.update({
            'dates': dates,
            'vehicules': vehicules, 
            'total_visit': total_visit,
            'total_entretien': total_entretien,
            'total_reparat': total_reparat,
            'total_autarret': total_autarret,
            'visit_mois': visit_mois,
            'visit_mois_format': visit_mois_format,
            'entretien_mois_format': entretien_mois_format,
            'reparat_mois_format': reparat_mois_format,
            'autarret_mois_format': autarret_mois_format,
            'form': form,
            'search_query': search_query,
        })
        return context

class SaisiComptaView(LoginRequiredMixin, CustomPermissionRequiredMixin,TemplateView):
    login_url = 'login'
    permission_url = 'saisi_compta'
    template_name = 'perfect/saisi_comptable.html'
    timeout_minutes = 600
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois = date.today().month
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()
        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        # Initialiser les variables de filtre par défaut (pour le filtrage du mois en cours)
        date_debut = None
        date_fin = None
        categorie_filter = None
        
        recette_queryset = Recette.objects.all()
        chargvariale_queryset = ChargeVariable.objects.all()
        chargefixe_queryset = ChargeFixe.objects.all()

        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            
        if categorie_filter:
            recette_queryset = recette_queryset.filter(vehicule__category__category=categorie_filter)
            chargvariale_queryset = chargvariale_queryset.filter(vehicule__category__category=categorie_filter)
            chargefixe_queryset = chargefixe_queryset.filter(vehicule__category__category=categorie_filter)
        # Appliquer le filtre par date : si dates fournies, les utiliser, sinon filtrer par mois en cours
        if date_debut and date_fin:
            filtre_recette = recette_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_chargvariale = chargvariale_queryset.filter(date_saisie__range=[date_debut, date_fin])
            filtre_chargefixe = chargefixe_queryset.filter(date_saisie__range=[date_debut, date_fin])
        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_recette = recette_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_chargvariale = chargvariale_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
            filtre_chargefixe = chargefixe_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
        
        recette_mois = filtre_recette.aggregate(somme=Sum('montant'))['somme'] or 0
        recette_mois_format ='{:,}'.format(recette_mois).replace(',', ' ')

        chargvariale_mois = filtre_chargvariale.aggregate(somme=Sum('montant'))['somme'] or 0
        chargvariale_mois_format ='{:,}'.format(chargvariale_mois).replace(',', ' ')

        chargefixe_mois = filtre_chargefixe.aggregate(somme=Sum('montant'))['somme'] or 0
        chargefixe_mois_format ='{:,}'.format(chargefixe_mois).replace(',', ' ')

        chargtot_mois = chargefixe_mois + chargvariale_mois
        chargtot_mois_format ='{:,}'.format(chargtot_mois).replace(',', ' ')

        margcontrib_mois = recette_mois - chargtot_mois
        margcontrib_mois_format ='{:,}'.format(margcontrib_mois).replace(',', ' ')

        if recette_mois == 0:
            taux_marge_mois = 0
        else:
            taux_marge_mois = (margcontrib_mois*100/(recette_mois))
        taux_marge_mois_format ='{:.2f}'.format(taux_marge_mois).replace(',', ' ')

        result_mois = recette_mois - chargtot_mois
        result_mois_format ='{:,}'.format(result_mois).replace(',', ' ')
        context={
            'dates':dates,
            'vehicules':vehicules,
            'recette_mois_format':recette_mois_format,
            'chargvariale_mois_format':chargvariale_mois_format,
            'chargefixe_mois_format':chargefixe_mois_format,
            'chargtot_mois_format':chargtot_mois_format,
            'margcontrib_mois_format':margcontrib_mois_format,
            'taux_marge_mois_format':taux_marge_mois_format,
            'result_mois_format':result_mois_format,
            
            'form':form,
        }
        return context

#########################################---COMPTABLE---#########################################

class AddRecetteView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_recettes'
    model = Recette
    form_class = RecetteForm
    template_name= "perfect/add_recet.html"
    success_message = 'Recette Ajoutée avec succès ✓✓'
    error_message = "Erreur de saisie ✘✘"
    timeout_minutes = 500
    
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)  
    
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse    
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        recette_queryset = Recette.objects.filter(vehicule=vehicule)
        form = DateForm(self.request.GET)
        forms = self.get_form()
        date_debut = date_fin = None
        search_query = self.request.GET.get("search", "").strip()
        vehicules = Vehicule.objects.all()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        # Traitement des filtres de date
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
        # Filtre de date : si fourni, utiliser la plage ; sinon par défaut = mois en cours (comme ListRecetView)
        today = date.today()
        if date_debut and date_fin:
            recette_queryset_filtre = recette_queryset.filter(date_saisie__range=[date_debut, date_fin])
        else:
            recette_queryset_filtre = recette_queryset.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # Statistiques : jours/mois sur la période affichée, année sur la base (toute l'année)
        recette_jours = recette_queryset_filtre.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        recette_jours_format = '{:,}'.format(recette_jours).replace(',', ' ')
        recette_mois = recette_queryset_filtre.aggregate(somme=Sum('montant'))['somme'] or 0
        recette_mois_format = '{:,}'.format(recette_mois).replace(',', ' ')
        recette_an = recette_queryset.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        recette_an_format = '{:,}'.format(recette_an).replace(',', ' ')
        liste_recette = recette_queryset_filtre.order_by('-id')

        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            "recette_jours_format": recette_jours_format,
            "recette_mois_format": recette_mois_format,
            "recette_an_format": recette_an_format,
            "liste_recette": liste_recette,
            'form': form,
            'forms': forms,
        }   
        return context  
    def get_success_url(self):
        return reverse('add_recettes', kwargs={'pk': self.kwargs['pk']})

class ListRecetView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_recet'
    model = Recette
    template_name = 'perfect/liste_recette.html'
    context_object = 'listrecet'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        # Initialiser les variables de filtre par défaut (pour le filtrage du mois en cours)
        date_debut = None
        date_fin = None
        categorie_filter = None
        immatriculation = None
        
        recette_queryset = Recette.objects.all()
        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')

            if categorie_filter:
                recette_queryset = recette_queryset.filter(vehicule__category__category=categorie_filter)

            if date_debut and date_fin:
                recette_queryset = recette_queryset.filter(date_saisie__range=[date_debut, date_fin])
            
            if immatriculation:
                recette_queryset = recette_queryset.filter(vehicule__immatriculation=immatriculation)
        # Appliquer le filtre par date : si dates fournies, les utiliser, sinon filtrer par mois en cours
        if date_debut and date_fin:
            filtre_recette = recette_queryset.filter(date_saisie__range=[date_debut, date_fin])
        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_recette = recette_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
        # ################################----Recettes----#############################
        recettes_jours = filtre_recette.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_jours_format ='{:,}'.format(recettes_jours).replace(',', ' ')
        recettes_mois = filtre_recette.aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_mois_format ='{:,}'.format(recettes_mois).replace(',', ' ')
        recettes_an = recette_queryset.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_an_format ='{:,}'.format(recettes_an).replace(',', ' ')
        liste_rec = filtre_recette
        context={
            'liste_rec':liste_rec,
            'recettes_an_format':recettes_an_format,
            'recettes_mois_format':recettes_mois_format,
            'recettes_jours_format':recettes_jours_format,
            
            'dates':dates,
            'annees':annee,
            'form':form,
            }
        return context 

class ExportRecetteExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        recettes = Recette.objects.all()
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')

        if date_debut and date_fin:
            recettes = recettes.filter(date_saisie__range=[date_debut, date_fin])
        if immatriculation:
            recettes = recettes.filter(vehicule__immatriculation=immatriculation)
        if categorie:
            recettes = recettes.filter(vehicule__category__category=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recettes"
        headers = ["Immatriculation", "Marque", "Catégorie", "Chauffeur", "Montant", "Date saisie", "Auteur"]
        ws.append(headers)

        for r in recettes:
            ws.append([
                r.vehicule.immatriculation,
                r.vehicule.marque,
                r.vehicule.category.category,
                r.chauffeur,
                r.montant,
                r.date_saisie.strftime("%d-%m-%Y"),
                r.auteur.username if r.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Recettes.xlsx'
        wb.save(response)
        return response

class UpdateRecetView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_recet'
    model = Recette
    form_class = UpdateRecetteForm
    template_name = "perfect/partials/recette_form.html" 
    success_url = reverse_lazy('list_recet')
    success_message = 'Recette modifiée avec succès✓✓'

    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response

    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class HistoriqueRecetteView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'historique_recette'
    template_name = 'perfect/historiq_recette.html'
    context_object_name = 'liste_rec'
    form_class = HistoriqueRecetteFilterForm
    timeout_minutes = 500
    
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        # Accéder au modèle historique via Recette.history.model
        HistoricalRecette = Recette.history.model
        queryset = HistoricalRecette.objects.all().order_by('-history_date').select_related('vehicule', 'auteur', 'history_user')
        
        form = self.form_class(self.request.GET)
        if form.is_valid():
            # Filtres sur les dates d'historique
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            if date_debut and date_fin:
                queryset = queryset.filter(history_date__range=[date_debut, date_fin])
            elif date_debut:
                queryset = queryset.filter(history_date__gte=date_debut)
            elif date_fin:
                queryset = queryset.filter(history_date__lte=date_fin)
            
            # Filtre sur la date de saisie
            date_saisie_debut = form.cleaned_data.get('date_saisie_debut')
            date_saisie_fin = form.cleaned_data.get('date_saisie_fin')
            if date_saisie_debut and date_saisie_fin:
                queryset = queryset.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
            elif date_saisie_debut:
                queryset = queryset.filter(date_saisie__gte=date_saisie_debut)
            elif date_saisie_fin:
                queryset = queryset.filter(date_saisie__lte=date_saisie_fin)
            
            # Filtre sur la catégorie
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                queryset = queryset.filter(vehicule__category__id=categorie.id)
            
            # Filtre sur l'immatriculation
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                queryset = queryset.filter(vehicule__immatriculation__icontains=immatriculation)
            
            # Filtre sur le chauffeur
            chauffeur = form.cleaned_data.get('chauffeur')
            if chauffeur:
                queryset = queryset.filter(chauffeur__icontains=chauffeur)
            
            # Filtre sur le compte comptable
            cpte_comptable = form.cleaned_data.get('cpte_comptable')
            if cpte_comptable:
                queryset = queryset.filter(cpte_comptable__icontains=cpte_comptable)
            
            # Filtre sur le numéro de facture
            numero_fact = form.cleaned_data.get('numero_fact')
            if numero_fact:
                queryset = queryset.filter(numero_fact__icontains=numero_fact)
            
            # Filtre sur le numéro de pièce
            Num_piece = form.cleaned_data.get('Num_piece')
            if Num_piece:
                queryset = queryset.filter(Num_piece__icontains=Num_piece)
            
            # Filtre sur le montant
            montant_min = form.cleaned_data.get('montant_min')
            montant_max = form.cleaned_data.get('montant_max')
            if montant_min is not None:
                queryset = queryset.filter(montant__gte=montant_min)
            if montant_max is not None:
                queryset = queryset.filter(montant__lte=montant_max)
            
            # Filtre sur l'auteur
            auteur = form.cleaned_data.get('auteur')
            if auteur:
                queryset = queryset.filter(auteur=auteur)
            
            # Filtre sur le type d'historique (Créé, Modifié, Supprimé)
            history_type = form.cleaned_data.get('history_type')
            if history_type:
                queryset = queryset.filter(history_type=history_type)
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        # Formulaire de filtre
        form = self.form_class(self.request.GET)
        context['form'] = form
        # Calculer les totaux (recettes actuelles pour les stats, pas l'historique)
        # Base : catégorie + immat uniquement (comme ListReparationView)
        recette_base = Recette.objects.all()
        if form.is_valid():
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                recette_base = recette_base.filter(vehicule__category__id=categorie.id)
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                recette_base = recette_base.filter(vehicule__immatriculation__icontains=immatriculation)
        # Période affichée : base + filtre date (plage si fournie, sinon mois en cours)
        date_saisie_debut = form.cleaned_data.get('date_saisie_debut') if form.is_valid() else None
        date_saisie_fin = form.cleaned_data.get('date_saisie_fin') if form.is_valid() else None
        if date_saisie_debut and date_saisie_fin:
            filtre_recette = recette_base.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
        elif date_saisie_debut:
            filtre_recette = recette_base.filter(date_saisie__gte=date_saisie_debut)
        elif date_saisie_fin:
            filtre_recette = recette_base.filter(date_saisie__lte=date_saisie_fin)
        else:
            filtre_recette = recette_base.filter(
                date_saisie__month=date.today().month,
                date_saisie__year=date.today().year
            )
        # Statistiques : Jours / Mois sur la période affichée, Année sur la base (toute l'année)
        recettes_jours = filtre_recette.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_jours_format = '{:,}'.format(recettes_jours).replace(',', ' ')
        recettes_mois = filtre_recette.aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_mois_format = '{:,}'.format(recettes_mois).replace(',', ' ')
        recettes_an = recette_base.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_an_format = '{:,}'.format(recettes_an).replace(',', ' ')
        context.update({
            'recettes_jours_format': recettes_jours_format,
            'recettes_mois_format': recettes_mois_format,
            'recettes_an_format': recettes_an_format,
            'dates': dates,
            'annees': annee,
        })
        # Ajouter les permissions groupées si nécessaire (comme dans ListRecetView)
        user = self.request.user
        if hasattr(user, 'custom_permissions'):
            permissions = user.custom_permissions.all().select_related('categorie')
            grouped_permissions = {}
            for perm in permissions:
                if perm.categorie not in grouped_permissions:
                    grouped_permissions[perm.categorie] = []
                grouped_permissions[perm.categorie].append(perm)
            context['grouped_permissions'] = grouped_permissions
        return context

class ExportHistoriqueRecetteExcelView(LoginRequiredMixin, View):
    """Vue pour exporter l'historique des recettes au format Excel"""
    login_url = 'login'
    
    def get(self, request, *args, **kwargs):
        # Accéder au modèle historique via Recette.history.model
        HistoricalRecette = Recette.history.model
        queryset = HistoricalRecette.objects.all().order_by('-history_date').select_related('vehicule', 'auteur', 'history_user')
        
        # Récupérer les paramètres de filtre depuis la requête GET
        form = HistoriqueRecetteFilterForm(request.GET)
        
        if form.is_valid():
            # Filtres sur les dates d'historique
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            if date_debut and date_fin:
                queryset = queryset.filter(history_date__range=[date_debut, date_fin])
            elif date_debut:
                queryset = queryset.filter(history_date__gte=date_debut)
            elif date_fin:
                queryset = queryset.filter(history_date__lte=date_fin)
            
            # Filtre sur la date de saisie
            date_saisie_debut = form.cleaned_data.get('date_saisie_debut')
            date_saisie_fin = form.cleaned_data.get('date_saisie_fin')
            if date_saisie_debut and date_saisie_fin:
                queryset = queryset.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
            elif date_saisie_debut:
                queryset = queryset.filter(date_saisie__gte=date_saisie_debut)
            elif date_saisie_fin:
                queryset = queryset.filter(date_saisie__lte=date_saisie_fin)
            
            # Filtre sur la catégorie
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                queryset = queryset.filter(vehicule__category__id=categorie.id)
            
            # Filtre sur l'immatriculation
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                queryset = queryset.filter(vehicule__immatriculation__icontains=immatriculation)
            
            # Filtre sur le chauffeur
            chauffeur = form.cleaned_data.get('chauffeur')
            if chauffeur:
                queryset = queryset.filter(chauffeur__icontains=chauffeur)
            
            # Filtre sur le compte comptable
            cpte_comptable = form.cleaned_data.get('cpte_comptable')
            if cpte_comptable:
                queryset = queryset.filter(cpte_comptable__icontains=cpte_comptable)
            
            # Filtre sur le numéro de facture
            numero_fact = form.cleaned_data.get('numero_fact')
            if numero_fact:
                queryset = queryset.filter(numero_fact__icontains=numero_fact)
            
            # Filtre sur le numéro de pièce
            Num_piece = form.cleaned_data.get('Num_piece')
            if Num_piece:
                queryset = queryset.filter(Num_piece__icontains=Num_piece)
            
            # Filtre sur le montant
            montant_min = form.cleaned_data.get('montant_min')
            montant_max = form.cleaned_data.get('montant_max')
            if montant_min is not None:
                queryset = queryset.filter(montant__gte=montant_min)
            if montant_max is not None:
                queryset = queryset.filter(montant__lte=montant_max)
            
            # Filtre sur l'auteur
            auteur = form.cleaned_data.get('auteur')
            if auteur:
                queryset = queryset.filter(auteur=auteur)
            
            # Filtre sur le type d'historique (Créé, Modifié, Supprimé)
            history_type = form.cleaned_data.get('history_type')
            if history_type:
                queryset = queryset.filter(history_type=history_type)
        
        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historique Recettes"
        
        # En-têtes
        headers = [
            "Type",
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Chauffeur",
            "Compte Comptable",
            "Numéro Facture",
            "Numéro Pièce",
            "Montant",
            "Date Saisie",
            "Auteur",
            "Date Historique",
            "Utilisateur Historique"
        ]
        ws.append(headers)
        
        # Style pour les en-têtes
        header_fill = PatternFill(start_color="06497C", end_color="06497C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[col_letter].width = 20
        # Données
        for h in queryset:
            # Déterminer le type d'historique
            history_type_label = ""
            if h.history_type == '+':
                history_type_label = "Créé"
            elif h.history_type == '~':
                history_type_label = "Modifié"
            elif h.history_type == '-':
                history_type_label = "Supprimé"
            
            ws.append([
                history_type_label,
                h.vehicule.immatriculation if h.vehicule else "N/A",
                h.vehicule.marque if h.vehicule else "N/A",
                h.vehicule.category.category if h.vehicule and h.vehicule.category else "N/A",
                h.chauffeur,
                h.cpte_comptable,
                h.numero_fact,
                h.Num_piece,
                h.montant,
                h.date_saisie.strftime("%d-%m-%Y") if h.date_saisie else "",
                h.auteur.username if h.auteur else "N/A",
                h.history_date.strftime("%d-%m-%Y %H:%M:%S") if h.history_date else "",
                h.history_user.username if h.history_user else "N/A"
            ])
        
        # Alignement des cellules de données
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Réponse HTTP
        filename = f"Historique-Recettes-{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

class AddAutrarretView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_autarrets'
    model = Autrarret
    form_class = AutrarretForm
    template_name= "perfect/add_autarret.html"
    success_message = "Autre d'arrêt Ajouté avec succès ✓✓"
    error_message = "Erreur de saisie ✘✘ "
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)  
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        now = datetime.now()
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()
        
        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()

        autarret_queryset = Autrarret.objects.filter(vehicule=vehicule)
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            
            if date_debut and date_fin:
                autarret_queryset = autarret_queryset.filter(
                    date_saisie__range=[date_debut, date_fin]
                )

        autarret_total = autarret_queryset.filter(date_saisie__month=date.today().month).count()
        autarret_jours = autarret_queryset.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 1
        autarret_jours_format ='{:,}'.format(autarret_jours).replace(',', ' ')
        autarret_mois = autarret_queryset.filter(date_saisie__month=date.today().month).aggregate(somme=Sum('montant'))['somme'] or 1
        autarret_mois_format ='{:,}'.format(autarret_mois).replace(',', ' ')
        autarret_an = autarret_queryset.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 1
        autarret_an_format ='{:,}'.format(autarret_an).replace(',', ' ')
        liste_autarret = autarret_queryset.filter(date_saisie__month=date.today().month).order_by('-id')    

        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            "autarret_total": autarret_total,
            "autarret_jours_format": autarret_jours_format,
            "autarret_mois_format": autarret_mois_format,
            "autarret_an_format": autarret_an_format,
            "liste_autarret": liste_autarret,
            
            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
        }   
        return context  
    def get_success_url(self):
        return reverse('add_autarrets', kwargs={'pk': self.kwargs['pk']})
    
class ListarretView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'liste_aut_arrets'
    model = Autrarret
    template_name = 'perfect/liste_arret.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        # Initialiser les variables de filtre par défaut (pour le filtrage du mois en cours)
        date_debut = None
        date_fin = None
        categorie_filter = None
        immatriculation = None
        
        autarret_queryset = Autrarret.objects.all()
        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')

            if categorie_filter:
                autarret_queryset = autarret_queryset.filter(vehicule__category__category=categorie_filter)
            if date_debut and date_fin:
                autarret_queryset = autarret_queryset.filter(date_saisie__range=[date_debut, date_fin])
            if immatriculation:
                autarret_queryset = autarret_queryset.filter(vehicule__immatriculation__icontains=immatriculation)

        # Appliquer le filtre par date : si dates fournies, les utiliser, sinon filtrer par mois en cours
        if date_debut and date_fin:
            filtre_autarret = autarret_queryset.filter(date_saisie__range=[date_debut, date_fin])
        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_autarret = autarret_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)

        autarret_total = filtre_autarret.count()
        
        autarret_jours = autarret_queryset.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 1
        autarret_jours_format ='{:,}'.format(autarret_jours).replace(',', ' ')
        autarret_mois = filtre_autarret.aggregate(somme=Sum('montant'))['somme'] or 1
        autarret_mois_format ='{:,}'.format(autarret_mois).replace(',', ' ')
        autarret_an = autarret_queryset.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 1
        autarret_an_format ='{:,}'.format(autarret_an).replace(',', ' ')

        liste_autarret = filtre_autarret
        context={
            'liste_autarret':liste_autarret,
            'autarret_total':autarret_total,
            'autarret_an_format':autarret_an_format,
            'autarret_mois_format':autarret_mois_format,
            'autarret_jours_format':autarret_jours_format,
            'dates':dates,
            'annees':annee,
            'form':form,
            }
        return context


class UpdateAutrarretView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'liste_aut_arrets'
    model = Autrarret
    form_class = UpdatAutrarretForm
    template_name = "perfect/partials/autarret_form.html"
    success_url = reverse_lazy('liste_aut_arrets')
    success_message = 'Autre arrêt modifié avec succès.'

    def get_success_url(self):
        # Si on vient de add_autarret (vehicule en contexte), rediriger vers cette page
        next_url = self.request.GET.get('next') or self.request.POST.get('next')
        if next_url:
            return next_url
        return str(self.success_url)

    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response

    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            from django.template.loader import render_to_string
            ctx = self.get_context_data(form=form)
            html = render_to_string(self.template_name, ctx, request=self.request)
            return HttpResponse(html, content_type="text/html", status=422)
        return super().form_invalid(form)

class ExportAutrarretExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        autrarrets = Autrarret.objects.all()
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        today = date.today()

        # Filtre par date : plage fournie par l'utilisateur, sinon mois en cours par défaut
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                autrarrets = autrarrets.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                autrarrets = autrarrets.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            autrarrets = autrarrets.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        if immatriculation:
            autrarrets = autrarrets.filter(vehicule__immatriculation__icontains=immatriculation)
        if categorie:
            autrarrets = autrarrets.filter(vehicule__category__category__icontains=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Autres Arrêts"

        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Libellé",
            "Date Arrêt",
            "Date Sortie",
            "Durée (jours)",
            "Numéro Fiche",
            "Montant",
            "Date Saisie",
            "Auteur"
        ]
        ws.append(headers)
        for a in autrarrets:
            duree = ""
            if a.date_arret and a.date_sortie:
                duree = (a.date_sortie.date() - a.date_arret.date()).days
            ws.append([
                a.vehicule.immatriculation if a.vehicule else "",
                a.vehicule.marque if a.vehicule else "",
                a.vehicule.category.category if a.vehicule and a.vehicule.category else "",
                a.libelle or "",
                a.date_arret.strftime("%d-%m-%Y %H:%M") if a.date_arret else "",
                a.date_sortie.strftime("%d-%m-%Y %H:%M") if a.date_sortie else "",
                duree,
                a.numfich or "",
                a.montant,
                a.date_saisie.strftime("%d-%m-%Y") if a.date_saisie else "",
                a.auteur.username if a.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Autres-Arrets.xlsx'
        wb.save(response)
        return response

@require_POST
@login_required
def delete_selected_autarret(request):
    selected_ids = request.POST.getlist('autarret_ids')
    if selected_ids:
        autarrets_to_delete = Autrarret.objects.filter(id__in=selected_ids)
        count = autarrets_to_delete.count()
        autarrets_to_delete.delete()
        messages.success(request, f"{count} arrêt(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune arrêt sélectionnée.")
    return redirect('liste_aut_arrets') 

class BestRecetView(LoginRequiredMixin, CustomPermissionRequiredMixin, TemplateView):
    login_url = 'login'
    permission_url = 'list_recet'
    model = Recette
    template_name = 'perfect/best_recette.html'
    context_object = 'listrecet'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois = date.today().month
        recette_queryset = Recette.objects.all()
        form = self.form_class(self.request.GET)
        date_debut = None
        date_fin = None
        # ------------------ FILTRES -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

            if categorie_filter:
                recette_queryset = recette_queryset.filter(
                    vehicule__category__category=categorie_filter
                )

            if date_debut and date_fin:
                recette_queryset = recette_queryset.filter(
                    date_saisie__range=[date_debut, date_fin]
                )
        if date_debut and date_fin:
            filtre_recette = recette_queryset.filter(date_saisie__range=[date_debut, date_fin])

        else:
            # Par défaut : filtrer par mois en cours et année en cours
            mois_en_cours = date.today().month
            annee_en_cours = date.today().year
            filtre_recette = recette_queryset.filter(date_saisie__month=mois_en_cours, date_saisie__year=annee_en_cours)
                
        best_recettes = (
            filtre_recette.values('vehicule__immatriculation', 'vehicule__category__category')
            .annotate(total_recette=Sum('montant'))
            .order_by('-total_recette')
        )
        
        recettes_jours = recette_queryset.filter(
            date_saisie=date.today()
        ).aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_jours_format ='{:,}'.format(recettes_jours).replace(',', ' ')

        recettes_mois = filtre_recette.aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_mois_format ='{:,}'.format(recettes_mois).replace(',', ' ')
        
        recettes_an = recette_queryset.filter(
            date_saisie__year=annee
        ).aggregate(somme=Sum('montant'))['somme'] or 0
        recettes_an_format ='{:,}'.format(recettes_an).replace(',', ' ')
        context.update({
            'form': form,
            'dates': dates,
            'annees': annee,
            'recettes_jours_format': recettes_jours_format,
            'recettes_mois_format': recettes_mois_format,
            'recettes_an_format': recettes_an_format,
            'best_recettes': best_recettes,
        })
        return context 

class ExportBestRecetteExcelView(LoginRequiredMixin, View):
    """Vue pour exporter les meilleures recettes au format Excel"""
    login_url = 'login'
    
    def get(self, request, *args, **kwargs):
        form = DateFormMJR(request.GET)
        recette_queryset = Recette.objects.all()
        
        # ------------------ FILTRES (même logique que BestRecetView) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

            if categorie_filter:
                recette_queryset = recette_queryset.filter(
                    vehicule__category__category=categorie_filter
                )
            if date_debut and date_fin:
                recette_queryset = recette_queryset.filter(
                    date_saisie__range=[date_debut, date_fin]
                )
        
        # ------------------ AGREGATION (même logique que BestRecetView) -------------------
        best_recettes = (
            recette_queryset.values('vehicule__immatriculation', 'vehicule__category__category')
            .annotate(total_recette=Sum('montant'))
            .order_by('-total_recette')  # Du plus grand au plus petit
        )
        
        # ------------------ CREATION DU FICHIER EXCEL -------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Meilleures Recettes"
        
        # En-têtes
        headers = [
            "IMMAT",
            "CATEGORIE",
            "MONTANT (FCFA)"
        ]
        ws.append(headers)
        
        # Style pour les en-têtes
        header_fill = PatternFill(start_color="06497C", end_color="06497C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[col_letter].width = 20
        
        # Données
        for rec in best_recettes:
            ws.append([
                rec.get('vehicule__immatriculation', ''),
                rec.get('vehicule__category__category', ''),
                rec.get('total_recette', 0),
            ])
        
        # Alignement des cellules de données
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatage numérique pour la colonne montant
        for row in range(2, ws.max_row + 1):
            cell = ws[f"C{row}"]
            cell.number_format = '#,##0'
        
        # Réponse HTTP
        filename = f"Meilleures-Recettes-{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

@require_POST
@login_required
def delete_selected_visite(request):
    selected_ids = request.POST.getlist('visite_ids')
    if selected_ids:
        visites_to_delete = VisiteTechnique.objects.filter(id__in=selected_ids)
        count = visites_to_delete.count()
        visites_to_delete.delete()
        messages.success(request, f"{count} visite(s) technique(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune visite technique sélectionnée.")
    return redirect('list_visit')

@require_POST
@login_required
def delete_selected_recettes(request):
    selected_ids = request.POST.getlist('recette_ids')
    if selected_ids:
        recettes_to_delete = Recette.objects.filter(id__in=selected_ids)
        count = recettes_to_delete.count()
        recettes_to_delete.delete()
        messages.success(request, f"{count} recette(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune recette sélectionnée.")
    return redirect('list_recet')

@require_POST
@login_required
def delete_selected_chargvar(request):
    selected_ids = request.POST.getlist('chargvar_ids')
    if selected_ids:
        chargvar_to_delete = ChargeVariable.objects.filter(id__in=selected_ids)
        count = chargvar_to_delete.count()
        chargvar_to_delete.delete()
        messages.success(request, f"{count} Charge variable(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune charge variable sélectionnée.")
    return redirect('list_charg_var')

@require_POST
@login_required
def delete_selected_chargfix(request):
    selected_ids = request.POST.getlist('chargfix_ids')
    if selected_ids:
        chargfix_to_delete = ChargeFixe.objects.filter(id__in=selected_ids)
        count = chargfix_to_delete.count()
        chargfix_to_delete.delete()
        messages.success(request, f"{count} charge fixe(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune charge fixe sélectionnée.")
    return redirect('list_charg_fix')
    
class AddChargeFixView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'addcharg_fix'
    model = ChargeFixe
    form_class = ChargeFixForm
    template_name= "perfect/add_charg_fixe.html"
    success_message = 'Charge fixe Ajoutée avec succès ✓✓'
    error_message = "Erreur de saisie ✘✘ "
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)  
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        now = datetime.now()
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        chargfixe_queryset = ChargeFixe.objects.filter(vehicule=vehicule)
        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()

        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        today = date.today()
        chargfixe_base = ChargeFixe.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_chargfixe = chargfixe_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_chargfixe = chargfixe_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        chargfixe_jours = filtre_chargfixe.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        chargfixe_jours_format = '{:,}'.format(chargfixe_jours).replace(',', ' ')
        chargfixe_mois = filtre_chargfixe.aggregate(somme=Sum('montant'))['somme'] or 0
        chargfixe_mois_format = '{:,}'.format(chargfixe_mois).replace(',', ' ')
        chargfixe_an = chargfixe_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        chargfixe_an_format = '{:,}'.format(chargfixe_an).replace(',', ' ')
        liste_chargfixe = filtre_chargfixe.order_by('-id')
        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            "chargfixe_jours_format": chargfixe_jours_format,
            "chargfixe_mois_format": chargfixe_mois_format,
            "chargfixe_an_format": chargfixe_an_format,
            "liste_chargfixe": liste_chargfixe,
            
            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
        }   
        return context  
    def get_success_url(self):
        return reverse('addcharg_fix', kwargs={'pk': self.kwargs['pk']})
 
class ListChargeFixView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_charg_fix'
    model = ChargeFixe
    template_name = 'perfect/liste_charg_fix.html'
    context_object = 'list_charg_fix'
    form_class = DateFormMJR
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        chargefix_base = ChargeFixe.objects.all()
        form = self.form_class(self.request.GET)
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')

            if categorie_filter:
                chargefix_base = chargefix_base.filter(vehicule__category__category=categorie_filter)
            if immatriculation:
                chargefix_base = chargefix_base.filter(vehicule__immatriculation__icontains=immatriculation)
            if date_debut and date_fin:
                filtre_chargefix = chargefix_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_chargefix = chargefix_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_chargefix = chargefix_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ STATISTIQUES -------------------
        chargefix_jours = filtre_chargefix.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        chargefix_jours_format = '{:,}'.format(chargefix_jours).replace(',', ' ')
        chargefix_mois = filtre_chargefix.aggregate(somme=Sum('montant'))['somme'] or 0
        chargefix_mois_format = '{:,}'.format(chargefix_mois).replace(',', ' ')
        chargefix_an = chargefix_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        chargefix_an_format = '{:,}'.format(chargefix_an).replace(',', ' ')
        context = {
            'liste_chargefix': filtre_chargefix,
            'chargefix_an_format': chargefix_an_format,
            'chargefix_mois_format': chargefix_mois_format,
            'chargefix_jours_format': chargefix_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context 
     
class UpdateChargFixView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'upd_charg_fix'
    model = ChargeFixe
    form_class = UpdatChargeFixForm
    template_name = "perfect/partials/chargfix_form.html"
    success_url = reverse_lazy('list_charg_fix')
    success_message = 'Charge fixe modifiée avec succès✓✓'
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class ExportChargeFixeExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        charges = ChargeFixe.objects.all()
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        # Appliquer les filtres
        if date_debut and date_fin:
            charges = charges.filter(date_saisie__range=[date_debut, date_fin])
        if immatriculation:
            charges = charges.filter(vehicule__immatriculation=immatriculation)
        if categorie:
            charges = charges.filter(vehicule__category__category=categorie)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Charges Fixes"
        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Libellé",
            "Compte comptable",
            "N° Pièce",
            "N° Facture",
            "Montant",
            "Date saisie",
            "Auteur"
        ]
        ws.append(headers)
        for c in charges:
            ws.append([
                c.vehicule.immatriculation,
                c.vehicule.marque,
                c.vehicule.category.category,
                c.libelle,
                c.cpte_comptable,
                c.Num_piece,
                c.Num_fact,
                c.montant,
                c.date_saisie.strftime("%d-%m-%Y") if c.date_saisie else "",
                c.auteur.username if c.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Charges-Fixes.xlsx'
        wb.save(response)
        return response

class HistoriqueChargeFixeView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    """Vue pour afficher l'historique des charges fixes avec filtres dynamiques"""
    login_url = 'login'
    permission_url = 'historique_charge_fixe'
    template_name = 'perfect/historiq_chargfix.html'
    context_object_name = 'liste_chargefix'
    form_class = HistoriqueChargeFixeFilterForm
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        HistoricalChargeFixe = ChargeFixe.history.model
        queryset = HistoricalChargeFixe.objects.all().order_by('-history_date').select_related('vehicule', 'auteur', 'history_user')
        form = self.form_class(self.request.GET)
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            if date_debut and date_fin:
                queryset = queryset.filter(history_date__range=[date_debut, date_fin])
            elif date_debut:
                queryset = queryset.filter(history_date__gte=date_debut)
            elif date_fin:
                queryset = queryset.filter(history_date__lte=date_fin)
            
            date_saisie_debut = form.cleaned_data.get('date_saisie_debut')
            date_saisie_fin = form.cleaned_data.get('date_saisie_fin')
            if date_saisie_debut and date_saisie_fin:
                queryset = queryset.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
            elif date_saisie_debut:
                queryset = queryset.filter(date_saisie__gte=date_saisie_debut)
            elif date_saisie_fin:
                queryset = queryset.filter(date_saisie__lte=date_saisie_fin)
            
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                queryset = queryset.filter(vehicule__category__id=categorie.id)
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                queryset = queryset.filter(vehicule__immatriculation__icontains=immatriculation)
            libelle = form.cleaned_data.get('libelle')
            if libelle:
                queryset = queryset.filter(libelle__icontains=libelle)
            cpte_comptable = form.cleaned_data.get('cpte_comptable')
            if cpte_comptable:
                queryset = queryset.filter(cpte_comptable__icontains=cpte_comptable)
            Num_piece = form.cleaned_data.get('Num_piece')
            if Num_piece:
                queryset = queryset.filter(Num_piece__icontains=Num_piece)
            Num_fact = form.cleaned_data.get('Num_fact')
            if Num_fact:
                queryset = queryset.filter(Num_fact__icontains=Num_fact)
            montant_min = form.cleaned_data.get('montant_min')
            montant_max = form.cleaned_data.get('montant_max')
            if montant_min is not None:
                queryset = queryset.filter(montant__gte=montant_min)
            if montant_max is not None:
                queryset = queryset.filter(montant__lte=montant_max)
            auteur = form.cleaned_data.get('auteur')
            if auteur:
                queryset = queryset.filter(auteur=auteur)
            history_type = form.cleaned_data.get('history_type')
            if history_type:
                queryset = queryset.filter(history_type=history_type)
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        form = self.form_class(self.request.GET)
        context['form'] = form
        chargefix_base = ChargeFixe.objects.all()
        if form.is_valid():
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                chargefix_base = chargefix_base.filter(vehicule__category__id=categorie.id)
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                chargefix_base = chargefix_base.filter(vehicule__immatriculation__icontains=immatriculation)
        date_saisie_debut = form.cleaned_data.get('date_saisie_debut') if form.is_valid() else None
        date_saisie_fin = form.cleaned_data.get('date_saisie_fin') if form.is_valid() else None
        if date_saisie_debut and date_saisie_fin:
            filtre_chargefix = chargefix_base.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
        elif date_saisie_debut:
            filtre_chargefix = chargefix_base.filter(date_saisie__gte=date_saisie_debut)
        elif date_saisie_fin:
            filtre_chargefix = chargefix_base.filter(date_saisie__lte=date_saisie_fin)
        else:
            filtre_chargefix = chargefix_base.filter(
                date_saisie__month=date.today().month,
                date_saisie__year=date.today().year
            )
        # Statistiques : Jours / Mois sur la période affichée, Année sur la base (toute l'année)
        chargefix_jours = filtre_chargefix.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
        chargefix_jours_format = '{:,}'.format(chargefix_jours).replace(',', ' ')
        chargefix_mois = filtre_chargefix.aggregate(somme=Sum('montant'))['somme'] or 0
        chargefix_mois_format = '{:,}'.format(chargefix_mois).replace(',', ' ')
        chargefix_an = chargefix_base.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 0
        chargefix_an_format = '{:,}'.format(chargefix_an).replace(',', ' ')
        
        context.update({
            'chargefix_jours_format': chargefix_jours_format,
            'chargefix_mois_format': chargefix_mois_format,
            'chargefix_an_format': chargefix_an_format,
            'dates': dates,
            'annees': annee,
        })
        # Ajouter les permissions groupées si nécessaire (comme dans ListChargeFixView)
        user = self.request.user
        if hasattr(user, 'custom_permissions'):
            permissions = user.custom_permissions.all().select_related('categorie')
            grouped_permissions = {}
            for perm in permissions:
                if perm.categorie not in grouped_permissions:
                    grouped_permissions[perm.categorie] = []
                grouped_permissions[perm.categorie].append(perm)
            context['grouped_permissions'] = grouped_permissions
        return context

class AddChargeVarView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'addcharg_var'
    model = ChargeVariable
    form_class = ChargeVarForm
    template_name= "perfect/add_charg_var.html"
    success_message = 'Charge variable Ajoutée avec succès ✓✓'
    error_message = "Erreur de saisie ✘✘"
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)  
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        now = datetime.now()
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        chargvar_queryset = ChargeVariable.objects.filter(vehicule=vehicule)
        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.filter(category="VTC").exists():
                    vehicules = Vehicule.objects.filter(category__category="VTC")
                else:
                    vehicules = Vehicule.objects.filter(category__category="TAXI")
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()  
                # No vehicles if no Gerant linked
        elif user:
            try:
                vehicules = Vehicule.objects.all()
            except:
                vehicules = Vehicule.objects.none()
        else:
            print()
        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        today = date.today()
        chargvar_base = ChargeVariable.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_chargvar = chargvar_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_chargvar = chargvar_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        chargvar_jours = filtre_chargvar.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        chargvar_jours_format = '{:,}'.format(chargvar_jours).replace(',', ' ')
        chargvar_mois = filtre_chargvar.aggregate(somme=Sum('montant'))['somme'] or 0
        chargvar_mois_format = '{:,}'.format(chargvar_mois).replace(',', ' ')
        chargvar_an = chargvar_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        chargvar_an_format = '{:,}'.format(chargvar_an).replace(',', ' ')
        liste_chargvar = filtre_chargvar.order_by('-id')

        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            'chargvar_jours_format': chargvar_jours_format,
            'chargvar_mois_format': chargvar_mois_format,
            'chargvar_an_format': chargvar_an_format,
            'liste_chargvar': liste_chargvar,
            'form': form,
            'forms': forms,
        }   
        return context  
    def get_success_url(self):
        return reverse('addcharg_var', kwargs={'pk': self.kwargs['pk']})

class ListChargeVarView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_charg_var'
    model = ChargeVariable
    template_name = 'perfect/liste_charg_var.html'
    context_object = 'list_charg_var'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        chargevar_base = ChargeVariable.objects.all()
        form = self.form_class(self.request.GET)
        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')

            if categorie_filter:
                chargevar_base = chargevar_base.filter(vehicule__category__category=categorie_filter)

            if immatriculation:
                chargevar_base = chargevar_base.filter(vehicule__immatriculation__icontains=immatriculation)

            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListRecetView)
            if date_debut and date_fin:
                filtre_chargevar = chargevar_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_chargevar = chargevar_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            # Formulaire invalide ou vide : par défaut mois en cours
            filtre_chargevar = chargevar_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        # ------------------ STATISTIQUES -------------------
        # Jours / Mois : sur la période affichée (filtre_chargevar)
        chargevar_jours = filtre_chargevar.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        chargevar_jours_format = '{:,}'.format(chargevar_jours).replace(',', ' ')
        chargevar_mois = filtre_chargevar.aggregate(somme=Sum('montant'))['somme'] or 0
        chargevar_mois_format = '{:,}'.format(chargevar_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        chargevar_an = chargevar_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        chargevar_an_format = '{:,}'.format(chargevar_an).replace(',', ' ')

        context = {
            'liste_chargevar': filtre_chargevar,
            'chargevar_an_format': chargevar_an_format,
            'chargevar_mois_format': chargevar_mois_format,
            'chargevar_jours_format': chargevar_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

class UpdateChargeVarView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_charg_var'
    model = ChargeVariable
    form_class = updatChargeVarForm
    template_name = "perfect/partials/chargvar_form.html"
    success_url = reverse_lazy('list_charg_var')
    success_message = 'Charge variable modifiée avec succès✓✓'

    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class ExportChargeVariableExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Récupérer toutes les charges variables
        charges = ChargeVariable.objects.all()

        # Récupérer les filtres GET
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')

        # Appliquer les filtres
        if date_debut and date_fin:
            charges = charges.filter(date_saisie__range=[date_debut, date_fin])
        if immatriculation:
            charges = charges.filter(vehicule__immatriculation=immatriculation)
        if categorie:
            charges = charges.filter(vehicule__category__category=categorie)

        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Charges Variables"

        # En-têtes de colonnes
        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Libellé",
            "Compte comptable",
            "N° Pièce",
            "N° Facture",
            "Montant",
            "Date saisie",
            "Auteur"
        ]
        ws.append(headers)

        # Lignes de données
        for c in charges:
            ws.append([
                c.vehicule.immatriculation,
                c.vehicule.marque,
                c.vehicule.category.category,
                c.libelle,
                c.cpte_comptable,
                c.Num_piece,
                c.Num_fact,
                c.montant,
                c.date_saisie.strftime("%d-%m-%Y") if c.date_saisie else "",
                c.auteur.username if c.auteur else "N/A"
            ])

        # Ajustement automatique des colonnes
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20

        # Réponse HTTP (fichier Excel téléchargeable)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Charges-Variables.xlsx'
        wb.save(response)
        return response

class HistoriqueChargeVariableView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    """Vue pour afficher l'historique des charges variables avec filtres dynamiques"""
    login_url = 'login'
    permission_url = 'historique_charge_variable'
    template_name = 'perfect/historiq_chargvar.html'
    context_object_name = 'liste_chargevar'
    form_class = HistoriqueChargeVariableFilterForm
    timeout_minutes = 500
    
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        # Accéder au modèle historique via ChargeVariable.history.model
        HistoricalChargeVariable = ChargeVariable.history.model
        queryset = HistoricalChargeVariable.objects.all().order_by('-history_date').select_related('vehicule', 'auteur', 'history_user')
        
        form = self.form_class(self.request.GET)
        if form.is_valid():
            # Filtres sur les dates d'historique
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            if date_debut and date_fin:
                queryset = queryset.filter(history_date__range=[date_debut, date_fin])
            elif date_debut:
                queryset = queryset.filter(history_date__gte=date_debut)
            elif date_fin:
                queryset = queryset.filter(history_date__lte=date_fin)
            
            # Filtre sur la date de saisie
            date_saisie_debut = form.cleaned_data.get('date_saisie_debut')
            date_saisie_fin = form.cleaned_data.get('date_saisie_fin')
            if date_saisie_debut and date_saisie_fin:
                queryset = queryset.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
            elif date_saisie_debut:
                queryset = queryset.filter(date_saisie__gte=date_saisie_debut)
            elif date_saisie_fin:
                queryset = queryset.filter(date_saisie__lte=date_saisie_fin)
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                queryset = queryset.filter(vehicule__category__id=categorie.id)
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                queryset = queryset.filter(vehicule__immatriculation__icontains=immatriculation)
            libelle = form.cleaned_data.get('libelle')
            if libelle:
                queryset = queryset.filter(libelle__icontains=libelle)
            cpte_comptable = form.cleaned_data.get('cpte_comptable')
            if cpte_comptable:
                queryset = queryset.filter(cpte_comptable__icontains=cpte_comptable)
            Num_piece = form.cleaned_data.get('Num_piece')
            if Num_piece:
                queryset = queryset.filter(Num_piece__icontains=Num_piece)
            Num_fact = form.cleaned_data.get('Num_fact')
            if Num_fact:
                queryset = queryset.filter(Num_fact__icontains=Num_fact)

            montant_min = form.cleaned_data.get('montant_min')
            montant_max = form.cleaned_data.get('montant_max')
            if montant_min is not None:
                queryset = queryset.filter(montant__gte=montant_min)
            if montant_max is not None:
                queryset = queryset.filter(montant__lte=montant_max)
            
            auteur = form.cleaned_data.get('auteur')
            if auteur:
                queryset = queryset.filter(auteur=auteur)
            
            # Filtre sur le type d'historique (Créé, Modifié, Supprimé)
            history_type = form.cleaned_data.get('history_type')
            if history_type:
                queryset = queryset.filter(history_type=history_type)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        # Formulaire de filtre
        form = self.form_class(self.request.GET)
        context['form'] = form
        chargevar_base = ChargeVariable.objects.all()
        if form.is_valid():
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                chargevar_base = chargevar_base.filter(vehicule__category__id=categorie.id)
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                chargevar_base = chargevar_base.filter(vehicule__immatriculation__icontains=immatriculation)
        # Période affichée : base + filtre date (plage si fournie, sinon mois en cours)
        date_saisie_debut = form.cleaned_data.get('date_saisie_debut') if form.is_valid() else None
        date_saisie_fin = form.cleaned_data.get('date_saisie_fin') if form.is_valid() else None
        if date_saisie_debut and date_saisie_fin:
            filtre_chargevar = chargevar_base.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
        elif date_saisie_debut:
            filtre_chargevar = chargevar_base.filter(date_saisie__gte=date_saisie_debut)
        elif date_saisie_fin:
            filtre_chargevar = chargevar_base.filter(date_saisie__lte=date_saisie_fin)
        else:
            filtre_chargevar = chargevar_base.filter(
                date_saisie__month=date.today().month,
                date_saisie__year=date.today().year
            )
        chargevar_jours = filtre_chargevar.filter(date_saisie=date.today()).aggregate(somme=Sum('montant'))['somme'] or 0
        chargevar_jours_format = '{:,}'.format(chargevar_jours).replace(',', ' ')
        chargevar_mois = filtre_chargevar.aggregate(somme=Sum('montant'))['somme'] or 0
        chargevar_mois_format = '{:,}'.format(chargevar_mois).replace(',', ' ')
        chargevar_an = chargevar_base.filter(date_saisie__year=date.today().year).aggregate(somme=Sum('montant'))['somme'] or 0
        chargevar_an_format = '{:,}'.format(chargevar_an).replace(',', ' ')
        context.update({
            'chargevar_jours_format': chargevar_jours_format,
            'chargevar_mois_format': chargevar_mois_format,
            'chargevar_an_format': chargevar_an_format,
            'dates': dates,
            'annees': annee,
        })
        user = self.request.user
        if hasattr(user, 'custom_permissions'):
            permissions = user.custom_permissions.all().select_related('categorie')
            grouped_permissions = {}
            for perm in permissions:
                if perm.categorie not in grouped_permissions:
                    grouped_permissions[perm.categorie] = []
                grouped_permissions[perm.categorie].append(perm)
            context['grouped_permissions'] = grouped_permissions
        return context

class ExportHistoriqueChargeVariableExcelView(LoginRequiredMixin, View):
    """Vue pour exporter l'historique des charges variables au format Excel"""
    login_url = 'login'
    
    def get(self, request, *args, **kwargs):
        # Accéder au modèle historique via ChargeVariable.history.model
        HistoricalChargeVariable = ChargeVariable.history.model
        queryset = HistoricalChargeVariable.objects.all().order_by('-history_date').select_related('vehicule', 'auteur', 'history_user')
        # Récupérer les paramètres de filtre depuis la requête GET
        form = HistoriqueChargeVariableFilterForm(request.GET)
        if form.is_valid():
            # Filtres sur les dates d'historique
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            if date_debut and date_fin:
                queryset = queryset.filter(history_date__range=[date_debut, date_fin])
            elif date_debut:
                queryset = queryset.filter(history_date__gte=date_debut)
            elif date_fin:
                queryset = queryset.filter(history_date__lte=date_fin)
            
            # Filtre sur la date de saisie
            date_saisie_debut = form.cleaned_data.get('date_saisie_debut')
            date_saisie_fin = form.cleaned_data.get('date_saisie_fin')
            if date_saisie_debut and date_saisie_fin:
                queryset = queryset.filter(date_saisie__range=[date_saisie_debut, date_saisie_fin])
            elif date_saisie_debut:
                queryset = queryset.filter(date_saisie__gte=date_saisie_debut)
            elif date_saisie_fin:
                queryset = queryset.filter(date_saisie__lte=date_saisie_fin)
            
            # Filtre sur la catégorie
            categorie = form.cleaned_data.get('categorie')
            if categorie:
                queryset = queryset.filter(vehicule__category__id=categorie.id)
            
            # Filtre sur l'immatriculation
            immatriculation = form.cleaned_data.get('immatriculation')
            if immatriculation:
                queryset = queryset.filter(vehicule__immatriculation__icontains=immatriculation)
            
            # Filtre sur le libellé
            libelle = form.cleaned_data.get('libelle')
            if libelle:
                queryset = queryset.filter(libelle__icontains=libelle)
            
            # Filtre sur le compte comptable
            cpte_comptable = form.cleaned_data.get('cpte_comptable')
            if cpte_comptable:
                queryset = queryset.filter(cpte_comptable__icontains=cpte_comptable)
            
            # Filtre sur le numéro de pièce
            Num_piece = form.cleaned_data.get('Num_piece')
            if Num_piece:
                queryset = queryset.filter(Num_piece__icontains=Num_piece)
            
            # Filtre sur le numéro de facture
            Num_fact = form.cleaned_data.get('Num_fact')
            if Num_fact:
                queryset = queryset.filter(Num_fact__icontains=Num_fact)
            
            # Filtre sur le montant
            montant_min = form.cleaned_data.get('montant_min')
            montant_max = form.cleaned_data.get('montant_max')
            if montant_min is not None:
                queryset = queryset.filter(montant__gte=montant_min)
            if montant_max is not None:
                queryset = queryset.filter(montant__lte=montant_max)
            
            # Filtre sur l'auteur
            auteur = form.cleaned_data.get('auteur')
            if auteur:
                queryset = queryset.filter(auteur=auteur)
            
            # Filtre sur le type d'historique (Créé, Modifié, Supprimé)
            history_type = form.cleaned_data.get('history_type')
            if history_type:
                queryset = queryset.filter(history_type=history_type)
        
        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historique Charges Variables"
        
        # En-têtes de colonnes
        headers = [
            'Type',
            'Immatriculation',
            'Marque',
            'Catégorie',
            'Libellé',
            'Compte Comptable',
            'N° Pièce',
            'N° Facture',
            'Montant',
            'Date Saisie',
            'Auteur',
            'Date Historique',
            'Utilisateur Historique'
        ]
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[col_letter].width = 20
 
        # Données
        for h in queryset:
            history_type_label = ''
            if h.history_type == '+':
                history_type_label = 'Créé'
            elif h.history_type == '~':
                history_type_label = 'Modifié'
            elif h.history_type == '-':
                history_type_label = 'Supprimé'
            
            ws.append([
                history_type_label,
                h.vehicule.immatriculation if h.vehicule else 'N/A',
                h.vehicule.marque if h.vehicule else 'N/A',
                h.vehicule.category.category if h.vehicule and h.vehicule.category else 'N/A',
                h.libelle or 'N/A',
                h.cpte_comptable or 'N/A',
                h.Num_piece or 'N/A',
                h.Num_fact or 'N/A',
                h.montant,
                h.date_saisie.strftime("%d-%m-%Y") if h.date_saisie else 'N/A',
                h.auteur.username if h.auteur else 'N/A',
                h.history_date.strftime("%d-%m-%Y %H:%M") if h.history_date else 'N/A',
                h.history_user.username if h.history_user else 'N/A'
            ])
        
        # Réponse HTTP (fichier Excel téléchargeable)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Historique-Charges-Variables.xlsx'
        wb.save(response)
        return response

class AddChargeAdminisView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_chargadminist'
    model = ChargeAdminis
    form_class = ChargeAdminisForm
    template_name = 'perfect/add_charg_admin.html'
    success_message = 'Charge administrative enregistrée avec succès✓✓'
    error_message = "Erreur de saisie ✘✘ "
    success_url = reverse_lazy ('add_chargadminist')
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        reponse =  super().form_valid(form)
        messages.success(self.request, self.success_message)
        return reponse
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        forms = self.get_form()
        annee_en_cours =date.today().year
        today = date.today()
        mois_en_cours =date.today().month

        libelle_mois_en_cours = calendar.month_name[mois_en_cours]

        # Querysets de base (pour stats annuelles)
        base_chargeadmin = ChargeAdminis.objects.all()
        base_recette = Recette.objects.all()
        base_chargvar = ChargeVariable.objects.all()
        base_chargfix = ChargeFixe.objects.all()
        base_piece = Piece.objects.all()
        base_piechan = PiecEchange.objects.all()

        date_debut = date_fin = None
        form = DateForm(self.request.GET)
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        # Filtre pour le mois en cours ou la plage de dates
        if date_debut and date_fin:
            filtre_chargeadmin = base_chargeadmin.filter(date_saisie__range=[date_debut, date_fin])
            filtre_recette = base_recette.filter(date_saisie__range=[date_debut, date_fin])
            filtre_piece = base_piece.filter(date_saisie__range=[date_debut, date_fin])
            filtre_piec_echange = base_piechan.filter(date_saisie__range=[date_debut, date_fin])
            filtre_chargfix = base_chargfix.filter(date_saisie__range=[date_debut, date_fin])
            filtre_chargvar = base_chargvar.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_chargeadmin = base_chargeadmin.filter(date_saisie__month=today.month, date_saisie__year=today.year)
            filtre_recette = base_recette.filter(date_saisie__month=today.month, date_saisie__year=today.year)
            filtre_piece = base_piece.filter(date_saisie__month=today.month, date_saisie__year=today.year)
            filtre_piec_echange = base_piechan.filter(date_saisie__month=today.month, date_saisie__year=today.year)
            filtre_chargfix = base_chargfix.filter(date_saisie__month=today.month, date_saisie__year=today.year)
            filtre_chargvar = base_chargvar.filter(date_saisie__month=today.month, date_saisie__year=today.year)

        # ---------- STATS ANNUELLES (info boxes) ----------
        base_an = base_recette.filter(date_saisie__year=today.year)
        base_chargadmin_an = base_chargeadmin.filter(date_saisie__year=today.year)
        base_chargfix_an = base_chargfix.filter(date_saisie__year=today.year)
        base_chargvar_an = base_chargvar.filter(date_saisie__year=today.year)

        total_recette_annuel = base_an.aggregate(somme=Sum('montant'))['somme'] or 0
        total_recette_annuel_format = '{:,}'.format(total_recette_annuel).replace(',', ' ')
        total_chargeadmin_annuel = base_chargadmin_an.aggregate(somme=Sum('montant'))['somme'] or 0
        total_chargeadmin_annuel_format = '{:,}'.format(total_chargeadmin_annuel).replace(',', ' ')
        total_chargfix_annuel = base_chargfix_an.aggregate(somme=Sum('montant'))['somme'] or 0
        total_chargvar_annuel = base_chargvar_an.aggregate(somme=Sum('montant'))['somme'] or 0

        marge_contribution_annuel = total_recette_annuel - total_chargvar_annuel
        
        marge_contribution_annuel_format = '{:,}'.format(marge_contribution_annuel).replace(',', ' ')
        total_charg_annuel = total_chargfix_annuel + total_chargvar_annuel
        marge_brute_annuel = total_recette_annuel - total_charg_annuel
        resultat_annuel = marge_brute_annuel - total_chargeadmin_annuel
        resultat_annuel_format = '{:,}'.format(resultat_annuel).replace(',', ' ')

        # ---------- STATS MENSUELLES (statistique mensuelle + tableau) ----------
        total_chargeadmin_mois = filtre_chargeadmin.aggregate(somme=Sum('montant'))['somme'] or 0
        total_chargeadmin_mois_format = '{:,}'.format(total_chargeadmin_mois).replace(',', ' ')

        list_charge_adminis = filtre_chargeadmin.order_by('-date_saisie')

        total_recettes_mois = filtre_recette.aggregate(somme=Sum('montant'))['somme'] or 0
        total_recette_mois_format = '{:,}'.format(total_recettes_mois).replace(',', ' ')
        total_piece = filtre_piece.aggregate(somme=Sum('montant'))['somme'] or 0
        total_piece_format = '{:,}'.format(total_piece).replace(',', ' ')
        total_piec_echange = filtre_piec_echange.aggregate(somme=Sum('montant'))['somme'] or 0
        total_piec_echange_format = '{:,}'.format(total_piec_echange).replace(',', ' ')
        total_chargfix_mois = filtre_chargfix.aggregate(somme=Sum('montant'))['somme'] or 0
        total_chargfix_format = '{:,}'.format(total_chargfix_mois).replace(',', ' ')
        total_chargvar_mois = filtre_chargvar.aggregate(somme=Sum('montant'))['somme'] or 0
        total_chargvar_format = '{:,}'.format(total_chargvar_mois).replace(',', ' ')

        total_charg = total_chargfix_mois + total_chargvar_mois
        total_charge_format = '{:,}'.format(total_charg).replace(',', ' ')
        marge_contribution = total_recettes_mois - total_chargvar_mois
        marge_brute = total_recettes_mois - total_charg
        marge_brute_format = '{:,}'.format(marge_brute).replace(',', ' ')
        resultat = marge_brute - total_chargeadmin_mois
        resultat_format = '{:,}'.format(resultat).replace(',', ' ')

        charg_admin_data = base_chargeadmin.filter(date_saisie__year=date.today().year)
        chargadmin_mois_data = {month: 0 for month in range(1, 13)}
        for commande in charg_admin_data:
            chargadmin_mois_data[commande.date_saisie.month] += commande.montant
        chargadmin_mois_data = [chargadmin_mois_data[month] for month in range(1, 13)]
        label = [calendar.month_name[month][:1] for month in range(1, 13)]
        context={
            'total_recette_format': total_recette_mois_format,
            'total_recette_annuel_format': total_recette_annuel_format,
            'marge_contribution_annuel_format': marge_contribution_annuel_format,
            'list_charge_adminis': list_charge_adminis,
            'total_chargfix_format': total_chargfix_format,
            'total_chargvar_format': total_chargvar_format,
            'total_piece_format': total_piece_format,
            'total_piec_echange_format': total_piec_echange_format,
            'marge_brute_format': marge_brute_format,
            'total_charge_format': total_charge_format,
            'label': label,
            'forms': forms,
            'form': form,
            'resultat_format': resultat_format,
            'resultat_annuel_format': resultat_annuel_format,
            'dates': today,
            'mois': libelle_mois_en_cours,
            'annee': annee_en_cours,
            'total_chargeadmin_format': total_chargeadmin_mois_format,
            'total_chargeadmin_annuel_format': total_chargeadmin_annuel_format,
            'chargadmin_data': chargadmin_mois_data,
        }  
        return context   
    
def delete_chargadmin(request, pk):
    try:
        chargadmin = get_object_or_404(ChargeAdminis, id=pk)
        chargadmin.delete()
        messages.success(request, f"la charge administrative {chargadmin.Num_fact}-{chargadmin.date_saisie} a été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('add_chargadminist')

class ExportChargeAdminisExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Récupérer les paramètres GET
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')

        # Récupérer les charges
        charges = ChargeAdminis.objects.all().order_by('-date_saisie')

        # Appliquer le filtre uniquement si les deux dates sont présentes
        if date_debut and date_fin:
            try:
                charges = charges.filter(date_saisie__range=[date_debut, date_fin])
            except Exception:
                pass  # ignore les erreurs de format de date

        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Charges administratives"

        # En-têtes
        headers = ["N° Facture","Libellé", "Montant", "Compte Comptable", "N° Pièce", "Date Saisie",  "Auteur"]
        ws.append(headers)

        # Lignes de données
        for c in charges:
            ws.append([
                c.Num_fact,
                c.libelle,
                c.montant,
                c.cpte_comptable,
                c.Num_piece,
                c.date_saisie.strftime("%d-%m-%Y") if c.date_saisie else "",
                c.auteur.username if c.auteur else "N/A",
            ])

        # Ajuster la largeur des colonnes
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22

        # Style de la première ligne (gras et couleur)
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")

        # Nom du fichier
        filename = f"Charges_Admin_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

        # Réponse HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

#--------------/-/---------------@-----------------/-/--------------Garage---------------/-/--------------@----------------/-/------------#

class AddCartStationnementView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_station'
    model = Stationnement
    form_class = CartStationForm
    template_name= "perfect/add_station.html"
    success_message = 'Carte de Stationnement enregistrée avec succès✓✓'
    error_message = "Erreur de saisie✘✘"
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)
    def form_invalid(self, form):
        messages.success(self.request,self.error_message)
        return super().form_invalid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        station_queryset = Stationnement.objects.filter(vehicule=vehicule)

        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None

        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none()
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        search_query = self.request.GET.get("search", "").strip()
        if search_query:
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()

        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        today = date.today()
        station_base = Stationnement.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_station = station_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_station = station_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        station_jours = filtre_station.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        station_jours_format = '{:,}'.format(station_jours).replace(',', ' ')
        station_mois = filtre_station.aggregate(somme=Sum('montant'))['somme'] or 0
        station_mois_format = '{:,}'.format(station_mois).replace(',', ' ')
        station_an = station_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        station_an_format = '{:,}'.format(station_an).replace(',', ' ')
        liste_station = filtre_station.order_by('-date_saisie')

        # Permissions groupées pour le template (liens Assurance, Vignette, etc.)
        grouped_permissions = {}
        if hasattr(user, 'custom_permissions'):
            permissions = user.custom_permissions.all().select_related('categorie')
            for perm in permissions:
                if perm.categorie not in grouped_permissions:
                    grouped_permissions[perm.categorie] = []
                grouped_permissions[perm.categorie].append(perm)

        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            'station_jours_format': station_jours_format,
            'station_mois_format': station_mois_format,
            'station_an_format': station_an_format,
            'liste_station': liste_station,
            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
            'search_query': search_query,
            'grouped_permissions': grouped_permissions,
        }
        return context  
    def get_success_url(self):
        return reverse('add_station', kwargs={'pk': self.kwargs['pk']})

class ListCartStationView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'liste_station'
    model = Stationnement
    template_name = 'perfect/liste_station.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        station_base = Stationnement.objects.all()
        form = self.form_class(self.request.GET)
        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            if categorie_filter:
                station_base = station_base.filter(vehicule__category__category=categorie_filter)

            if immatriculation:
                station_base = station_base.filter(vehicule__immatriculation__icontains=immatriculation)

            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListRecetView)
            if date_debut and date_fin:
                filtre_station = station_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_station = station_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_station = station_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ STATISTIQUES -------------------
        # Jours / Mois : sur la période affichée (filtre_station)
        station_jours = filtre_station.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        station_jours_format = '{:,}'.format(station_jours).replace(',', ' ')
        station_mois = filtre_station.aggregate(somme=Sum('montant'))['somme'] or 0
        station_mois_format = '{:,}'.format(station_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        station_an = station_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        station_an_format = '{:,}'.format(station_an).replace(',', ' ')

        context = {
            'liste_station': filtre_station,
            'station_an_format': station_an_format,
            'station_mois_format': station_mois_format,
            'station_jours_format': station_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

class UpdatCartStationView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_station'
    model = Stationnement
    form_class = UpdatCartStationForm
    template_name = "perfect/partials/station_form.html"
    success_url = reverse_lazy('liste_station')
    success_message = 'Carte de stationnement modifiée avec succès✓✓'
    timeout_minutes = 600
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class ExportStationnementExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        stationnements = Stationnement.objects.all().select_related('vehicule', 'auteur')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        today = date.today()

        # Filtre par date : plage fournie par l'utilisateur, sinon mois en cours par défaut
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                stationnements = stationnements.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                stationnements = stationnements.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            stationnements = stationnements.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        if immatriculation:
            stationnements = stationnements.filter(vehicule__immatriculation__icontains=immatriculation)
        if categorie:
            stationnements = stationnements.filter(vehicule__category__category__icontains=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stationnements"

        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Montant (FCFA)",
            "Date Saisie",
            "Date Prochaine",
            "Jours Restants",
            "Auteur"
        ]
        ws.append(headers)

        # 🔹 Remplissage des données
        for s in stationnements:
            ws.append([
                s.vehicule.immatriculation if s.vehicule else "",
                s.vehicule.marque if s.vehicule else "",
                s.vehicule.category.category if s.vehicule and s.vehicule.category else "",
                s.montant,
                s.date_saisie.strftime("%d-%m-%Y") if s.date_saisie else "",
                s.date_proch.strftime("%d-%m-%Y") if s.date_proch else "",
                s.jours_cartsta_restant if s.date_proch else "",
                s.auteur.username if s.auteur else "N/A"
            ])

        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Stationnements.xlsx'

        wb.save(response)
        return response

class AddPatenteView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_patente'
    model = Patente
    form_class = PatenteForm
    template_name= "perfect/add_patente.html"
    success_message = 'Patente enregistrée avec succès✓✓'
    error_message = "Erreur de saisie✘✘"
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)
    def form_invalid(self, form):
        messages.success(self.request,self.error_message)
        return super().form_invalid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        patent_queryset = Patente.objects.filter(vehicule=vehicule)

        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None

        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        today = date.today()
        patente_base = Patente.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_patente = patente_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_patente = patente_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        patente_jours = filtre_patente.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        patente_jours_format = '{:,}'.format(patente_jours).replace(',', ' ')
        patente_mois = filtre_patente.aggregate(somme=Sum('montant'))['somme'] or 0
        patente_mois_format = '{:,}'.format(patente_mois).replace(',', ' ')
        patente_an = patente_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        patente_an_format = '{:,}'.format(patente_an).replace(',', ' ')
        liste_patente = filtre_patente.order_by('-date_saisie')

        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            'patente_jours_format': patente_jours_format,
            'patente_mois_format': patente_mois_format,
            'patente_an_format': patente_an_format,
            'liste_patente': liste_patente,
            'form': form,
            'forms': forms,
        }   
        return context  
    def get_success_url(self):
        return reverse('add_patente', kwargs={'pk': self.kwargs['pk']})

@require_POST
@login_required
def delete_selected_stat(request):
    selected_ids = request.POST.getlist('stat_ids')
    if selected_ids:
        stats_to_delete = Stationnement.objects.filter(id__in=selected_ids)
        count = stats_to_delete.count()
        stats_to_delete.delete()
        messages.success(request, f"{count} carte(s) de stationnement supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune carte de stationnement sélectionnée.")
    return redirect('liste_station')

class ListPatenteView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'liste_patente'
    model = Patente
    template_name = 'perfect/liste_patente.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        patente_base = Patente.objects.all()
        form = self.form_class(self.request.GET)
        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            if categorie_filter:
                patente_base = patente_base.filter(vehicule__category__category=categorie_filter)
            if immatriculation:
                patente_base = patente_base.filter(vehicule__immatriculation__icontains=immatriculation)
            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListRecetView)
            if date_debut and date_fin:
                filtre_patente = patente_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_patente = patente_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_patente = patente_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ STATISTIQUES -------------------
        # Jours / Mois : sur la période affichée (filtre_patente)
        patente_jours = filtre_patente.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        patente_jours_format = '{:,}'.format(patente_jours).replace(',', ' ')
        patente_mois = filtre_patente.aggregate(somme=Sum('montant'))['somme'] or 0
        patente_mois_format = '{:,}'.format(patente_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        patente_an = patente_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        patente_an_format = '{:,}'.format(patente_an).replace(',', ' ')
        context = {
            'liste_patente': filtre_patente,
            'patente_an_format': patente_an_format,
            'patente_mois_format': patente_mois_format,
            'patente_jours_format': patente_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

@require_POST
@login_required
def delete_selected_patente(request):
    selected_ids = request.POST.getlist('patente_ids')
    if selected_ids:
        patentes_to_delete = Patente.objects.filter(id__in=selected_ids)
        count = patentes_to_delete.count()
        patentes_to_delete.delete()
        messages.success(request, f"{count} patente(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune patente sélectionnée.")
    return redirect('liste_patente')

class ExportPatenteExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        patentes = Patente.objects.all().select_related('vehicule', 'auteur')

        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                patentes = patentes.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pass
        if immatriculation:
            patentes = patentes.filter(vehicule__immatriculation__icontains=immatriculation)
        if categorie:
            patentes = patentes.filter(vehicule__category__category__icontains=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patentes"
        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Montant",
            "Date Saisie",
            "Prochaine Échéance",
            "Jours Restants",
            "Auteur"
        ]
        ws.append(headers)

        for p in patentes:
            jours_restants = (p.date_proch - timezone.now().date()).days if p.date_proch else ""
            ws.append([
                p.vehicule.immatriculation if p.vehicule else "",
                p.vehicule.marque if p.vehicule else "",
                p.vehicule.category.category if p.vehicule and p.vehicule.category else "",
                p.montant,
                p.date_saisie.strftime("%d-%m-%Y") if p.date_saisie else "",
                p.date_proch.strftime("%d-%m-%Y") if p.date_proch else "",
                jours_restants,
                p.auteur.username if p.auteur else "N/A"
            ])

        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Patentes.xlsx'
        wb.save(response)
        return response

class AddVignetteView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_vignet'
    model = Vignette
    form_class = VignetteForm
    template_name= "perfect/add_vignette.html"
    success_message = 'Vignette enregistrée avec succès✓✓'
    error_message = "Erreur de saisie✘✘"
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)
    def form_invalid(self, form):
        messages.success(self.request,self.error_message)
        return super().form_invalid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        vignette_queryset = Vignette.objects.filter(vehicule=vehicule)

        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None

        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()
        
        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()

        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        today = date.today()
        vignette_base = Vignette.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_vignette = vignette_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_vignette = vignette_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        vignette_jours = filtre_vignette.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        vignette_jours_format = '{:,}'.format(vignette_jours).replace(',', ' ')
        vignette_mois = filtre_vignette.aggregate(somme=Sum('montant'))['somme'] or 0
        vignette_mois_format = '{:,}'.format(vignette_mois).replace(',', ' ')
        vignette_an = vignette_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        vignette_an_format = '{:,}'.format(vignette_an).replace(',', ' ')
        liste_vignette = filtre_vignette.order_by('-date_saisie')

        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            'vignette_jours_format':vignette_jours_format,
            'vignette_mois_format':vignette_mois_format,
            'vignette_an_format':vignette_an_format,
            'liste_vignette':liste_vignette,
            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
        }     
        return context  
    def get_success_url(self):
        return reverse('add_vignet', kwargs={'pk': self.kwargs['pk']})

class ListVignetteView(LoginRequiredMixin, CustomPermissionRequiredMixin,ListView):
    login_url = 'login'
    permission_url = 'liste_vignette'
    model = Vignette
    template_name = 'perfect/liste_vignette.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        vignet_base = Vignette.objects.all()
        form = self.form_class(self.request.GET)
        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            if categorie_filter:
                vignet_base = vignet_base.filter(vehicule__category__category=categorie_filter)

            if immatriculation:
                vignet_base = vignet_base.filter(vehicule__immatriculation__icontains=immatriculation)

            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListRecetView)
            if date_debut and date_fin:
                filtre_vignet = vignet_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_vignet = vignet_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_vignet = vignet_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ STATISTIQUES -------------------
        # Jours / Mois : sur la période affichée (filtre_vignet)
        vignet_jours = filtre_vignet.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        vignet_jours_format = '{:,}'.format(vignet_jours).replace(',', ' ')
        vignet_mois = filtre_vignet.aggregate(somme=Sum('montant'))['somme'] or 0
        vignet_mois_format = '{:,}'.format(vignet_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        vignet_an = vignet_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        vignet_an_format = '{:,}'.format(vignet_an).replace(',', ' ')
        context = {
            'liste_vignet': filtre_vignet,
            'vignet_an_format': vignet_an_format,
            'vignet_mois_format': vignet_mois_format,
            'vignet_jours_format': vignet_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

@require_POST
@login_required
def delete_selected_vignette(request):
    selected_ids = request.POST.getlist('vignette_ids')
    if selected_ids:
        vignet_to_delete = Vignette.objects.filter(id__in=selected_ids)
        count = vignet_to_delete.count()
        vignet_to_delete.delete()
        messages.success(request, f"{count} vignette(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune vignette sélectionnée.")
    return redirect('liste_vignette')

class UpdatVignetteView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_vignet'
    model = Vignette
    form_class = UpdatVignetteForm
    template_name = "perfect/partials/vignet_form.html"
    success_url = reverse_lazy('liste_vignette')
    success_message = 'Vignette modifiée avec succès✓✓'
    timeout_minutes = 200
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class ExportVignetteExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        vignettes = Vignette.objects.all().select_related('vehicule', 'auteur')

        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                vignettes = vignettes.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pass
        if immatriculation:
            vignettes = vignettes.filter(vehicule__immatriculation__icontains=immatriculation)

        if categorie:
            vignettes = vignettes.filter(vehicule__category__category__icontains=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vignettes"
        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Montant (FCFA)",
            "Date Saisie",
            "Date Prochaine",
            "Jours Restants",
            "Auteur"
        ]
        ws.append(headers)
        for v in vignettes:
            ws.append([
                v.vehicule.immatriculation if v.vehicule else "",
                v.vehicule.marque if v.vehicule else "",
                v.vehicule.category.category if v.vehicule and v.vehicule.category else "",
                v.montant,
                v.date_saisie.strftime("%d-%m-%Y") if v.date_saisie else "",
                v.date_proch.strftime("%d-%m-%Y") if v.date_proch else "",
                v.jours_vign_restant if v.date_proch else "",
                v.auteur.username if v.auteur else "N/A"
            ])

        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Vignettes.xlsx'
        wb.save(response)
        return response

class AddVisitView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_visit'
    model = VisiteTechnique
    form_class = VisiteTechniqueForm
    template_name= "perfect/add_visite.html"
    success_message = 'Visite Ajoutée avec succès ✓✓'
    error_message = "Erreur de saisie ✘✘ "
    # success_url = reverse_lazy('journal_compta')
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)  
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        now = datetime.now()
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        visitech_queryset = VisiteTechnique.objects.filter(vehicule=vehicule)
        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.filter(category="VTC").exists():
                    vehicules = Vehicule.objects.filter(category__category="VTC")
                else:
                    vehicules = Vehicule.objects.filter(category__category="TAXI")
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()  
                # No vehicles if no Gerant linked
        elif user:
            try:
                vehicules = Vehicule.objects.all()
            except:
                vehicules = Vehicule.objects.none() 
        else:
            print()

        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()

        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        today = date.today()
        visitech_base = VisiteTechnique.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_visitech = visitech_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_visitech = visitech_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        visitech_jours = filtre_visitech.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        visitech_jours_format = '{:,}'.format(visitech_jours).replace(',', ' ')
        visitech_mois = filtre_visitech.aggregate(somme=Sum('montant'))['somme'] or 0
        visitech_mois_format = '{:,}'.format(visitech_mois).replace(',', ' ')
        visitech_an = visitech_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        visitech_an_format = '{:,}'.format(visitech_an).replace(',', ' ')
        liste_visitech = filtre_visitech.order_by('-date_saisie')
        
        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            'liste_visitech':liste_visitech,
            'visitech_an_format':visitech_an_format,
            'visitech_mois_format':visitech_mois_format,
            'visitech_jours_format':visitech_jours_format,
            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
        }   
        return context  
    def get_success_url(self):
        return reverse('add_visit', kwargs={'pk': self.kwargs['pk']})

class ListVisitTechniqueView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_visit'
    model = VisiteTechnique
    template_name = 'perfect/liste_visites.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        visitech_base = VisiteTechnique.objects.all()
        form = self.form_class(self.request.GET)
        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            if categorie_filter:
                visitech_base = visitech_base.filter(vehicule__category__category=categorie_filter)
            if immatriculation:
                visitech_base = visitech_base.filter(vehicule__immatriculation__icontains=immatriculation)
            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListReparationView)
            if date_debut and date_fin:
                filtre_visitech = visitech_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_visitech = visitech_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_visitech = visitech_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ STATISTIQUES -------------------
        # Jours / Mois : sur la période affichée (filtre_visitech)
        visitech_jours = filtre_visitech.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        visitech_jours_format = '{:,}'.format(visitech_jours).replace(',', ' ')
        visitech_mois = filtre_visitech.aggregate(somme=Sum('montant'))['somme'] or 0
        visitech_mois_format = '{:,}'.format(visitech_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListReparationView
        visitech_an = visitech_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        visitech_an_format = '{:,}'.format(visitech_an).replace(',', ' ')
        #------------------------Permissions groupées pour le template (boutons Exporter, sup visite, modifier visite)------------------------#
        grouped_permissions = {}
        user = self.request.user
        if hasattr(user, 'custom_permissions'):
            permissions = user.custom_permissions.all().select_related('categorie')
            for perm in permissions:
                if perm.categorie not in grouped_permissions:
                    grouped_permissions[perm.categorie] = []
                grouped_permissions[perm.categorie].append(perm)
        context.update({
            'liste_visitech': filtre_visitech,
            'visitech_an_format': visitech_an_format,
            'visitech_mois_format': visitech_mois_format,
            'visitech_jours_format': visitech_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
            'grouped_permissions': grouped_permissions,
        })
        return context

class UpdateVisiteView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_visit'
    model = VisiteTechnique
    form_class = UpdatVisiteTechniqueForm
    template_name = "perfect/partials/visite_form.html"
    success_url = reverse_lazy('list_visit')
    success_message = 'Visite Techniques modifiée avec succès✓✓'

    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class ExportVisiteTechniqueExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        visites = VisiteTechnique.objects.all()
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        today = date.today()

        # Filtre par date : plage fournie par l'utilisateur, sinon mois en cours par défaut
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                visites = visites.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                visites = visites.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            visites = visites.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        if immatriculation:
            visites = visites.filter(vehicule__immatriculation__icontains=immatriculation)
        if categorie:
            visites = visites.filter(vehicule__category__category__icontains=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visites Techniques"
        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Date Visite",
            "Date Sortie",
            "Date Prochaine",
            "Montant",
            "Date Saisie",
            "Auteur"
        ]
        ws.append(headers)

        for v in visites:
            ws.append([
                v.vehicule.immatriculation if v.vehicule else "",
                v.vehicule.marque if v.vehicule else "",
                v.vehicule.category.category if v.vehicule and v.vehicule.category else "",
                v.date_vis.strftime("%d-%m-%Y %H:%M") if v.date_vis else "",
                v.date_sortie.strftime("%d-%m-%Y %H:%M") if v.date_sortie else "",
                v.date_proch.strftime("%d-%m-%Y") if v.date_proch else "",
                v.montant,
                v.date_saisie.strftime("%d-%m-%Y") if v.date_saisie else "",
                v.auteur.username if v.auteur else "N/A"
            ])

        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Visites-Techniques.xlsx'
        wb.save(response)
        return response

class AddAssuranceView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_assurance'
    model = Assurance
    form_class = AssuranceForm
    template_name= "perfect/add_assurance.html"
    success_message = 'Assurance enregistré avec succès✓✓'
    error_message = "Erreur de saisie✘✘"
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)
    def form_invalid(self, form):
        messages.success(self.request,self.error_message)
        return super().form_invalid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        assurance_queryset = Assurance.objects.filter(vehicule=vehicule)

        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None

        if user.user_type == "4":
            try:
                gerant = user.gerants.get()  
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')

        today = date.today()
        assurance_base = Assurance.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_assurance = assurance_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_assurance = assurance_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        assurance_jours = filtre_assurance.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        assurance_jours_format = '{:,}'.format(assurance_jours).replace(',', ' ')
        assurance_mois = filtre_assurance.aggregate(somme=Sum('montant'))['somme'] or 0
        assurance_mois_format = '{:,}'.format(assurance_mois).replace(',', ' ')
        assurance_an = assurance_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        assurance_an_format = '{:,}'.format(assurance_an).replace(',', ' ')
        liste_recette = filtre_assurance.order_by('-date_saisie')

        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            "assurance_an_format": assurance_an_format,
            "assurance_mois_format": assurance_mois_format,
            "assurance_jours_format": assurance_jours_format,
            "liste_recette": liste_recette,
            
            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
        }    
        return context  
    def get_success_url(self):
        return reverse('add_assurance', kwargs={'pk': self.kwargs['pk']})

class ListAssuranceView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'liste_assurance'
    model = Assurance
    template_name = 'perfect/liste_assurance.html'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        assur_base = Assurance.objects.all()
        form = self.form_class(self.request.GET)
        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')

            if categorie_filter:
                assur_base = assur_base.filter(vehicule__category__category=categorie_filter)

            if immatriculation:
                assur_base = assur_base.filter(vehicule__immatriculation__icontains=immatriculation)

            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListRecetView)
            if date_debut and date_fin:
                filtre_assur = assur_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_assur = assur_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_assur = assur_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        # ------------------ STATISTIQUES -------------------
        # Jours / Mois : sur la période affichée (filtre_assur)
        assur_jours = filtre_assur.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        assur_jours_format = '{:,}'.format(assur_jours).replace(',', ' ')
        assur_mois = filtre_assur.aggregate(somme=Sum('montant'))['somme'] or 0
        assur_mois_format = '{:,}'.format(assur_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        assur_an = assur_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        assur_an_format = '{:,}'.format(assur_an).replace(',', ' ')

        context = {
            'liste_assur': filtre_assur,
            'assur_an_format': assur_an_format,
            'assur_mois_format': assur_mois_format,
            'assur_jours_format': assur_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

@require_POST
@login_required
def delete_selected_assurance(request):
    selected_ids = request.POST.getlist('assurance_ids')
    if selected_ids:
        assurance_to_delete = Assurance.objects.filter(id__in=selected_ids)
        count = assurance_to_delete.count()
        assurance_to_delete.delete()
        messages.success(request, f"{count} assurance(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune assurance sélectionnée.")
    return redirect('liste_assurance')

class UpdateAssuranceView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_assurance'
    model = Assurance
    form_class = UpdatAssuranceForm
    template_name = "perfect/partials/assurance_form.html"
    success_url = reverse_lazy('liste_assurance')
    success_message = 'Assurance modifiée avec succès ✓✓'
    timeout_minutes = 600
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class ExportAssuranceExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # assurances = Assurance.objects.all().select_related('vehicule', 'auteur')
        assurances = Assurance.objects.all()
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                assurances = assurances.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pass
        if immatriculation:
            assurances = assurances.filter(vehicule__immatriculation__icontains=immatriculation)
        if categorie:
            assurances = assurances.filter(vehicule__category__category__icontains=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assurances"

        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Montant",
            "Date Saisie",
            "Date Prochaine",
            "Jours Restants",
            "Auteur"
        ]
        ws.append(headers)

        for a in assurances:
            ws.append([
                a.vehicule.immatriculation if a.vehicule else "",
                a.vehicule.marque if a.vehicule else "",
                a.vehicule.category.category if a.vehicule and a.vehicule.category else "",
                a.montant,
                a.date_saisie.strftime("%d-%m-%Y") if a.date_saisie else "",
                a.date_proch.strftime("%d-%m-%Y") if a.date_proch else "",
                a.jours_assu_restant if a.date_proch else "",
                a.auteur.username if a.auteur else "N/A"
            ])

        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Assurances.xlsx'
        wb.save(response)
        return response

class AddReparationView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_reparation'
    model = Reparation
    form_class = ReparationForm
    template_name= "perfect/add_reparation.html"
    success_message = 'Réparation enregistrée avec succès✓✓'
    error_message = "Erreur de saisie ✘✘"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        # now = datetime.now()
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        form = DateFormRepar(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        if self.request.POST:
            piece_formset = PieceFormSet(self.request.POST, instance=self.object)
        else:
            piece_formset = PieceFormSet(instance=self.object)
        forms = self.get_form()
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()  
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()

        reparat_base = Reparation.objects.filter(vehicule=vehicule)
        date_debut = date_fin = None
        today = date.today()
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            motif_filter = self.request.GET.get('motif')

            if motif_filter:
                reparat_base = reparat_base.filter(motif=motif_filter)
        if date_debut and date_fin:
            filtre_reparat = reparat_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_reparat = reparat_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        reparat_total = filtre_reparat.count()
        reparat_jours = filtre_reparat.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_jours_format = '{:,}'.format(reparat_jours).replace(',', ' ')
        reparat_mois = filtre_reparat.aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_mois_format = '{:,}'.format(reparat_mois).replace(',', ' ')
        reparat_an = reparat_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_an_format = '{:,}'.format(reparat_an).replace(',', ' ')
        liste_reparat = filtre_reparat.order_by('-date_saisie')
        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            "piece_formset": piece_formset,
            "liste_reparat": liste_reparat,
            
            "reparat_total": reparat_total,
            "reparat_jours_format": reparat_jours_format,
            "reparat_mois_format": reparat_mois_format,
            "reparat_an_format": reparat_an_format,

            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
        }
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        piece_formset = context['piece_formset']
        # Associer véhicule et auteur AVANT la validation finale
        form.instance.vehicule_id = self.kwargs['pk']
        form.instance.auteur = self.request.user  

        if form.is_valid() and piece_formset.is_valid():
            self.object = form.save(commit=False)
            total_pieces = 0
            for piece_form in piece_formset:
                piece = piece_form.save(commit=False)
                if piece.libelle:  # éviter les formulaires vides
                    piece.auteur = self.request.user  
                    piece.reparation = self.object
                    total_pieces += (piece.montant or 0) * (piece.quantite or 1)

            # calcul du montant total
            prestation = form.cleaned_data.get("prestation", 0)
            self.object.montant = total_pieces + prestation
            self.object.save()
            # maintenant on rattache les pièces à la réparation
            piece_formset.instance = self.object
            piece_formset.save()

            messages.success(self.request, self.success_message)
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)
    def form_invalid(self, form):
        context = self.get_context_data()
        piece_formset = context['piece_formset']
        print("=== ERREURS REPARATION ===")
        print(form.errors)
        print("=== ERREURS PIECES ===")
        print(piece_formset.errors)
        messages.error(self.request, self.error_message)
        return super().form_invalid(form)
    def get_success_url(self):
        return reverse('add_reparation', kwargs={'pk': self.kwargs['pk']})

class UpdateReparationView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'update_reparation'
    model = Reparation
    form_class = ReparationForm
    template_name = "perfect/update_reparation.html"
    success_message = 'Réparation modifiée avec succès✓✓'
    error_message = "Erreur lors de la modification ✘✘"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['piece_formset'] = PieceFormSet(self.request.POST, self.request.FILES or None, instance=self.object)
        else:
            context['piece_formset'] = PieceFormSet(instance=self.object)
        context['vehicule'] = self.object.vehicule
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('partial'):
            context = self.get_context_data(object=self.object)
            return render(request, 'perfect/partials/reparation_update_form.html', context)
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        context = self.get_context_data()
        piece_formset = context['piece_formset']
        form.instance.auteur = self.request.user

        if form.is_valid() and piece_formset.is_valid():
            with transaction.atomic():
                self.object = form.save(commit=False)
                self.object.save()

                for piece_form in piece_formset:
                    piece = piece_form.save(commit=False)
                    if piece.libelle:
                        piece.auteur = self.request.user
                        piece.reparation = self.object
                piece_formset.instance = self.object
                piece_formset.save()

                total_pieces = 0
                for piece in self.object.pieces.all():
                    total_pieces += (piece.montant or 0) * (piece.quantite or 1)
                prestation = form.cleaned_data.get("prestation", 0)
                self.object.montant = total_pieces + prestation
                self.object.save()

            messages.success(self.request, self.success_message)
            return redirect(self.get_success_url())

        return self.form_invalid(form)

    def form_invalid(self, form):
        context = self.get_context_data()
        piece_formset = context.get("piece_formset")
        print("=== ERREURS UPDATE REPARATION ===")
        print(form.errors)
        if piece_formset is not None:
            print("=== ERREURS PIECES (UPDATE) ===")
            print(piece_formset.errors)
        messages.error(self.request, self.error_message)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'piece_errors': piece_formset.errors if piece_formset else {},
            }, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse("list_repa")

@require_POST
@login_required
def delete_selected_reparation(request):
    selected_ids = request.POST.getlist('reparation_ids')
    if selected_ids:
        reparation_to_delete = Reparation.objects.filter(id__in=selected_ids)
        count = reparation_to_delete.count()
        reparation_to_delete.delete()
        messages.success(request, f"{count} réparation(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune réparation sélectionnée.")
    return redirect('list_repa') 

class AddPiecEchangeView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_piechange'
    model = PiecEchange
    form_class = PiecEchangeForm
    template_name = "perfect/add_piecechange.html"
    success_message = 'Pièces enregistrées avec succès✓✓'
    error_message = "Erreur de saisie✘✘"
    timeout_minutes = 500

    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        context = self.get_context_data()
        lignes = context['lignes_formset']
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']

        with transaction.atomic():
            self.object = form.save(commit=False)
            self.object.save()

            if lignes.is_valid():
                lignes.instance = self.object
                lignes.save()
            # 🔹 Calcul du montant total
            total = sum([ligne.prix_total for ligne in self.object.lignes.all()])
            self.object.montant = total
            self.object.save()

        messages.success(self.request, self.success_message)
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        context = self.get_context_data()
        lignes = context['lignes_formset']
        print("=== ERREURS PIECHANGE ===")
        print(form.errors)
        print("=== ERREURS LIGNES ===")
        print(lignes.errors)
        messages.error(self.request, self.error_message)
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse('add_piechange', kwargs={'pk': self.kwargs['pk']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])

        # Formset des lignes
        if self.request.POST:
            context['lignes_formset'] = LignePiecEchangeFormSet(self.request.POST, instance=self.object)
        else:
            context['lignes_formset'] = LignePiecEchangeFormSet(instance=self.object)

        piechange_queryset = PiecEchange.objects.filter(vehicule=vehicule)
        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None

        # 🔹 Gestion des véhicules par rôle
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()  
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()

        # 🔹 Recherche véhicule
        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()
        # 🔹 Filtres date
        today = date.today()
        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
        piechange_base = PiecEchange.objects.filter(vehicule=vehicule)
        if date_debut and date_fin:
            filtre_piechange = piechange_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_piechange = piechange_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        # 🔹 Statistiques (coûts)
        piechange_jours = filtre_piechange.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        piechange_jours_format = '{:,}'.format(piechange_jours).replace(',', ' ')
        piechange_mois = filtre_piechange.aggregate(somme=Sum('montant'))['somme'] or 0
        piechange_mois_format = '{:,}'.format(piechange_mois).replace(',', ' ')
        piechange_an = piechange_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        piechange_an_format = '{:,}'.format(piechange_an).replace(',', ' ')
        liste_piechange = filtre_piechange.annotate(
            total_quantite=Coalesce(Sum('lignes__quantite'), Value(0))
        ).order_by('-date_saisie')

        # 🔹 Nombre de pièces (somme des quantités LignePiecEchange)
        piechange_mois_nb = LignePiecEchange.objects.filter(piecec__in=filtre_piechange).aggregate(s=Sum('quantite'))['s'] or 0
        piechange_an_nb = LignePiecEchange.objects.filter(
            piecec__in=piechange_base.filter(date_saisie__year=today.year)
        ).aggregate(s=Sum('quantite'))['s'] or 0

        context.update({
            "vehicules": vehicules,
            "vehicule": vehicule,
            "liste_piechange": liste_piechange,

            "piechange_jours_format": piechange_jours_format,
            "piechange_mois_format": piechange_mois_format,
            "piechange_an_format": piechange_an_format,
            "piechange_mois_nb": piechange_mois_nb,
            "piechange_an_nb": piechange_an_nb,

            "dates": dates,
            "mois": libelle_mois,
            "annee": annee,
            "form": form,
            "forms": forms,
        })
        return context

@require_POST
@login_required
def delete_selected_piechange(request):
    selected_ids = request.POST.getlist('piechange_ids')
    if selected_ids:
        piechange_to_delete = PiecEchange.objects.filter(id__in=selected_ids)
        count = piechange_to_delete.count()
        piechange_to_delete.delete()
        messages.success(request, f"{count} pièce changé(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucune pièce changé sélectionnée.")
    return redirect('list_piechange')

# class DetailReparatView(LoginRequiredMixin, DetailView):
class DetailReparatView(LoginRequiredMixin, CustomPermissionRequiredMixin, DetailView):
    login_url = 'login'
    permission_url = 'detail_reparat'
    model = Reparation
    template_name = 'perfect/detail_reparat.html'
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def calculate_effective_hours(self, start, end):
        total_seconds = 0
        current = start
        while current < end:
            next_hour = (current + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
            if next_hour > end:
                next_hour = end
            if time(5, 0) <= current.time() < time(22, 0):
                total_seconds += (next_hour - current).total_seconds()
            current = next_hour
        return total_seconds / 3600
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        mois = date.today().month
        mois_en_cours =date.today().month
        reparation = get_object_or_404(Reparation, pk=self.kwargs['pk'])
        list_piece= reparation.pieces.all()
        vehicule = reparation.vehicule
        user = self.request.user
        if user.user_type == "4":
            try:
                gerant = user.gerants.get()  
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.exists():  
                    vehicules = Vehicule.objects.filter(category__in=categories_gerant)
                else:
                    vehicules = Vehicule.objects.none() 
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()
        else:
            vehicules = Vehicule.objects.all()
            
        date_entree = reparation.date_entree
        date_sortie = reparation.date_sortie
        duree_effective_heures = round(self.calculate_effective_hours(date_entree, date_sortie), 2)
        perte_par_30min = vehicule.category.perte_par_30min or 0
        recette_categorie = vehicule.category.recette_defaut or 0
        perte = round((duree_effective_heures * 2) * perte_par_30min, 2)
        recette_nette = round(recette_categorie - perte, 2)
        total_piece= Piece.objects.filter(reparation = reparation,).aggregate(somme=Sum('montant'))['somme'] or 0
        
        context={
            'reparation':reparation,
            'duree_effective_heures':duree_effective_heures,
            'perte':perte,
            'recette_nette':recette_nette,
            'vehicules':vehicules,
            'total_piece':total_piece,
            'list_piece':list_piece,
            # 'list_reparation':list_reparation,
            'dates':dates
        }
        return context

class ExportReparationPDFView(LoginRequiredMixin, CustomPermissionRequiredMixin, View):
    login_url = 'login'
    permission_url = 'detail_reparat'
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    
    def calculate_effective_hours(self, start, end):
        total_seconds = 0
        current = start
        while current < end:
            next_hour = (current + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
            if next_hour > end:
                next_hour = end
            if time(5, 0) <= current.time() < time(22, 0):
                total_seconds += (next_hour - current).total_seconds()
            current = next_hour
        return total_seconds / 3600
    
    def get(self, request, pk):
        try:
            from xhtml2pdf import pisa
            from django.conf import settings
            import os
        except ImportError:
            messages.error(request, "La bibliothèque xhtml2pdf n'est pas installée. Veuillez installer: pip install xhtml2pdf")
            return redirect('detail_reparat', pk=pk)
        
        reparation = get_object_or_404(Reparation, pk=pk)
        list_piece = reparation.pieces.all()
        vehicule = reparation.vehicule
        
        date_entree = reparation.date_entree
        date_sortie = reparation.date_sortie
        duree_effective_heures = round(self.calculate_effective_hours(date_entree, date_sortie), 2)
        perte_par_30min = vehicule.category.perte_par_30min or 0
        recette_categorie = vehicule.category.recette_defaut or 0
        perte = round((duree_effective_heures * 2) * perte_par_30min, 2)
        recette_nette = round(recette_categorie - perte, 2)
        total_piece = Piece.objects.filter(reparation=reparation).aggregate(somme=Sum('montant'))['somme'] or 0
        
        # Préparer le chemin de l'image si elle existe
        image_path = None
        if reparation.image:
            image_path = os.path.join(settings.MEDIA_ROOT, str(reparation.image))
            if not os.path.exists(image_path):
                image_path = None
        
        # Préparer le chemin du logo
        logo_path = None
        if hasattr(settings, 'STATIC_ROOT') and settings.STATIC_ROOT:
            logo_path = os.path.join(settings.STATIC_ROOT, 'pb_logo.png')
        elif hasattr(settings, 'STATICFILES_DIRS') and settings.STATICFILES_DIRS:
            logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'pb_logo.png')
        
        if logo_path and not os.path.exists(logo_path):
            logo_path = None
        
        context = {
            'reparation': reparation,
            'duree_effective_heures': duree_effective_heures,
            'perte': perte,
            'recette_nette': recette_nette,
            'total_piece': total_piece,
            'list_piece': list_piece,
            'dates': date.today(),
            'image_path': image_path,
            'logo_path': logo_path,
        }
        
        # Rendre le template HTML
        html_content = render(request, 'perfect/detail_reparat_pdf.html', context).content.decode('utf-8')
        
        # Créer le PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f"Fiche_Reparation_{reparation.num_fich}_{reparation.vehicule.immatriculation}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Générer le PDF avec gestion des erreurs
        def link_callback(uri, rel):
            """
            Convertit les URIs HTML en chemins système pour les images
            """
            if uri.startswith('http://') or uri.startswith('https://'):
                return uri
            if uri.startswith(settings.MEDIA_URL):
                path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
            elif uri.startswith(settings.STATIC_URL):
                path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], uri.replace(settings.STATIC_URL, ''))
            else:
                return uri
            if os.path.exists(path):
                return path
            return uri
        # Générer le PDF
        pisa_status = pisa.CreatePDF(
            html_content,
            dest=response,
            link_callback=link_callback,
            encoding='utf-8'
        )
        if pisa_status.err:
            messages.error(request, f"Erreur lors de la génération du PDF: {pisa_status.err}")
            return redirect('detail_reparat', pk=pk)
        
        return response

class ListPiechangeView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_piechange'
    model = PiecEchange
    template_name = 'perfect/liste_piechange.html'
    ordering = ['date_saisie']
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        piechange_base = PiecEchange.objects.all()
        form = self.form_class(self.request.GET)

        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')

            if categorie_filter:
                piechange_base = piechange_base.filter(vehicule__category__category=categorie_filter)

            if immatriculation:
                piechange_base = piechange_base.filter(vehicule__immatriculation__icontains=immatriculation)

            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListRecetView)
            if date_debut and date_fin:
                filtre_piechange = piechange_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_piechange = piechange_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_piechange = piechange_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        # Queryset affiché : annoter avec la quantité totale de pièces par enregistrement
        liste_piechange = filtre_piechange.annotate(
            total_quantite=Coalesce(Sum('lignes__quantite'), Value(0))
        )

        # ------------------ STATISTIQUES (coûts) -------------------
        # Jours / Mois : sur la période affichée
        piechange_jours = filtre_piechange.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        piechange_jours_format = '{:,}'.format(piechange_jours).replace(',', ' ')
        piechange_mois = filtre_piechange.aggregate(somme=Sum('montant'))['somme'] or 0
        piechange_mois_format = '{:,}'.format(piechange_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        piechange_an = piechange_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        piechange_an_format = '{:,}'.format(piechange_an).replace(',', ' ')

        # ------------------ NOMBRE DE PIÈCES (somme des quantités LignePiecEchange) -------------------
        filtre_jours = filtre_piechange.filter(date_saisie=today)
        piechange_jours_nb = LignePiecEchange.objects.filter(piecec__in=filtre_jours).aggregate(s=Sum('quantite'))['s'] or 0
        piechange_mois_nb = LignePiecEchange.objects.filter(piecec__in=filtre_piechange).aggregate(s=Sum('quantite'))['s'] or 0
        piechange_an_nb = LignePiecEchange.objects.filter(
            piecec__in=piechange_base.filter(date_saisie__year=today.year)
        ).aggregate(s=Sum('quantite'))['s'] or 0

        context = {
            'liste_piechange': liste_piechange,
            'piechange_an_format': piechange_an_format,
            'piechange_mois_format': piechange_mois_format,
            'piechange_jours_format': piechange_jours_format,
            'piechange_jours_nb': piechange_jours_nb,
            'piechange_mois_nb': piechange_mois_nb,
            'piechange_an_nb': piechange_an_nb,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

class ListPieceView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_piece'
    model = Piece
    template_name = 'perfect/liste_pieces.html'
    ordering = ['date_saisie']
    timeout_minutes = 500
    form_class = DateFormPiece
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        piece_queryset = Piece.objects.all()
        form = self.form_class(self.request.GET)

        # ------------------ FILTRES -------------------
        piece_base = Piece.objects.all()
        date_debut = date_fin = None
        if form.is_valid():
            lieu_filter = form.cleaned_data.get('lieu')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            categorie = form.cleaned_data.get('categorie')

            if lieu_filter:
                piece_base = piece_base.filter(lieu=lieu_filter)
            if immatriculation:
                piece_base = piece_base.filter(reparation__vehicule__immatriculation__icontains=immatriculation)
            if categorie:
                piece_base = piece_base.filter(reparation__vehicule__category__category=categorie)

        # Filtre de date : si fourni, utiliser la plage ; sinon par défaut = mois en cours
        if date_debut and date_fin:
            filtre_piece = piece_base.filter(date_saisie__range=[date_debut, date_fin])
        else:
            filtre_piece = piece_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        # ------------------ STATISTIQUES -------------------
        piece_jours = filtre_piece.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        piece_jours_format = '{:,}'.format(piece_jours).replace(',', ' ')
        piece_mois = filtre_piece.aggregate(somme=Sum('montant'))['somme'] or 0
        piece_mois_format = '{:,}'.format(piece_mois).replace(',', ' ')
        piece_mois_count = filtre_piece.count()
        piece_an = piece_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        piece_an_format = '{:,}'.format(piece_an).replace(',', ' ')
        piece_an_count = piece_base.filter(date_saisie__year=today.year).count()

        context = {
            'liste_piece': filtre_piece,
            'piece_an_format': piece_an_format,
            'piece_an_count': piece_an_count,
            'piece_mois_format': piece_mois_format,
            'piece_mois_count': piece_mois_count,
            'piece_jours_format': piece_jours_format,
            'form': form,
        }
        return context

class ExportPieceExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        pieces = Piece.objects.select_related('reparation__vehicule', 'auteur')
        lieu_filter = request.GET.get('lieu')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        today = date.today()

        # Filtre par date : plage fournie par l'utilisateur, sinon mois en cours par défaut
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                pieces = pieces.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pieces = pieces.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            pieces = pieces.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        if immatriculation:
            pieces = pieces.filter(reparation__vehicule__immatriculation__icontains=immatriculation)

        if categorie:
            pieces = pieces.filter(reparation__vehicule__category__category__icontains=categorie)

        if lieu_filter:
                pieces = pieces.filter(lieu=lieu_filter)
        # --- Création du fichier Excel ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pièces utilisées"

        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Libellé Pièce",
            "Lieu",
            "Quantité",
            "Montant Unitaire",
            "Prix Total",
            "Auteur",
            "Date Saisie",
        ]
        ws.append(headers)

        # --- Ajout des données ---
        for piece in pieces:
            vehicule = piece.reparation.vehicule if piece.reparation and piece.reparation.vehicule else None
            ws.append([
                vehicule.immatriculation if vehicule else "",
                vehicule.marque if vehicule else "",
                vehicule.category.category if vehicule and vehicule.category else "",
                piece.libelle or "",
                piece.lieu or "",
                piece.quantite,
                piece.montant,
                piece.prix_total,
                piece.auteur.username if piece.auteur else "N/A",
                piece.date_saisie.strftime("%d-%m-%Y") if piece.date_saisie else "",
            ])

        # --- Ajustement largeur colonnes ---
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22

        # --- Réponse HTTP ---
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Pieces-Reparations.xlsx'
        wb.save(response)
        return response

class ExportPiecEchangeExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        pieces = PiecEchange.objects.all().select_related('vehicule', 'auteur')

        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')

        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                pieces = pieces.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                pass
        if immatriculation:
            pieces = pieces.filter(vehicule__immatriculation__icontains=immatriculation)

        if categorie:
            pieces = pieces.filter(vehicule__category__category__icontains=categorie)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pièces Échangées"

        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Montant",
            "Date Saisie",
            "Auteur"
        ]
        ws.append(headers)

        for p in pieces:
            ws.append([
                p.vehicule.immatriculation if p.vehicule else "",
                p.vehicule.marque if p.vehicule else "",
                p.vehicule.category.category if p.vehicule and p.vehicule.category else "",
                p.montant,
                p.date_saisie.strftime("%d-%m-%Y") if p.date_saisie else "",
                p.auteur.username if p.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Pieces-Echangees.xlsx'
        wb.save(response)
        return response

class ListReparationView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_repa'
    model = Reparation
    template_name = 'perfect/liste_reparations.html'
    ordering = ['date_saisie']
    context_object = 'listereparation'
    timeout_minutes = 300
    form_class = DateFormListRepar
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        reparat_base = Reparation.objects.all()
        form = self.form_class(self.request.GET)

        # ------------------ FILTRES (base : catégorie + motif + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            motif_filter = form.cleaned_data.get('motif')
            immatriculation = form.cleaned_data.get('immatriculation')

            if categorie_filter:
                reparat_base = reparat_base.filter(vehicule__category__category=categorie_filter)
            if motif_filter:
                reparat_base = reparat_base.filter(motif=motif_filter)
            if immatriculation:
                reparat_base = reparat_base.filter(vehicule__immatriculation__icontains=immatriculation)
            if date_debut and date_fin:
                filtre_reparat = reparat_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_reparat = reparat_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_reparat = reparat_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )

        # ------------------ STATISTIQUES -------------------
        reparat_total = filtre_reparat.count()
        # Jours / Mois : sur la période affichée (filtre_reparat)
        reparat_jours = filtre_reparat.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_jours_format = '{:,}'.format(reparat_jours).replace(',', ' ')
        reparat_mois = filtre_reparat.aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_mois_format = '{:,}'.format(reparat_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        reparat_an = reparat_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_an_format = '{:,}'.format(reparat_an).replace(',', ' ')

        context = {
            'liste_reparat': filtre_reparat,
            'reparat_total': reparat_total,
            'reparat_an_format': reparat_an_format,
            'reparat_mois_format': reparat_mois_format,
            'reparat_jours_format': reparat_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

class ExportReparationExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        reparations = Reparation.objects.all().select_related('vehicule', 'auteur')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        motif = request.GET.get('motif')
        today = date.today()
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                reparations = reparations.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                reparations = reparations.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            reparations = reparations.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        if immatriculation:
            reparations = reparations.filter(vehicule__immatriculation__icontains=immatriculation)
        if categorie:
            reparations = reparations.filter(vehicule__category__category__icontains=categorie)
        if motif:
            reparations = reparations.filter(motif=motif)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Réparations"
        headers = [
            "Immatriculation",
            "Marque",
            "Catégorie",
            "Motif",
            "Numéro Fiche",
            "Description",
            "Date Entrée",
            "Date Sortie",
            "Durée (jj/hh:mm)",
            "Montant (FCFA)",
            "Prestation (FCFA)",
            "Date Saisie",
            "Auteur"
        ]
        ws.append(headers)
        for r in reparations:
            duree = ""
            if r.date_entree and r.date_sortie:
                # Calculer la différence totale
                delta = r.date_sortie - r.date_entree
                jours = delta.days
                heures = delta.seconds // 3600
                minutes = (delta.seconds % 3600) // 60
                # Format: jour-heure:minute
                duree = f"{jours}/{heures:02d}:{minutes:02d}"
            ws.append([
                r.vehicule.immatriculation if r.vehicule else "",
                r.vehicule.marque if r.vehicule else "",
                r.vehicule.category.category if r.vehicule and r.vehicule.category else "",
                r.motif,
                r.num_fich,
                r.description or "",
                r.date_entree.strftime("%d-%m-%Y %H:%M") if r.date_entree else "",
                r.date_sortie.strftime("%d-%m-%Y %H:%M") if r.date_sortie else "",
                duree,
                r.montant,
                r.prestation,
                r.date_saisie.strftime("%d-%m-%Y") if r.date_saisie else "",
                r.auteur.username if r.auteur else "N/A"
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Reparations.xlsx'
        wb.save(response)
        return response

class BestReparationView(LoginRequiredMixin, CustomPermissionRequiredMixin, TemplateView):
    login_url = 'login'
    permission_url = 'list_repa'
    model = Reparation
    template_name = 'perfect/best_reparation.html'
    context_object = 'listreparation'
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = self.form_class(self.request.GET)
        reparat_base = Reparation.objects.all()
        today = date.today()
        # ------------------ FILTRES (base : catégorie + motif uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            motif_filter = form.cleaned_data.get('motif')

            if categorie_filter:
                reparat_base = reparat_base.filter(
                    vehicule__category__category=categorie_filter
                )
            if motif_filter:
                reparat_base = reparat_base.filter(motif=motif_filter)
            if date_debut and date_fin:
                filtre_reparat = reparat_base.filter(
                    date_saisie__range=[date_debut, date_fin]
                )
            else:
                filtre_reparat = reparat_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_reparat = reparat_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ STATISTIQUES (comme ListReparationView) -------------------
        reparat_total = filtre_reparat.count()
        # Jours / Mois : sur la période affichée (filtre_reparat)
        reparat_jours = filtre_reparat.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_jours_format = '{:,}'.format(reparat_jours).replace(',', ' ')
        reparat_mois = filtre_reparat.aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_mois_format = '{:,}'.format(reparat_mois).replace(',', ' ')
        # Année : sur la base, toute l'année en cours
        reparat_an = reparat_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        reparat_an_format = '{:,}'.format(reparat_an).replace(',', ' ')
        # ------------------ AGREGATION (Panne, Visite, Accident) sur la période affichée -------------------
        best_reparations = (
            filtre_reparat.values(
                'vehicule__id',
                'vehicule__immatriculation',
                'vehicule__category__category'
            )
            .annotate(
                total_reparations=Count('id'),
                total_visite=Count('id', filter=Q(motif="Visite")),
                total_panne=Count('id', filter=Q(motif="Panne")),
                total_accident=Count('id', filter=Q(motif="Accident")),
            )
            .order_by('-total_reparations')  # classé du plus au moins
        )
        context.update({
            'form': form,
            'best_reparations': best_reparations,
            'reparat_total': reparat_total,
            'reparat_jours_format': reparat_jours_format,
            'reparat_mois_format': reparat_mois_format,
            'reparat_an_format': reparat_an_format,
        })
        return context

class ExportBestReparationExcelView(LoginRequiredMixin, View):
    """Vue pour exporter le top des réparations au format Excel"""
    login_url = 'login'
    
    def get(self, request, *args, **kwargs):
        form = DateFormMJR(request.GET)
        reparation_queryset = Reparation.objects.all()
        
        # ------------------ FILTRES (même logique que BestReparationView) -------------------
        today = date.today()
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            motif_filter = form.cleaned_data.get('motif')

            if categorie_filter:
                reparation_queryset = reparation_queryset.filter(
                    vehicule__category__category=categorie_filter
                )
            if motif_filter:
                reparation_queryset = reparation_queryset.filter(
                    motif=motif_filter
                )
            has_date_filter = bool(date_debut and date_fin)
            if has_date_filter:
                reparation_queryset = reparation_queryset.filter(
                    date_saisie__range=[date_debut, date_fin]
                )
            else:
                reparation_queryset = reparation_queryset.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            reparation_queryset = reparation_queryset.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ AGREGATION (même logique que BestReparationView) -------------------
        best_reparations = (
            reparation_queryset.values(
                'vehicule__id',
                'vehicule__immatriculation',
                'vehicule__category__category'
            )
            .annotate(
                total_reparations=Count('id'),
                total_visite=Count('id', filter=Q(motif="Visite")),
                total_panne=Count('id', filter=Q(motif="Panne")),
                total_accident=Count('id', filter=Q(motif="Accident")),
            )
            .order_by('-total_reparations')  # classé du plus au moins
        )
        
        # ------------------ CREATION DU FICHIER EXCEL -------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top Réparations"
        
        # En-têtes
        headers = [
            "IMMAT",
            "CATEGORIE",
            "VISITE",
            "PANNE",
            "ACCIDENT",
            "TOTAL"
        ]
        ws.append(headers)
        
        # Style pour les en-têtes
        header_fill = PatternFill(start_color="06497C", end_color="06497C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[col_letter].width = 18
        
        # Données
        for rep in best_reparations:
            ws.append([
                rep.get('vehicule__immatriculation', ''),
                rep.get('vehicule__category__category', ''),
                rep.get('total_visite', 0),
                rep.get('total_panne', 0),
                rep.get('total_accident', 0),
                rep.get('total_reparations', 0),
            ])
        
        # Alignement des cellules de données
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Réponse HTTP
        filename = f"Top-Reparations-{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

class AddEntretienView(LoginRequiredMixin, CustomPermissionRequiredMixin, CreateView):
    login_url = 'login'
    permission_url = 'add_entretien'
    model = Entretien
    form_class = EntretienForm
    template_name= "perfect/add_entre.html"
    success_message = 'Entretien Ajouté avec succès ✓✓'
    error_message = "Erreur de saisie ✘✘ "
    # success_url = reverse_lazy('journal_compta')
    timeout_minutes = 500
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        form.instance.vehicule_id = self.kwargs['pk']
        messages.success(self.request, self.success_message)
        return super().form_valid(form)  
    def form_invalid(self, form):
        reponse =  super().form_invalid(form)
        messages.success(self.request, self.error_message)
        return reponse    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dates = date.today()
        annee = date.today().year
        now = datetime.now()
        mois_en_cours = date.today().month
        libelle_mois = calendar.month_name[mois_en_cours]
        vehicule = get_object_or_404(Vehicule, pk=self.kwargs['pk'])
        entret_queryset = Entretien.objects.filter(vehicule=vehicule)
        form = DateForm(self.request.GET)
        forms = self.get_form()
        user = self.request.user
        date_debut = date_fin = None

        if user.user_type == "4":
            try:
                gerant = user.gerants.get()
                categories_gerant = gerant.gerant_voiture.all()
                if categories_gerant.filter(category="VTC").exists():
                    vehicules = Vehicule.objects.filter(category__category="VTC")
                else:
                    vehicules = Vehicule.objects.filter(category__category="TAXI")
            except Gerant.DoesNotExist:
                vehicules = Vehicule.objects.none()  
                # No vehicles if no Gerant linked
        elif user:
            try:
                vehicules = Vehicule.objects.all()
            except:
                vehicules = Vehicule.objects.none() 
        else:
            print()

        search_query = self.request.GET.get("search", "").strip()
        if search_query:  
            vehicules = search_vehicules(vehicules, search_query)
        else:
            vehicules = Vehicule.objects.none()

        if form.is_valid():
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
        today = date.today()
        if date_debut and date_fin:
            entret_queryset = entret_queryset.filter(date_saisie__range=[date_debut, date_fin])
        else:
            entret_queryset = entret_queryset.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        entret_jours = entret_queryset.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        entret_jours_format = '{:,}'.format(entret_jours).replace(',', ' ')
        entret_mois = entret_queryset.aggregate(somme=Sum('montant'))['somme'] or 0
        entret_mois_format = '{:,}'.format(entret_mois).replace(',', ' ')
        entret_an = entret_queryset.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        entret_an_format = '{:,}'.format(entret_an).replace(',', ' ')
        liste_entret = entret_queryset.order_by('-date_saisie')
        context = {
            "vehicules": vehicules,
            "vehicule": vehicule,
            'entret_jours_format': entret_jours_format,
            'entret_mois_format': entret_mois_format,
            'entret_an_format': entret_an_format,
            'liste_entret': liste_entret,
            'dates': dates,
            'mois': libelle_mois,
            'annee': annee,
            'form': form,
            'forms': forms,
        }   
        return context  
    def get_success_url(self):
        return reverse('add_entretien', kwargs={'pk': self.kwargs['pk']})

class ListEntretienView(LoginRequiredMixin, CustomPermissionRequiredMixin, ListView):
    login_url = 'login'
    permission_url = 'list_entretien'
    model = Entretien
    template_name = 'perfect/liste_entretien.html'
    ordering = ['date_saisie']
    timeout_minutes = 500
    form_class = DateFormMJR
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        entret_base = Entretien.objects.all()
        form = self.form_class(self.request.GET)
        # ------------------ FILTRES (base : catégorie + immat uniquement) -------------------
        if form.is_valid():
            categorie_filter = form.cleaned_data.get('categorie')
            date_debut = form.cleaned_data.get('date_debut')
            date_fin = form.cleaned_data.get('date_fin')
            immatriculation = form.cleaned_data.get('immatriculation')
            if categorie_filter:
                entret_base = entret_base.filter(vehicule__category__category=categorie_filter)
            if immatriculation:
                entret_base = entret_base.filter(vehicule__immatriculation__icontains=immatriculation)
            # Filtre de date pour l'affichage : plage si fournie, sinon mois en cours (comme ListRecetView)
            if date_debut and date_fin:
                filtre_entret = entret_base.filter(date_saisie__range=[date_debut, date_fin])
            else:
                filtre_entret = entret_base.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            filtre_entret = entret_base.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        # ------------------ STATISTIQUES -------------------
        entret_jours = filtre_entret.filter(date_saisie=today).aggregate(somme=Sum('montant'))['somme'] or 0
        entret_jours_format = '{:,}'.format(entret_jours).replace(',', ' ')
        entret_mois = filtre_entret.aggregate(somme=Sum('montant'))['somme'] or 0
        entret_mois_format = '{:,}'.format(entret_mois).replace(',', ' ')
        # Année : sur la base (toute l'année en cours), comme ListRecetView
        entret_an = entret_base.filter(date_saisie__year=today.year).aggregate(somme=Sum('montant'))['somme'] or 0
        entret_an_format = '{:,}'.format(entret_an).replace(',', ' ')
        context = {
            'liste_entret': filtre_entret,
            'entret_an_format': entret_an_format,
            'entret_mois_format': entret_mois_format,
            'entret_jours_format': entret_jours_format,
            'dates': today,
            'annees': today.year,
            'form': form,
        }
        return context

@require_POST
@login_required
def delete_selected_entretien(request):
    selected_ids = request.POST.getlist('entretien_ids')
    if selected_ids:
        entretiens_to_delete = Entretien.objects.filter(id__in=selected_ids)
        count = entretiens_to_delete.count()
        entretiens_to_delete.delete()
        messages.success(request, f"{count} entretien(s) supprimée(s) avec succès.")
    else:
        messages.warning(request, "Aucun entretien sélectionné.")
    return redirect('list_entretien')

class ExportEntretienExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        entretiens = Entretien.objects.all().select_related('vehicule', 'auteur')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        immatriculation = request.GET.get('immatriculation')
        categorie = request.GET.get('categorie')
        today = date.today()
        if date_debut and date_fin:
            try:
                date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
                date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()
                entretiens = entretiens.filter(date_saisie__range=[date_debut_obj, date_fin_obj])
            except ValueError:
                entretiens = entretiens.filter(
                    date_saisie__month=today.month,
                    date_saisie__year=today.year
                )
        else:
            entretiens = entretiens.filter(
                date_saisie__month=today.month,
                date_saisie__year=today.year
            )
        if immatriculation:
            entretiens = entretiens.filter(vehicule__immatriculation__icontains=immatriculation)
        if categorie:
            entretiens = entretiens.filter(vehicule__category__category__icontains=categorie)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entretiens"
        headers = [
            "Véhicule",
            "Date Entrée",
            "Date Sortie",
            "Date Prochain Entretien",
            "Jours Restants Avant Prochain Entretien",
            "Montant",
            "Auteur",
            "Date de Saisie",
        ]
        ws.append(headers)
        for e in entretiens:
            ws.append([
                e.vehicule.immatriculation if e.vehicule else "",
                e.date_Entret.strftime("%d/%m/%Y %H:%M") if e.date_Entret else "",
                e.date_sortie.strftime("%d/%m/%Y %H:%M") if e.date_sortie else "",
                e.date_proch.strftime("%d/%m/%Y") if e.date_proch else "",
                e.jours_ent_restant,
                e.montant,
                e.auteur.username if e.auteur else "",
                e.date_saisie.strftime("%d/%m/%Y") if e.date_saisie else "",
            ])
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 22
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=Entretiens.xlsx'
        wb.save(response)
        return response

class UpdatPatenteView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_patente'
    model = Patente
    form_class = UpdatPatenteForm
    template_name = "perfect/partials/patente_form.html"
    success_url = reverse_lazy('liste_patente')
    success_message = 'Patente modifiée avec succès ✓✓'
    timeout_minutes = 600
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

class UpdatEntretienView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_entretien'
    model = Entretien
    form_class = UpdatEntretienForm
    template_name = "perfect/partials/entretien_form.html"
    success_url = reverse_lazy('list_entretien')
    success_message = 'Entretien modifiée avec succès ✓✓'
    timeout_minutes = 600
    def get_success_url(self):
        next_url = self.request.GET.get('next') or self.request.POST.get('next')
        if next_url:
            return next_url
        return str(self.success_url)
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            from django.template.loader import render_to_string
            ctx = self.get_context_data(form=form)
            html = render_to_string(self.template_name, ctx, request=self.request)
            return HttpResponse(html, content_type="text/html", status=422)
        return super().form_invalid(form)

class UpdateChargeAdminView(LoginRequiredMixin, CustomPermissionRequiredMixin, UpdateView):
    login_url = 'login'
    permission_url = 'updat_charg_administ'
    model = ChargeAdminis
    form_class = updatChargeAdminisForm
    template_name = "perfect/partials/chargadmin_form.html"
    success_url = reverse_lazy('add_chargadminist')
    success_message = 'Charge Administrative modifiée avec succès✓✓'
    timeout_minutes = 600
    def dispatch(self, request, *args, **kwargs):
        last_activity = request.session.get('last_activity')
        if last_activity:
            last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
            if datetime.now() - last_activity > timedelta(minutes=self.timeout_minutes):
                logout(request)
                messages.warning(request, "Vous avez été déconnecté ")
                return redirect("login")
        return super().dispatch(request, *args, **kwargs)
    def form_valid(self, form):
        form.instance.auteur = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": self.success_message})
        return response
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors})
        return super().form_invalid(form)

@login_required(login_url='login')
def historique(request):
    current_year = now().year
    selected_year = int(request.GET.get('annee', current_year))
    selected_categorie_id = request.GET.get('categorie', '').strip()
    immatriculation_filter = request.GET.get('immatriculation', '').strip()
    try:
        selected_categorie_id = int(selected_categorie_id) if selected_categorie_id else None
    except (ValueError, TypeError):
        selected_categorie_id = None

    models_with_amount = {
        'recette': (Recette, 'vehicule'),
        'chargefixe': (ChargeFixe, 'vehicule'),
        'chargevariable': (ChargeVariable, 'vehicule'),
        'chargeadminis': (ChargeAdminis, None),
        'encaissement': (Encaissement, None),
        'decaissement': (Decaissement, None),
        'vignette': (Vignette, 'vehicule'),
        'reparation': (Reparation, 'vehicule'),
        'piece': (Piece, 'reparation__vehicule'),
        'piecechange': (PiecEchange, 'vehicule'),
        'entretien': (Entretien, 'vehicule'),
        'autrarret': (Autrarret, 'vehicule'),
        'visitetechnique': (VisiteTechnique, 'vehicule'),
        'patente': (Patente, 'vehicule'),
        'stationnement': (Stationnement, 'vehicule'),
        'assurance': (Assurance, 'vehicule'),
        'soldejour': (SoldeJour, None),
    }
    context = {}
    for name, (model, vehicule_path) in models_with_amount.items():
        qs = model.history.filter(history_date__year=selected_year)
        if vehicule_path and (selected_categorie_id or immatriculation_filter):
            if selected_categorie_id:
                qs = qs.filter(**{f'{vehicule_path}__category_id': selected_categorie_id})
            if immatriculation_filter:
                qs = qs.filter(**{f'{vehicule_path}__immatriculation__icontains': immatriculation_filter})
        total = qs.aggregate(Sum('montant'))['montant__sum'] or 0
        context[f"total_{name}"] = '{:,}'.format(total).replace(',', ' ')
    last_10_years = [current_year - i for i in range(10)][::-1]
    categories_list = CategoVehi.objects.all().order_by('category')
    context["selected_year"] = selected_year
    context["last_10_years"] = last_10_years
    context["categories_list"] = categories_list
    context["selected_categorie_id"] = selected_categorie_id
    context["immatriculation_filter"] = immatriculation_filter
    return render(request, "perfect/audit_annuel.html", context)
    