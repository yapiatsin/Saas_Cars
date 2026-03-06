import random
import string
from typing import Any
from urllib import request
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse_lazy
from .utils import send_email_with_html_body
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from userauths.forms import EditUserProfileForm, CustomUserCreationForm, PasswordChangingForm, CreateUserProfileForm
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.conf import settings 
from userauths.models import CustomUser
from datetime import datetime
from django.views.generic import ListView, DetailView,CreateView, DeleteView, UpdateView, TemplateView
from django.views import View, generic
from django.contrib.auth.forms import UserChangeForm, PasswordChangeForm
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.mixins import LoginRequiredMixin
#User = settings.AUTH_USER_MODEL
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from .models import UserProfile
from .models import *
from userauths.forms import *
from .models import CustomUser
from Gest_saas.models import Gerant
# Create your views here.
# from .utils import generate_greeting, generate_goodbye
from django.core.mail import send_mail
# from .utils import send_email_with_html_body
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.contrib.auth.hashers import make_password
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth import get_user_model
# Vue pour vérifier l'OTP envoyé par email
from django.utils import timezone
from .forms import ChangePasswordForm

from userauths.forms import CustomPermissionForm, TypeCustomPermissionForm
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
import pandas as pd
CustomUser = get_user_model()


def list_users(request):
    users = CustomUser.objects.filter(is_superuser=False)
    return render(request, 'liste_compte.html', {'users': users})

@login_required
def edit_user_permissions(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if request.method == 'POST':
        form = UserPermissionForm(request.POST)
        if form.is_valid():
            permissions = form.cleaned_data['permissions']
            user.custom_permissions.set(permissions)
            messages.success(request, 'Permissions mises à jour avec succès.')
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': 'Permissions mises à jour avec succès.'
                })
            return redirect('edit_user_permissions', user.id)  
    else:
        form = UserPermissionForm(initial={
            'permissions': user.custom_permissions.all()
        })
    
    # Si c'est une requête AJAX, retourner seulement le contenu de la modal
    if is_ajax:
        html = render_to_string('modif_user_perm_modal.html', {
            'form': form,
            'user': user
        }, request=request)
        return JsonResponse({'html': html})
    
    return render(request, 'modif_user_perm.html', {
        'form': form,
        'user': user
    })

def generate_random_password(length=8):
    characters = string.ascii_letters + string.digits 
    return ''.join(random.choice(characters) for i in range(length))

@login_required(login_url='/login/')
def add_administrateur(request):
    cxt = {}
    adm = Administ.objects.all()
    if request.method == 'POST':
        userform = CustomUserCreationForm(request.POST)
        adminform = AdministForm(request.POST)
        permission_form = UserPermissionForm(request.POST)
        if userform.is_valid() and adminform.is_valid() and permission_form.is_valid():
            try:
                user = userform.save(commit=False)
                password = generate_random_password()
                user.set_password(password)
                user.user_type = "1"
                user.save()
                adminst = adminform.save(commit=False)
                adminst.user = user
                adminst.save()

                permissions = permission_form.cleaned_data['permissions']
                user.custom_permissions.set(permissions)
                
                subjet = 'Création de Compte Administrateur'
                receivers = [user.email]
                template = 'compte_success.html'
                context = {
                    'username': user.username,
                    'password': password,
                    'date': datetime.today().date,
                    'user_email':user.email
                }
                has_send=send_email_with_html_body(
                    subjet=subjet, 
                    receivers= receivers, 
                    template= template, 
                    context=context
                )
                if has_send:
                    messages.success(request, 'Compte créé avec succès. Un email a été envoyé.')
                else:
                    messages.error(request, 'Centre de santé enregistré avec succès, mais l\'email n\'a pas pu être envoyé.')
                return redirect('addadministrateur')
            except Exception as e:
                messages.error(request, f"Erreur: {str(e)}")
        else:
            for field, errors in userform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
            for field, errors in adminform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        userform = CustomUserCreationForm()
        adminform = AdministForm()
        permission_form = UserPermissionForm()
    return render(request, 'add_admin.html', {
        'user_form': userform,
        'admin_form': adminform,
        'cxt': cxt,
        'admins': adm,
        'permission_form': permission_form,
    })

def delete_admin(request, pk):
    try:
        admin = get_object_or_404(Administ, id=pk)
        user = admin.user  
        admin.delete()
        user.delete()
        messages.success(request, f"le compte administrateur de {admin.user.username} et le profile associé ont été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('addadministrateur')

@login_required(login_url='/login/')
def add_chefexploit(request):
    user=request.user
    # Vérifier que l'utilisateur connecté est un administrateur
    if user.user_type != "1":
        messages.error(request, "Seuls les administrateurs peuvent créer des comptes de chef d'exploitation.")
        return redirect('home')
    
    # S'assurer que l'utilisateur connecté a un profil Administ
    # Si le profil n'existe pas, le créer avec des valeurs par défaut
    admin_profile, created = Administ.objects.get_or_create(
        user=user,
        defaults={
            'nom': user.username,
            'prenom': '',
            'commune': '',
        }
    )
    if created:
        messages.info(request, f"Votre profil administrateur a été créé automatiquement. Veuillez le compléter dans votre profil.")
    
    try:
        chefexp = Chefexploitation.objects.all()
    except Administ.DoesNotExist:
        chefexp = Chefexploitation.objects.none()
    cxt = {}
    employ = CustomUser.objects.all()
    if request.method == 'POST':
        userform = CustomUserCreationForm(request.POST)
        chefexploitform = ChefexploitationForm(request.POST)
        permission_form = UserPermissionForm(request.POST)
        if userform.is_valid() and chefexploitform.is_valid() and permission_form.is_valid():
            try:
                user = userform.save(commit=False)
                password = generate_random_password()
                user.set_password(password)
                user.user_type = "2" 
                user.save()
                
                chefexploitation = chefexploitform.save(commit=False)
                chefexploitation.user = user
                
                # Utiliser le profil Administ de l'utilisateur connecté
                chefexploitation.create_by = admin_profile
                chefexploitation.save()

                ################ Ajouter des permissions #################
                permissions = permission_form.cleaned_data['permissions']
                user.custom_permissions.set(permissions)
                
                subjet = "Création de Compte de chef d'exploitation"
                receivers = [user.email]
                template = 'compte_success.html'
                context = {
                    'username': user.username,
                    'password': password,
                    'date': datetime.today().date,
                    'user_email':user.email
                }
                has_send=send_email_with_html_body(
                    subjet=subjet, 
                    receivers= receivers, 
                    template= template, 
                    context=context
                )
                if has_send:
                    messages.success(request, 'Compte créé avec succès. Un email a été envoyé.')
                else:
                    messages.error(request, 'Centre de santé enregistré avec succès, mais l\'email n\'a pas pu être envoyé.')
                return redirect('addchefexploit')
            except Exception as e:
                messages.error(request, f"Erreur: {str(e)}")
        else:
            for field, errors in userform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
            for field, errors in chefexploitform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        userform = CustomUserCreationForm()
        chefexploitform = AdministForm()
        permission_form = UserPermissionForm()
    return render(request, 'add_chef_exploitation.html', {
        'user_form': userform,
        'chefexp_form': chefexploitform,
        'permission_form': permission_form,
        'cxt': cxt,
        'employes': employ,
        'list_chefexp': chefexp,
    })

def delete_chefexploit(request, pk):
    try:
        chefexploit = get_object_or_404(Chefexploitation, id=pk)
        user = chefexploit.user  
        chefexploit.delete()
        user.delete()
        messages.success(request, f"le compte Chef exploitation de {chefexploit.user.username} et le profile associé ont été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('addchefexploit')

@login_required(login_url='/login/')
def add_comptable(request):
    user=request.user
    # Vérifier que l'utilisateur connecté est un administrateur
    if user.user_type != "1":
        messages.error(request, "Seuls les administrateurs peuvent créer des comptes comptable.")
        return redirect('home')
    admin_profile, created = Administ.objects.get_or_create(
        user=user,
        defaults={
            'nom': user.username,
            'prenom': '',
            'commune': '',
        }
    )
    if created:
        messages.info(request, f"Votre profil administrateur a été créé automatiquement. Veuillez le compléter dans votre profil.")
    
    try:
        # admins = Administ.objects.get(user=user)
        compt = Comptable.objects.all()
    except Administ.DoesNotExist:
        compt = Comptable.objects.none()
    cxt = {}
    employ = CustomUser.objects.all()
    if request.method == 'POST':
        userform = CustomUserCreationForm(request.POST)
        comptableform = ComptableForm(request.POST)
        permission_form = UserPermissionForm(request.POST)
        if userform.is_valid() and comptableform.is_valid()and permission_form.is_valid():
            try:
                user = userform.save(commit=False)
                password = generate_random_password()
                user.set_password(password)
                user.user_type = "3"  
                user.save()
                
                comptable = comptableform.save(commit=False)
                comptable.user = user
                
                # Utiliser le profil Administ de l'utilisateur connecté
                comptable.create_by = admin_profile
                comptable.save()
                permissions = permission_form.cleaned_data['permissions']
                user.custom_permissions.set(permissions)
                
                subjet = 'Création de Compte Comptable'
                receivers = [user.email]
                template = 'compte_success.html'
                context = {
                    'username': user.username,
                    'password': password,
                    'date': datetime.today().date,
                    'user_email':user.email
                }
                has_send=send_email_with_html_body(
                    subjet=subjet, 
                    receivers= receivers, 
                    template= template, 
                    context=context
                )
                if has_send:
                    messages.success(request, 'Compte créé avec succès. Un email a été envoyé.')
                else:
                    messages.error(request, 'Centre de santé enregistré avec succès, mais l\'email n\'a pas pu être envoyé.')
                return redirect('addcomptable')
            except Exception as e:
                messages.error(request, f"Erreur: {str(e)}")
        else:
            for field, errors in userform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
            for field, errors in comptableform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        userform = CustomUserCreationForm()
        comptableform = AdministForm()
        permission_form = UserPermissionForm()
    return render(request, 'add_comptable.html', {
        'user_form': userform,
        'comptable_form': comptableform,
        'cxt': cxt,
        'employes': employ,
        'list_compt': compt,
        'permission_form': permission_form,
    })

def delete_comptable(request, pk):
    try:
        comptable = get_object_or_404(Comptable, id=pk)
        user = comptable.user  
        comptable.delete()
        user.delete()
        messages.success(request, f"le compte comptable de {comptable.user.username} et le profile associé ont été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('addcomptable')

@login_required(login_url='/login/')
def add_gerant(request):
    user=request.user
    # Vérifier que l'utilisateur connecté est un administrateur
    if user.user_type != "1":
        messages.error(request, "Seuls les administrateurs peuvent créer des comptes de gérant.")
        return redirect('home')
    # S'assurer que l'utilisateur connecté a un profil Administ
    # Si le profil n'existe pas, le créer avec des valeurs par défaut
    admin_profile, created = Administ.objects.get_or_create(
        user=user,
        defaults={
            'nom': user.username,
            'prenom': '',
            'commune': '',
        }
    )
    if created:
        messages.info(request, f"Votre profil administrateur a été créé automatiquement. Veuillez le compléter dans votre profil.")
    
    try:
        list_gerant = Gerant.objects.all()
    except Administ.DoesNotExist:
        list_gerant = Gerant.objects.none()
    cxt = {}
    employ = CustomUser.objects.all()
    if request.method == 'POST':
        userform = CustomUserCreationForm(request.POST)
        gerantform = GerantForm(request.POST)
        permission_form = UserPermissionForm(request.POST)
        if userform.is_valid() and gerantform.is_valid() and permission_form.is_valid():
            try:
                user = userform.save(commit=False)
                password = generate_random_password()
                user.set_password(password)
                user.user_type = "4"  
                user.save()
                gerant = gerantform.save(commit=False)
                gerant.user = user
                # Utiliser le profil Administ de l'utilisateur connecté
                gerant.create_by = admin_profile
                gerant.save()
                # Enregistrer les catégories de véhicules (ManyToMany)
                categories_vehicules = gerantform.cleaned_data.get('gerant_voiture', [])
                gerant.gerant_voiture.set(categories_vehicules)
                permissions = permission_form.cleaned_data['permissions']
                user.custom_permissions.set(permissions)
                
                subjet = 'Création de Compte de Gérante'
                receivers = [user.email]
                template = 'compte_success.html'
                context = {
                    'username': user.username,
                    'password': password,
                    'date': datetime.today().date,
                    'user_email':user.email
                }
                has_send=send_email_with_html_body(
                    subjet=subjet, 
                    receivers= receivers, 
                    template= template, 
                    context=context
                )
                if has_send:
                    messages.success(request, 'Compte créé avec succès. Un email a été envoyé.')
                else:
                    messages.error(request, 'Centre de santé enregistré avec succès, mais l\'email n\'a pas pu être envoyé.')
                return redirect('addgerant')
            except Exception as e:
                messages.error(request, f"Erreur: {str(e)}")
        else:
            for field, errors in userform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
            for field, errors in gerantform.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        userform = CustomUserCreationForm()
        gerantform = GerantForm()
        permission_form = UserPermissionForm()
    return render(request, 'add_gerant.html', {
        'user_form': userform,
        'gerantform_form': gerantform,
        'cxt': cxt,
        'employes': employ,
        'list_gerant': list_gerant,
        'permission_form': permission_form,
    })

def delete_gerant(request, pk):
    try:
        gerant = get_object_or_404(Gerant, id=pk)
        user = gerant.user  
        gerant.delete()
        user.delete()
        messages.success(request, f"Le gérant {user.username} été supprimés avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    return redirect('addgerant')

@login_required(login_url='/login/')
def edit_gerant(request, pk):
    """Vue pour charger le formulaire d'édition d'un gérant dans un modal"""
    try:
        gerant = get_object_or_404(Gerant, id=pk)
        user = gerant.user
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if request.method == 'POST':
            form = GerantEditForm(request.POST, instance=gerant)
            if form.is_valid():
                try:
                    # Mettre à jour CustomUser
                    user.username = form.cleaned_data['username']
                    user.email = form.cleaned_data['email']
                    user.gender = form.cleaned_data['gender']
                    user.save()
                    
                    # Mettre à jour Gerant
                    gerant.nom = form.cleaned_data['nom']
                    gerant.prenom = form.cleaned_data['prenom']
                    gerant.commune = form.cleaned_data['commune']
                    gerant.tel1 = form.cleaned_data['tel1']
                    gerant.tel2 = form.cleaned_data['tel2']
                    gerant.save()
                    
                    # Mettre à jour les catégories de véhicules (ManyToMany)
                    categories_vehicules = form.cleaned_data.get('gerant_voiture', [])
                    gerant.gerant_voiture.set(categories_vehicules)
                    
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'message': 'Compte gérant modifié avec succès.'
                        })
                    messages.success(request, 'Compte gérant modifié avec succès.')
                    return redirect('addgerant')
                except Exception as e:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': f"Erreur: {str(e)}"
                        })
                    messages.error(request, f"Erreur: {str(e)}")
        else:
            form = GerantEditForm(instance=gerant)
        
        # Si c'est une requête AJAX, retourner seulement le contenu de la modal
        if is_ajax:
            try:
                html = render_to_string('edit_gerant_modal.html', {
                    'form': form,
                    'gerant': gerant,
                    'user': user
                }, request=request)
                return JsonResponse({'html': html})
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                return JsonResponse({
                    'html': f'<div class="alert alert-danger">Erreur lors du rendu du formulaire: {str(e)}<br><small>{error_detail}</small></div>'
                }, status=500)
        
        return render(request, 'edit_gerant.html', {
            'form': form,
            'gerant': gerant,
            'user': user
        })
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({
                'html': f'<div class="alert alert-danger">Erreur lors du chargement: {str(e)}<br><small>{error_detail}</small></div>'
            }, status=500)
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('addgerant')

@login_required(login_url='/login/')
def edit_gerant_by_user(request, user_id):
    """Vue pour charger le formulaire d'édition d'un gérant à partir de son user_id (pour liste_compte.html)"""
    try:
        user = get_object_or_404(CustomUser, id=user_id)
        if user.user_type != "4":
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Cet utilisateur n\'est pas un gérant.'
                })
            messages.error(request, "Cet utilisateur n'est pas un gérant.")
            return redirect('compte')
        
        try:
            gerant = user.gerants.get()
        except Gerant.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Profil gérant introuvable.'
                })
            messages.error(request, "Profil gérant introuvable.")
            return redirect('compte')
        
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if request.method == 'POST':
            form = GerantEditForm(request.POST, instance=gerant)
            if form.is_valid():
                try:
                    # Mettre à jour CustomUser
                    user.username = form.cleaned_data['username']
                    user.email = form.cleaned_data['email']
                    user.gender = form.cleaned_data['gender']
                    user.save()
                    
                    # Mettre à jour Gerant
                    gerant.nom = form.cleaned_data['nom']
                    gerant.prenom = form.cleaned_data['prenom']
                    gerant.commune = form.cleaned_data['commune']
                    gerant.tel1 = form.cleaned_data['tel1']
                    gerant.tel2 = form.cleaned_data['tel2']
                    gerant.save()
                    
                    # Mettre à jour les catégories de véhicules (ManyToMany)
                    categories_vehicules = form.cleaned_data.get('gerant_voiture', [])
                    gerant.gerant_voiture.set(categories_vehicules)
                    
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'message': 'Compte gérant modifié avec succès.'
                        })
                    messages.success(request, 'Compte gérant modifié avec succès.')
                    return redirect('compte')
                except Exception as e:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': f"Erreur: {str(e)}"
                        })
                    messages.error(request, f"Erreur: {str(e)}")
        else:
            form = GerantEditForm(instance=gerant)
        
        # Si c'est une requête AJAX, retourner seulement le contenu de la modal
        if is_ajax:
            try:
                html = render_to_string('edit_gerant_by_user_modal.html', {
                    'form': form,
                    'gerant': gerant,
                    'user': user
                }, request=request)
                return JsonResponse({'html': html})
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                return JsonResponse({
                    'html': f'<div class="alert alert-danger">Erreur lors du rendu du formulaire: {str(e)}<br><small>{error_detail}</small></div>'
                }, status=500)
        
        return render(request, 'edit_gerant.html', {
            'form': form,
            'gerant': gerant,
            'user': user
        })
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({
                'html': f'<div class="alert alert-danger">Erreur lors du chargement: {str(e)}<br><small>{error_detail}</small></div>'
            }, status=500)
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('compte')

def password_success(request):
    return render(request,'userauths/success.html')

def pb_home(request):
    return render(request,'no_acces.html')
    # return render(request,'perfect/pb_home.html')
    
def loginview(request):
    if request.user.is_authenticated:
        messages.warning(request,f"hey you are already logged In")
        return redirect("home")
    if request.method == "POST": 
        email = request.POST.get("email")  
        password = request.POST.get("password")
        try:
            user = CustomUser.objects.get(email=email)
            user = authenticate(request, email=email, password=password)
            if user is not None:
                login(request, user)
                user_type=user.user_type
                if user_type == '1':
                    messages.success(request, f"Bienvenue Administrateur {user.username}")
                    return redirect('dash')
                elif user_type == '2':
                    messages.success(request, f"Bienvenue Chef d'exploitation {user.username}")
                    return redirect('dash')
                elif user_type == '3':
                    messages.success(request, f"Bienvenue Comptable {user.username}")
                    return redirect('dash')
                elif user_type == '4':
                    messages.success(request,  f"Bienvenue Gérant {user.username}")
                    return redirect('dashgarage')
                else:
                   return redirect('login')
            else:
                messages.error(request,  f"Mot de passe ou email invalide")
        except:
            messages.error(request, "Détails de connexion invalides!!!")
    return render(request, "perfect/logins.html")

def logout_view(request):
    user=request.user
    logout(request)
    messages.success(request, f"Vous êtes deconnecté {user.username}")
    return redirect("home")

@require_POST
@login_required(login_url="login")
def toggle_active_user(request, pk):
    user = get_object_or_404(CustomUser, id=pk)
    user.is_active = not user.is_active
    user.save()
    return JsonResponse({"success": True, "is_active": user.is_active})    

def interneView(request):
    return render(request,"userauths/interne.html")
# Vue pour afficher le formulaire de saisie de l'email
class ForgotPasswordView(View):
    def get(self, request):
        # return render(request, 'saisie_otp.html')
        return render(request, 'perfect/forgot_password.html')    

class RequestEmailView(View):
    def post(self, request):
        email = request.POST.get('email')
        try:
            user = CustomUser.objects.get(email=email)
            otp = ''.join(random.choices('0123456789', k=4))
            PWD_FORGET.objects.create(user_id=user, otp=otp, status='0')

            subject = 'Votre OTP de réinitialisation de mot de passe'
            from_email = settings.EMAIL_HOST_USER
            to = [email]

            # HTML message
            html_content = f'''
           <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Réinitialisation du mot de passe</title>
            </head>
            <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <div style="max-width: 100%; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="text-align: center; color: #333;">REINITIALISATION DE MOT DE PASSE</h2>
                    <p>Bonjour cher {user.username},</p>
                    <p>Vous avez demandé à réinitialiser votre mot de passe. Veuillez utiliser le code OTP ci-dessous pour compléter cette action :</p>
                    
                    <div style="text-align: center; margin: 20px 0;">
                        <p style="font-size: 20px; font-weight: bold; color: #2c3e50;">Votre code OTP : <span style="color: #e74c3c;">{otp}</span></p>
                    </div>
                    
                    <p style="color: #555;">Ce code est valable pour une durée limitée. Si vous n'êtes pas à l'origine de cette demande, vous pouvez ignorer cet e-mail. Pour toute assistance, veuillez nous contacter immédiatement.</p>
            
                    <p>Merci de votre confiance.</p>
            
                    <p>Cordialement,<br>L'équipe support</p>
            
                    <footer style="margin-top: 20px; text-align: center; font-size: 12px; color: #999;">
                        &copy; 2024-2025 P&BEntreprise - Thrive One. Tous droits réservés.
                    </footer>
                </div>
            </body>
            </html>
            '''
            # Create email message with HTML content
            email_message = EmailMultiAlternatives(subject, 'Votre OTP est {otp}', from_email, to)
            email_message.attach_alternative(html_content, 'text/html')
            email_message.send()
            return redirect("otp")
        except CustomUser.DoesNotExist:
            messages.error(request, "Utilisateur non trouvé.")
            return  render(request, "perfect/forgot_password.html")

# Vue pour vérifier l'OTP envoyé par email
CustomUser = get_user_model()
from django.utils import timezone
class VerifyOtpView(View):
    def get(self, request):
        return render(request, 'perfect/reinitialise.html')
    def post(self, request):
        otp = request.session.get('otp')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not otp or not new_password or not confirm_password:
            return JsonResponse({'error': 'Tous les champs sont obligatoires.'}, status=400)
        
        if new_password != confirm_password:
            messages.error(request, "Les mots de passe ne correspondent pas.")
            return  render(request, "perfect/reinitialise.html")
        try:
            reset_request = PWD_FORGET.objects.get(otp=otp, status='0')
            # Vérifiez si l'OTP à expirer
            if (timezone.now() - reset_request.creat_at).total_seconds() > 300:  # 2 minutes
                messages.error(request, "OTP expiré")
                return redirect('otp')
            # Marquer l'OTP comme utilisé
            reset_request.status = '1'
            reset_request.save()

            # Réinitialiser le mot de passe
            user = reset_request.user_id
            user.password = make_password(new_password)
            user.save()
            messages.success(request, 'Mot de passe réinitialisé avec succès.')
            return redirect('login')
        
        except PWD_FORGET.DoesNotExist:
            messages.error(request, 'OTP non valide.')
            return  render(request, "perfect/otp.html")

class OptValid(View):
    def get(self, request):
        return render(request, 'perfect/otp.html')
    def post(self, request):
        otp = request.POST.get('otp')
        try :
            reset_request = PWD_FORGET.objects.get(otp=otp, status='0')
            request.session['otp'] = otp
            if reset_request :
                return redirect("verify_otp")
        except PWD_FORGET.DoesNotExist:
                 messages.error(request, "OTP non valide.")
                 return  render(request, "perfect/otp.html")
                     
class PasswordChangeView(PasswordChangeView):
    form_class = ChangePasswordForm
    template_name = 'profil.html'
    success_message = "Mot de passe réinitialisé avec succès👍✓✓"
    error_message = "Erreur de saisie ✘✘"
    success_url = reverse_lazy('change_password')
    def form_valid(self, form):
        reponse = super().form_valid(form)
        messages.success(self.request, self.success_message)
        return reponse
    def form_invalid(self, form):
        reponse = super().form_invalid(form)
        messages.error(self.request, self.error_message)
        return reponse 
    def get(self, request, *args, **kwargs):
        form = self.get_form()
        user = get_object_or_404(CustomUser, id=request.user.id)
        admin_profil = None
        chefexploit_profil = None
        comptable_profil = None
        gerant_profil = None
        
        if user.user_type == "1":
            try:
                admin_profil = Administ.objects.get(user=user)
            except Administ.DoesNotExist:
                admin_profil = None
        elif user.user_type == "2":
            try:
                chefexploit_profil = Chefexploitation.objects.get(user=user)
            except Chefexploitation.DoesNotExist:
                chefexploit_profil = None
                
        elif user.user_type == "3":
            try:
                comptable_profil = Comptable.objects.get(user=user)
            except Comptable.DoesNotExist:
                comptable_profil = None
        
        elif user.user_type == "4":
            try:
                gerant_profil = Gerant.objects.get(user=user)
            except Gerant.DoesNotExist:
                gerant_profil = None
        else: 
            print()

        # Récupérer les permissions personnalisées de l'utilisateur
        from userauths.models import TypeCustomPermission
        grouped_permissions = {}
        for category in TypeCustomPermission.objects.all():
            perms = category.cat_permis.filter(users=user)
            if perms.exists():
                grouped_permissions[category] = perms
        
        # Récupérer toutes les permissions personnalisées (sans groupement)
        custom_permissions = user.custom_permissions.all()
        
        # Récupérer les permissions système (Django permissions)
        system_permissions = user.user_permissions.all()

        # Passer les informations récupérées au contexte
        context = {
            'form': form,
            'user': user,
            'admin_profil': admin_profil,
            'chefexploit_profil': chefexploit_profil,
            'comptable_profil': comptable_profil,
            'gerant_profil': gerant_profil,
            'grouped_permissions': grouped_permissions,
            'custom_permissions': custom_permissions,
            'system_permissions': system_permissions,
        }
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        form = self.get_form()
        user = get_object_or_404(CustomUser, id=request.user.id)
        admin_profil = None
        chefexploit_profil = None
        comptable_profil = None
        gerant_profil = None
        
        if user.user_type == "1":
            try:
                admin_profil = Administ.objects.get(user=user)
            except Administ.DoesNotExist:
                admin_profil = None
        elif user.user_type == "2":
            try:
                chefexploit_profil = Chefexploitation.objects.get(user=user)
            except Chefexploitation.DoesNotExist:
                chefexploit_profil = None
        elif user.user_type == "3":
            try:
                comptable_profil = Comptable.objects.get(user=user)
            except Comptable.DoesNotExist:
                comptable_profil = None
        elif user.user_type == "4":
            try:
                gerant_profil = Gerant.objects.get(user=user)
            except Gerant.DoesNotExist:
                gerant_profil = None
        # Récupérer les permissions personnalisées de l'utilisateur
        from userauths.models import TypeCustomPermission
        grouped_permissions = {}
        for category in TypeCustomPermission.objects.all():
            perms = category.cat_permis.filter(users=user)
            if perms.exists():
                grouped_permissions[category] = perms
        # Récupérer toutes les permissions personnalisées (sans groupement)
        custom_permissions = user.custom_permissions.all()
        # Récupérer les permissions système (Django permissions)
        system_permissions = user.user_permissions.all()

        if form.is_valid():
            return self.form_valid(form)
        else:
            # Ajouter le contexte pour afficher les erreurs
            context = {
                'form': form,
                'user': user,
                'admin_profil': admin_profil,
                'chefexploit_profil': chefexploit_profil,
                'comptable_profil': comptable_profil,
                'gerant_profil': gerant_profil,
                'grouped_permissions': grouped_permissions,
                'custom_permissions': custom_permissions,
                'system_permissions': system_permissions,
            }
            return self.render_to_response(context)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = get_object_or_404(CustomUser, id=self.request.user.id)
        admin_profil = None
        chefexploit_profil = None
        comptable_profil = None
        gerant_profil = None
        
        if user.user_type == "1":
            try:
                admin_profil = Administ.objects.get(user=user)
            except Administ.DoesNotExist:
                admin_profil = None
        elif user.user_type == "2":
            try:
                chefexploit_profil = Chefexploitation.objects.get(user=user)
            except Chefexploitation.DoesNotExist:
                chefexploit_profil = None
        elif user.user_type == "3":
            try:
                comptable_profil = Comptable.objects.get(user=user)
            except Comptable.DoesNotExist:
                comptable_profil = None
        elif user.user_type == "4":
            try:
                gerant_profil = Gerant.objects.get(user=user)
            except Gerant.DoesNotExist:
                gerant_profil = None

        # Récupérer les permissions personnalisées de l'utilisateur
        from userauths.models import TypeCustomPermission
        grouped_permissions = {}
        for category in TypeCustomPermission.objects.all():
            perms = category.cat_permis.filter(users=user)
            if perms.exists():
                grouped_permissions[category] = perms
        
        # Récupérer toutes les permissions personnalisées (sans groupement)
        custom_permissions = user.custom_permissions.all()
        
        # Récupérer les permissions système (Django permissions)
        system_permissions = user.user_permissions.all()

        context.update({
            'user': user,
            'admin_profil': admin_profil,
            'chefexploit_profil': chefexploit_profil,
            'comptable_profil': comptable_profil,
            'gerant_profil': gerant_profil,
            'grouped_permissions': grouped_permissions,
            'custom_permissions': custom_permissions,
            'system_permissions': system_permissions,
        })
        return context

class PasswordChangeDoneView(View):
    def get(self, request):
         return render(request, 'password_change_done.html')

# ==================== GESTION DES PERMISSIONS ====================
class PermissionListView(LoginRequiredMixin, ListView):
    model = CustomPermission
    template_name = 'perfect/permission.html'
    context_object_name = 'permissions'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = CustomPermission.objects.select_related('categorie').all()
        search = self.request.GET.get('search', '')
        categorie_filter = self.request.GET.get('categorie', '')
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        if categorie_filter:
            queryset = queryset.filter(categorie__id=categorie_filter)
        
        return queryset.order_by('id','categorie__categorie', 'name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = TypeCustomPermission.objects.all()
        context['form'] = CustomPermissionForm()
        return context

class PermissionCreateView(LoginRequiredMixin, CreateView):
    model = CustomPermission
    form_class = CustomPermissionForm
    template_name = 'perfect/permission.html'
    success_url = reverse_lazy('list_permissions')
    
    def form_valid(self, form):
        messages.success(self.request, 'Permission créée avec succès ✓✓')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Erreur lors de la création de la permission ✘✘')
        return super().form_invalid(form)

class PermissionUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomPermission
    form_class = CustomPermissionForm
    template_name = 'perfect/partials/permission_form.html'
    success_url = reverse_lazy('list_permissions')
    success_message = 'Permission modifiée avec succès ✓✓'
    
    def form_valid(self, form):
        messages.success(self.request, self.success_message)
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Erreur lors de la modification ✘✘')
        return super().form_invalid(form)

@login_required
def delete_permission(request, pk):
    permission = get_object_or_404(CustomPermission, pk=pk)
    permission.delete()
    messages.success(request, 'Permission supprimée avec succès ✓✓')
    return redirect('list_permissions')

class ExportPermissionExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        permissions = CustomPermission.objects.select_related('categorie').all().order_by('categorie__categorie', 'name')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Permissions"
        
        # En-tête
        headers = ['ID', 'Nom', 'Catégorie', 'URL']
        ws.append(headers)
        
        # Style des en-têtes
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for perm in permissions:
            ws.append([
                perm.id,
                perm.name,
                perm.categorie.categorie,
                perm.url
            ])
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="Permissions.xlsx"'
        wb.save(response)
        return response

class ImportPermissionExcelView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Aucun fichier sélectionné ✘✘')
            return redirect('list_permissions')
        
        try:
            file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created_count = 0
            updated_count = 0
            errors = []
            # Ignorer la première ligne (en-têtes)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Ignorer les lignes vides
                    continue
                try:
                    perm_id = row[0]
                    name = row[1]
                    categorie_name = row[2]
                    url = row[3]
                    # Récupérer ou créer la catégorie
                    categorie, _ = TypeCustomPermission.objects.get_or_create(
                        categorie=categorie_name
                    )
                    # Créer ou mettre à jour la permission
                    permission, created = CustomPermission.objects.update_or_create(
                        id=perm_id,
                        defaults={
                            'name': name,
                            'categorie': categorie,
                            'url': url
                        }
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                        
                except Exception as e:
                    errors.append(f"Ligne {row}: {str(e)}")
            
            if created_count > 0 or updated_count > 0:
                messages.success(
                    request, 
                    f'Import réussi: {created_count} créé(s), {updated_count} mis à jour ✓✓'
                )
            if errors:
                messages.warning(request, f'Erreurs: {len(errors)} ligne(s) en erreur')
                
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'import: {str(e)} ✘✘')
        
        return redirect('list_permissions')

# ==================== GESTION DES CATÉGORIES ====================

class CategorieListView(LoginRequiredMixin, ListView):
    model = TypeCustomPermission
    template_name = 'perfect/categorie.html'
    context_object_name = 'categories'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = TypeCustomPermissionForm()
        return context

class CategorieCreateView(LoginRequiredMixin, View):
    """Vue combinée pour créer des catégories via formulaire ou import Excel"""
    login_url = 'login'
    success_url = reverse_lazy('list_categories')
    
    def post(self, request, *args, **kwargs):
        # Vérifier si c'est un import Excel
        if request.FILES.get('excel_file'):
            return self.handle_excel_import(request)
        else:
            # Sinon, traiter comme un formulaire normal
            return self.handle_form_submit(request)
    
    def handle_form_submit(self, request):
        """Gère la soumission du formulaire classique"""
        form = TypeCustomPermissionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie créée avec succès ✓✓')
            return redirect(self.success_url)
        else:
            messages.error(request, 'Erreur lors de la création de la catégorie ✘✘')
            return redirect('list_categories')
    
    def handle_excel_import(self, request):
        """Gère l'import depuis Excel"""
        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
            df = df.fillna('')
            
            # Vérifier les colonnes requises
            # Accepter soit 'categorie' soit une seule colonne sans en-tête
            if 'categorie' not in df.columns:
                # Si pas de colonne nommée 'categorie', prendre la première colonne
                if len(df.columns) == 1:
                    df.columns = ['categorie']
                else:
                    messages.error(request, "Format attendu : une seule colonne nommée 'categorie' ou une colonne unique")
                    return redirect('list_categories')
            
            created_count = 0
            skipped_count = 0
            errors = []
            
            for index, row in df.iterrows():
                categorie_name = str(row.get('categorie', '')).strip()
                if not categorie_name:
                    continue
                
                try:
                    # Vérifier si la catégorie existe déjà
                    categorie, created = TypeCustomPermission.objects.get_or_create(
                        categorie=categorie_name
                    )
                    if created:
                        created_count += 1
                    else:
                        skipped_count += 1
                except Exception as e:
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            if created_count > 0:
                messages.success(request, f"✅ {created_count} catégorie(s) créée(s) avec succès")
            if skipped_count > 0:
                messages.info(request, f"ℹ️ {skipped_count} catégorie(s) déjà existante(s)")
            if errors:
                messages.warning(request, f"⚠️ {len(errors)} erreur(s) rencontrée(s)")
            
            return redirect('list_categories')
        
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {str(e)} ✘✘")
            return redirect('list_categories')

class CategorieUpdateView(LoginRequiredMixin, UpdateView):
    model = TypeCustomPermission
    form_class = TypeCustomPermissionForm
    template_name = 'perfect/partials/categorie_perm_form.html'
    success_url = reverse_lazy('list_categories')
    success_message = 'Catégorie modifiée avec succès ✓✓'

class ExportCategoriesExcelView(LoginRequiredMixin, View):
    """Vue pour exporter les catégories de permissions au format Excel (une seule colonne)"""
    login_url = 'login'
    
    def get(self, request, *args, **kwargs):
        categories = TypeCustomPermission.objects.all().order_by('categorie')
        
        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catégories de Permissions"
        
        # En-tête (une seule colonne)
        headers = ["categorie"]
        ws.append(headers)
        
        # Style pour l'en-tête
        header_fill = PatternFill(start_color="06497C", end_color="06497C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell = ws["A1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions["A"].width = 30
        
        # Données (une seule colonne avec les noms des catégories)
        for cat in categories:
            ws.append([cat.categorie])
        
        # Alignement des cellules de données
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Réponse HTTP
        filename = f"Categories-Permissions-{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


@login_required
def delete_categorie(request, pk):
    categorie = get_object_or_404(TypeCustomPermission, pk=pk)
    # Vérifier si la catégorie est utilisée
    if CustomPermission.objects.filter(categorie=categorie).exists():
        messages.error(request, 'Impossible de supprimer: cette catégorie contient des permissions ✘✘')
    else:
        categorie.delete()
        messages.success(request, 'Catégorie supprimée avec succès ✓✓')
    return redirect('list_categories')